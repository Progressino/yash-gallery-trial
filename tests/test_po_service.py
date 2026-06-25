"""PO engine smoke test with minimal sales + inventory."""

import pandas as pd
import pytest

from backend.services.helpers import get_parent_sku
from backend.services.po_engine import calculate_po_base, calculate_quarterly_history


def _minimal_sales():
    return pd.DataFrame(
        {
            "Sku": ["TEST-SKU-1"] * 30,
            "TxnDate": pd.date_range("2025-11-01", periods=30, freq="D"),
            "Transaction Type": ["Shipment"] * 30,
            "Quantity": [2] * 30,
            "Units_Effective": [2] * 30,
            "Source": ["Myntra"] * 30,
        }
    )


def _minimal_inventory():
    return pd.DataFrame(
        {
            "OMS_SKU": ["TEST-SKU-1"],
            "Total_Inventory": [50],
        }
    )


def test_calculate_quarterly_history_returns_rows():
    sales = _minimal_sales()
    pivot = calculate_quarterly_history(
        sales_df=sales,
        mtr_df=None,
        myntra_df=None,
        sku_mapping=None,
        group_by_parent=False,
        n_quarters=4,
    )
    assert not pivot.empty
    assert "OMS_SKU" in pivot.columns


def test_calculate_quarterly_history_uses_sku_mapping():
    """Quarterly pivot keys must match PO engine OMS_SKU (PL strip + mapping)."""
    days = pd.date_range("2025-06-01", periods=20, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": ["1001PLYKBEIGE-M"] * 20,
            "TxnDate": days,
            "Transaction Type": ["Shipment"] * 20,
            "Quantity": [3] * 20,
        }
    )
    mapping = {"1001YKBEIGE-M": "1001YKBEIGE-M"}
    pivot = calculate_quarterly_history(
        sales_df=sales,
        sku_mapping=mapping,
        group_by_parent=False,
        n_quarters=2,
    )
    assert not pivot.empty
    assert (pivot["OMS_SKU"] == "1001YKBEIGE-M").any()


def test_calculate_quarterly_history_merges_older_platform_rows():
    """Platform frames must fill older quarter columns even when sales_df exists."""
    recent = pd.DataFrame(
        {
            "Sku": ["1001YKBEIGE-M"] * 5,
            "TxnDate": pd.date_range("2026-05-01", periods=5, freq="D"),
            "Transaction Type": ["Shipment"] * 5,
            "Quantity": [2] * 5,
        }
    )
    mtr = pd.DataFrame(
        {
            "Date": pd.to_datetime(["2024-11-15", "2025-02-10"]),
            "SKU": ["1001YKBEIGE-M", "1001YKBEIGE-M"],
            "Transaction_Type": ["Shipment", "Shipment"],
            "Quantity": [40, 25],
        }
    )
    pivot = calculate_quarterly_history(
        sales_df=recent,
        mtr_df=mtr,
        sku_mapping=None,
        group_by_parent=False,
        n_quarters=8,
    )
    assert not pivot.empty
    row = pivot.loc[pivot["OMS_SKU"] == "1001YKBEIGE-M"].iloc[0]
    assert int(row.get("Oct-Dec 2024", 0)) == 40
    assert int(row.get("Jan-Mar 2025", 0)) == 25


def test_calculate_quarterly_history_shipment_type_case_insensitive():
    days = pd.date_range("2025-06-01", periods=10, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": ["CASE-SHIP"] * 10,
            "TxnDate": days,
            "Transaction Type": ["SHIPMENT"] * 10,
            "Quantity": [2] * 10,
        }
    )
    pivot = calculate_quarterly_history(
        sales_df=sales,
        mtr_df=None,
        myntra_df=None,
        sku_mapping=None,
        group_by_parent=False,
        n_quarters=2,
    )
    assert not pivot.empty
    assert (pivot["OMS_SKU"] == "CASE-SHIP").any()


def test_sheet_lead_on_parent_applies_to_variant_inventory_sku():
    """Style-level lead row should propagate to size variants (e.g. STYLE-M)."""
    days = pd.date_range("2025-11-01", periods=25, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": ["PARENTLEAD-M"] * 25,
            "TxnDate": days,
            "Transaction Type": ["Shipment"] * 25,
            "Quantity": [2] * 25,
            "Units_Effective": [2] * 25,
            "Source": ["Amazon"] * 25,
        }
    )
    inv = pd.DataFrame({"OMS_SKU": ["PARENTLEAD-M"], "Total_Inventory": [80]})
    sku_status = pd.DataFrame(
        {
            "OMS_SKU": ["PARENTLEAD"],
            "SKU_Sheet_Status": ["Open"],
            "Lead_Time_From_Sheet": [33.0],
            "SKU_Sheet_Closed": [False],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=7,
        target_days=60,
        demand_basis="Sold",
        safety_pct=0.0,
        sku_status_df=sku_status,
    )
    row = po.iloc[0]
    assert int(row["Lead_Time_Days"]) == 33


def test_sheet_lead_numeric_style_row_applies_to_full_oms_sku():
    """Sheet lists style ``1657`` only; inventory / sales use ``1657YKWHITE-M`` (common in Excel)."""
    days = pd.date_range("2025-11-01", periods=25, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": ["1657YKWHITE-M"] * 25,
            "TxnDate": days,
            "Transaction Type": ["Shipment"] * 25,
            "Quantity": [2] * 25,
            "Units_Effective": [2] * 25,
            "Source": ["Amazon"] * 25,
        }
    )
    inv = pd.DataFrame({"OMS_SKU": ["1657YKWHITE-M"], "Total_Inventory": [80]})
    sku_status = pd.DataFrame(
        {
            "OMS_SKU": ["1657"],
            "SKU_Sheet_Status": ["Open"],
            "Lead_Time_From_Sheet": [30.0],
            "SKU_Sheet_Closed": [False],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=7,
        target_days=60,
        demand_basis="Sold",
        safety_pct=0.0,
        sku_status_df=sku_status,
    )
    row = po[po["OMS_SKU"] == "1657YKWHITE-M"].iloc[0]
    assert int(row["Lead_Time_Days"]) == 30


def test_eff_days_uses_active_demand_span_not_trailing_calendar():
    """Eff_Days counts first→last shipment in the ADS window, not empty days to max_date."""
    rows = []
    for d in pd.date_range("2025-11-01", periods=10, freq="D"):
        rows.append(
            {
                "Sku": "ACTIVEDAYS-SKU",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 3,
                "Units_Effective": 3,
                "Source": "Amazon",
            }
        )
    sales = pd.DataFrame(rows)
    inv = pd.DataFrame({"OMS_SKU": ["ACTIVEDAYS-SKU"], "Total_Inventory": [500]})
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=7,
        target_days=60,
        demand_basis="Sold",
        safety_pct=0.0,
        min_denominator=7,
    )
    row = po.iloc[0]
    assert int(row["Eff_Days"]) == 10
    # 10 days × 3 units = 30 sold; Recent_ADS = 30 / 10
    assert abs(float(row["Recent_ADS"]) - 3.0) < 0.02


def test_eff_days_not_forced_to_min_denominator_for_single_active_day():
    """One active shipment day should keep Eff_Days=1 (not forced to 7/30)."""
    sales = pd.DataFrame(
        [
            {
                "Sku": "ONEDAY-SKU",
                "TxnDate": pd.Timestamp("2025-11-10"),
                "Transaction Type": "Shipment",
                "Quantity": 21,
                "Units_Effective": 21,
                "Source": "Amazon",
            }
        ]
    )
    inv = pd.DataFrame({"OMS_SKU": ["ONEDAY-SKU"], "Total_Inventory": [100]})
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=7,
        target_days=60,
        demand_basis="Sold",
        safety_pct=0.0,
        min_denominator=7,
    )
    row = po.iloc[0]
    assert int(row["Eff_Days"]) == 1
    assert abs(float(row["Recent_ADS"]) - 21.0) < 0.02


def test_calculate_po_base_non_empty():
    sales = _minimal_sales()
    inv = _minimal_inventory()
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=7,
        target_days=60,
        demand_basis="Sold",
        safety_pct=0.0,
    )
    assert not po.empty
    assert "PO_Qty" in po.columns
    assert "ADS" in po.columns
    row = po[po["OMS_SKU"] == "TEST-SKU-1"].iloc[0]
    assert row["Sold_Units"] >= 30


def test_calculate_po_ignores_sales_outside_catalog():
    """Historical sales for SKUs not in inventory must not slow or skew PO math."""
    sales = _minimal_sales()
    days = pd.date_range("2025-11-01", periods=30, freq="D")
    noise = []
    for i in range(200):
        for d in days:
            noise.append(
                {
                    "Sku": f"LEGACY-SKU-{i}",
                    "TxnDate": d,
                    "Transaction Type": "Shipment",
                    "Quantity": 3,
                    "Units_Effective": 3,
                }
            )
    sales = pd.concat([sales, pd.DataFrame(noise)], ignore_index=True)
    inv = _minimal_inventory()
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=7,
        target_days=60,
        demand_basis="Sold",
        safety_pct=0.0,
    )
    row = po[po["OMS_SKU"] == "TEST-SKU-1"].iloc[0]
    assert int(row["Sold_Units"]) >= 30


def _sales_two_sizes_same_parent():
    """Shared parent CUTPARENT; L moves more units than XL."""
    days = pd.date_range("2025-11-01", periods=30, freq="D")
    rows = []
    for d in days:
        rows.append(
            {"Sku": "CUTPARENT-L", "TxnDate": d, "Transaction Type": "Shipment", "Quantity": 5, "Units_Effective": 5}
        )
        rows.append(
            {"Sku": "CUTPARENT-XL", "TxnDate": d, "Transaction Type": "Shipment", "Quantity": 1, "Units_Effective": 1}
        )
    return pd.DataFrame(rows)


def test_cutting_ratio_single_gross_size_no_ads_split():
    """One size has PO need: cut ratio stays on that size only."""
    sales = _sales_two_sizes_same_parent()
    # L is understocked (gross PO > 0); XL is flush — only L has gross requirement.
    inv = pd.DataFrame(
        {
            "OMS_SKU": ["CUTPARENT-L", "CUTPARENT-XL"],
            "Total_Inventory": [0, 99_999],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=7,
        target_days=60,
        demand_basis="Sold",
        safety_pct=0.0,
        enforce_two_size_minimum=False,
    )
    row_l = po[po["OMS_SKU"] == "CUTPARENT-L"].iloc[0]
    row_xl = po[po["OMS_SKU"] == "CUTPARENT-XL"].iloc[0]
    assert row_l["PO_Qty"] > 0 and row_xl["PO_Qty"] == 0
    assert row_l["Gross_PO_Qty"] > 0
    assert row_xl["Gross_PO_Qty"] == 0
    # Only one variant should carry the cutting share (not ADS-split across both).
    assert float(row_l["Cutting_Ratio"]) == 1.0
    assert float(row_xl["Cutting_Ratio"]) == 0.0


def test_cutting_ratio_two_gross_sizes_uses_demand_split():
    """Two sizes with gross need and no net PO: split by ADS (or gross) across the family."""
    sales = _sales_two_sizes_same_parent()
    inv = pd.DataFrame(
        {
            "OMS_SKU": ["CUTPARENT-L", "CUTPARENT-XL"],
            "Total_Inventory": [0, 0],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=7,
        target_days=60,
        demand_basis="Sold",
        safety_pct=0.0,
        enforce_two_size_minimum=False,
    )
    row_l = po[po["OMS_SKU"] == "CUTPARENT-L"].iloc[0]
    row_xl = po[po["OMS_SKU"] == "CUTPARENT-XL"].iloc[0]
    assert row_l["Gross_PO_Qty"] > 0 and row_xl["Gross_PO_Qty"] > 0
    assert row_l["Cutting_Ratio"] > row_xl["Cutting_Ratio"]
    assert abs(float(row_l["Cutting_Ratio"]) + float(row_xl["Cutting_Ratio"]) - 1.0) < 0.02


def test_sibling_cut_suggested_when_overstock_has_pending_cutting():
    """1061YK-style: PO sizes with low cover; siblings >target days with pending cut → adjust note."""
    days = pd.date_range("2026-05-01", periods=30, freq="D")
    sales_rows = []
    for d in days:
        sales_rows.append(
            {"Sku": "1061YKBLUE-3XL", "TxnDate": d, "Transaction Type": "Shipment", "Quantity": 1, "Units_Effective": 1}
        )
        sales_rows.append(
            {"Sku": "1061YKBLUE-L", "TxnDate": d, "Transaction Type": "Shipment", "Quantity": 1, "Units_Effective": 1}
        )
    sales = pd.DataFrame(sales_rows)
    inv = pd.DataFrame(
        {
            "OMS_SKU": ["1061YKBLUE-3XL", "1061YKBLUE-L"],
            "Total_Inventory": [5, 200],
        }
    )
    existing_po = pd.DataFrame(
        {
            "OMS_SKU": ["1061YKBLUE-3XL", "1061YKBLUE-L"],
            "PO_Pipeline_Total": [0, 110],
            "Pending_Cutting": [0, 110],
            "Balance_to_Dispatch": [0, 0],
            "PO_Qty_Ordered": [0, 0],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=180,
        grace_days=0,
        demand_basis="Sold",
        safety_pct=0.0,
        enforce_two_size_minimum=False,
        enforce_lead_time_release_gate=False,
        existing_po_df=existing_po,
    )
    need = po[po["OMS_SKU"] == "1061YKBLUE-3XL"].iloc[0]
    donor = po[po["OMS_SKU"] == "1061YKBLUE-L"].iloc[0]
    assert int(need["PO_Qty"]) > 0
    assert float(donor["Projected_Running_Days"]) > 180
    assert int(donor["Pending_Cutting"]) == 110
    assert "ADJUST FROM OTHER SIZES" in str(need["PO_Cutting_Note"])
    assert "L" in str(need["Cut_From_Siblings"])
    assert "Donor" in str(donor["Cutting_Source"])


def test_sku_status_pick_lead_time_days_column_name():
    from backend.services.sku_status_lead import parse_sku_status_lead_dataframe

    raw = pd.DataFrame(
        {
            "SKU": ["A-1"],
            "Status": ["Open"],
            "Lead_Time_Days": [33],
        }
    )
    out = parse_sku_status_lead_dataframe(raw, None)
    assert int(out.iloc[0]["Lead_Time_From_Sheet"]) == 33


def test_sku_status_optional_status_column_defaults_empty():
    from backend.services.sku_status_lead import parse_sku_status_lead_dataframe

    raw = pd.DataFrame({"SKU": ["Z-9"], "Lead_Time_Days": [44]})
    out = parse_sku_status_lead_dataframe(raw, None)
    assert out.iloc[0]["SKU_Sheet_Status"] == ""
    assert int(out.iloc[0]["Lead_Time_From_Sheet"]) == 44


def test_sku_status_detects_lt_column_alias():
    from backend.services.sku_status_lead import parse_sku_status_lead_dataframe

    raw = pd.DataFrame({"SKU": ["LT-A"], "LT": [51], "Status": ["Open"]})
    out = parse_sku_status_lead_dataframe(raw, None)
    assert int(out.iloc[0]["Lead_Time_From_Sheet"]) == 51


def test_existing_po_individual_skus_merge_pipeline_to_po_rows():
    """Per-size SKUs in the existing PO sheet must show on matching PO Engine rows."""
    days = pd.date_range("2026-05-01", periods=20, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": ["1917YKBLUE-3XL"] * 20,
            "TxnDate": days,
            "Transaction Type": ["Shipment"] * 20,
            "Quantity": [3] * 20,
            "Units_Effective": [3] * 20,
            "Source": ["Amazon"] * 20,
        }
    )
    inv = pd.DataFrame({"OMS_SKU": ["1917YKBLUE-3XL"], "Total_Inventory": [23]})
    existing_po = pd.DataFrame(
        {"OMS_SKU": ["1917YKBLUE-3XL"], "PO_Pipeline_Total": [130], "Pending_Cutting": [0], "Balance_to_Dispatch": [130]}
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=135,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing_po,
    )
    row = po.loc[po["OMS_SKU"] == "1917YKBLUE-3XL"].iloc[0]
    assert int(row["PO_Pipeline_Total"]) == 130


def test_existing_po_keeps_bundled_and_individual_sizes_separate():
    """Bundled inventory rows keep only their sheet qty; individual sizes stay separate."""
    days = pd.date_range("2026-05-01", periods=20, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": ["1917YKBLUE-3XL"] * 10 + ["1917YKBLUE-4XL-5XL"] * 10,
            "TxnDate": list(days),
            "Transaction Type": ["Shipment"] * 20,
            "Quantity": [2] * 20,
            "Units_Effective": [2] * 20,
            "Source": ["Amazon"] * 20,
        }
    )
    inv = pd.DataFrame(
        {
            "OMS_SKU": ["1917YKBLUE-3XL", "1917YKBLUE-4XL-5XL"],
            "Total_Inventory": [23, 22],
        }
    )
    existing_po = pd.DataFrame(
        {
            "OMS_SKU": [
                "1917YKBLUE-3XL",
                "1917YKBLUE-4XL",
                "1917YKBLUE-5XL",
                "1917YKBLUE-4XL-5XL",
            ],
            "PO_Pipeline_Total": [130, 170, 150, 4],
            "Balance_to_Dispatch": [130, 170, 150, 4],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=135,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing_po,
    )
    assert len(po[po["OMS_SKU"] == "1917YKBLUE-4XL-5XL"]) == 1
    row_3xl = po.loc[po["OMS_SKU"] == "1917YKBLUE-3XL"].iloc[0]
    row_bund = po.loc[po["OMS_SKU"] == "1917YKBLUE-4XL-5XL"].iloc[0]
    row_4xl = po.loc[po["OMS_SKU"] == "1917YKBLUE-4XL"].iloc[0]
    row_5xl = po.loc[po["OMS_SKU"] == "1917YKBLUE-5XL"].iloc[0]
    assert int(row_3xl["PO_Pipeline_Total"]) == 130
    assert int(row_bund["PO_Pipeline_Total"]) == 4
    assert int(row_4xl["PO_Pipeline_Total"]) == 170
    assert int(row_5xl["PO_Pipeline_Total"]) == 150


def test_po_output_dedupes_duplicate_oms_sku_rows():
    from backend.services.existing_po import dedupe_po_rows_by_sku

    df = pd.DataFrame(
        {
            "OMS_SKU": ["1917YKBLUE-S-M", "1917YKBLUE-S-M"],
            "PO_Pipeline_Total": [0, 115],
            "Total_Inventory": [0, 19],
            "ADS": [0.0, 0.37],
        }
    )
    out = dedupe_po_rows_by_sku(df)
    assert len(out) == 1
    assert int(out.iloc[0]["PO_Pipeline_Total"]) == 115


def test_po_output_dedupes_coalesces_pipeline_when_inventory_row_wins():
    from backend.services.existing_po import dedupe_po_rows_by_sku

    df = pd.DataFrame(
        {
            "OMS_SKU": ["1488YKWHITE-3XL", "1488YKWHITE-3XL"],
            "PO_Pipeline_Total": [0, 300],
            "Total_Inventory": [84, 0],
            "Sold_Units": [200, 0],
            "ADS": [3.0, 0.0],
            "PO_Qty": [470, 0],
        }
    )
    out = dedupe_po_rows_by_sku(df)
    assert len(out) == 1
    row = out.iloc[0]
    assert int(row["PO_Pipeline_Total"]) == 300
    assert int(row["Total_Inventory"]) == 84
    assert int(row["PO_Qty"]) == 470


def test_existing_po_1488ykwhite_pipeline_on_per_size_not_bundled_ghost():
    from pathlib import Path

    from backend.services.existing_po import parse_existing_po

    po_path = Path("/Users/samraisinghani/Downloads/Po 13-Jun-26.xlsx")
    if not po_path.is_file():
        pytest.skip("operator PO fixture not available locally")
    existing = parse_existing_po(po_path.read_bytes(), po_path.name)
    inv_map = {
        "1488YKWHITE-3XL": 84,
        "1488YKWHITE-4XL": 74,
        "1488YKWHITE-3XL-4XL": 0,
        "1488YKWHITE-5XL-6XL": 0,
    }
    days = pd.date_range("2026-05-01", periods=30, freq="D")
    sales_rows = []
    for sku in inv_map:
        for d in days:
            sales_rows.append(
                {
                    "Sku": sku,
                    "TxnDate": d,
                    "Transaction Type": "Shipment",
                    "Quantity": 3,
                    "Units_Effective": 3,
                    "Source": "Amazon",
                }
            )
    sales = pd.DataFrame(sales_rows)
    inv = pd.DataFrame(
        {"OMS_SKU": list(inv_map.keys()), "Total_Inventory": list(inv_map.values())}
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=180,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing,
        enforce_lead_time_release_gate=True,
    )
    row_3xl = po.loc[po["OMS_SKU"] == "1488YKWHITE-3XL"].iloc[0]
    assert int(row_3xl["PO_Pipeline_Total"]) == 300
    assert float(row_3xl["Projected_Running_Days"]) > 80.0
    bundled = po.loc[po["OMS_SKU"] == "1488YKWHITE-3XL-4XL"]
    if not bundled.empty:
        assert int(bundled.iloc[0]["PO_Pipeline_Total"]) == 24


def test_existing_po_pipeline_survives_sku_mapping_to_bundled_listing():
    from backend.services.existing_po import parse_existing_po
    from pathlib import Path

    po_path = Path("/Users/samraisinghani/Downloads/Po 13-Jun-26.xlsx")
    if not po_path.is_file():
        pytest.skip("operator PO fixture not available locally")
    existing = parse_existing_po(po_path.read_bytes(), po_path.name)
    sku_map = {
        "1488YKWHITE-3XL": "1488YKWHITE-3XL-4XL",
        "1488YKWHITE-4XL": "1488YKWHITE-3XL-4XL",
    }
    days = pd.date_range("2026-05-01", periods=30, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": ["1488YKWHITE-3XL"] * 30 + ["1488YKWHITE-4XL"] * 30,
            "TxnDate": list(days) + list(days),
            "Transaction Type": ["Shipment"] * 60,
            "Quantity": [3] * 60,
            "Units_Effective": [3] * 60,
            "Source": ["Amazon"] * 60,
        }
    )
    inv = pd.DataFrame(
        {
            "OMS_SKU": ["1488YKWHITE-3XL", "1488YKWHITE-4XL", "1488YKWHITE-3XL-4XL"],
            "Total_Inventory": [84, 74, 0],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=180,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing,
        sku_mapping=sku_map,
        enforce_lead_time_release_gate=True,
    )
    bundled = po.loc[po["OMS_SKU"] == "1488YKWHITE-3XL-4XL"].iloc[0]
    assert int(bundled["PO_Pipeline_Total"]) >= 600
    assert int(bundled["PO_Qty"]) == 0


def test_po_4_jun_fixture_pending_cutting_flows_to_calculate():
    """Uploaded Po 4-Jun-26 values must appear on Calculate PO rows (not stale cache)."""
    from pathlib import Path

    fixture = Path(__file__).resolve().parent / "fixtures" / "Po_4-Jun-26.xlsx"
    if not fixture.is_file():
        return
    from backend.services.existing_po import parse_existing_po

    existing_po = parse_existing_po(fixture.read_bytes(), fixture.name)
    days = pd.date_range("2026-05-01", periods=20, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": ["1003YKMUSTARD-3XL"] * 20,
            "TxnDate": days,
            "Transaction Type": ["Shipment"] * 20,
            "Quantity": [2] * 20,
            "Units_Effective": [2] * 20,
            "Source": ["Amazon"] * 20,
        }
    )
    inv = pd.DataFrame({"OMS_SKU": ["1003YKMUSTARD-3XL"], "Total_Inventory": [30]})
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=135,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing_po,
    )
    row = po.loc[po["OMS_SKU"] == "1003YKMUSTARD-3XL"].iloc[0]
    assert int(row["Pending_Cutting"]) == 40
    assert int(row["Balance_to_Dispatch"]) == 3
    assert int(row["PO_Pipeline_Total"]) == 43


def test_calculate_po_dedupes_bundled_sku_after_pipeline_merge():
    """Duplicate inventory rows collapse; bundled listing keeps stock, per-size gets pipeline."""
    days = pd.date_range("2026-05-01", periods=20, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": ["1917YKBLUE-4XL-5XL"] * 20,
            "TxnDate": days,
            "Transaction Type": ["Shipment"] * 20,
            "Quantity": [2] * 20,
            "Units_Effective": [2] * 20,
            "Source": ["Amazon"] * 20,
        }
    )
    inv = pd.DataFrame(
        {
            "OMS_SKU": ["1917YKBLUE-4XL-5XL", "1917YKBLUE-4XL-5XL"],
            "Total_Inventory": [22, 22],
        }
    )
    existing_po = pd.DataFrame(
        {
            "OMS_SKU": ["1917YKBLUE-4XL", "1917YKBLUE-5XL"],
            "PO_Pipeline_Total": [85, 77],
            "Balance_to_Dispatch": [85, 77],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=135,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing_po,
    )
    bundled = po.loc[po["OMS_SKU"] == "1917YKBLUE-4XL-5XL"].iloc[0]
    assert float(bundled["Total_Inventory"]) == 22
    assert int(bundled["PO_Pipeline_Total"]) == 0
    assert int(po.loc[po["OMS_SKU"] == "1917YKBLUE-4XL", "PO_Pipeline_Total"].iloc[0]) == 85
    assert int(po.loc[po["OMS_SKU"] == "1917YKBLUE-5XL", "PO_Pipeline_Total"].iloc[0]) == 77


def test_bundled_and_per_size_pipeline_not_double_counted():
    """When sheet + inventory list both band and per-size rows, count pipeline once."""
    days = pd.date_range("2026-05-01", periods=20, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": ["1488YKWHITE-XL"] * 10 + ["1488YKWHITE-XXL"] * 10,
            "TxnDate": list(days)[:10] + list(days)[:10],
            "Transaction Type": ["Shipment"] * 20,
            "Quantity": [2] * 20,
            "Units_Effective": [2] * 20,
            "Source": ["Amazon"] * 20,
        }
    )
    inv = pd.DataFrame(
        {
            "OMS_SKU": ["1488YKWHITE-XL", "1488YKWHITE-XXL", "1488YKWHITE-XL-XXL"],
            "Total_Inventory": [10, 8, 0],
        }
    )
    existing_po = pd.DataFrame(
        {
            "OMS_SKU": ["1488YKWHITE-XL", "1488YKWHITE-XXL", "1488YKWHITE-XL-XXL"],
            "PO_Pipeline_Total": [44, 177, 221],
            "Balance_to_Dispatch": [44, 177, 221],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=135,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing_po,
    )
    xl = po.loc[po["OMS_SKU"] == "1488YKWHITE-XL"].iloc[0]
    xxl = po.loc[po["OMS_SKU"] == "1488YKWHITE-XXL"].iloc[0]
    band = po.loc[po["OMS_SKU"] == "1488YKWHITE-XL-XXL"].iloc[0]
    assert int(xl["PO_Pipeline_Total"]) == 44
    assert int(xxl["PO_Pipeline_Total"]) == 177
    assert int(band["PO_Pipeline_Total"]) == 0
    assert int(po["PO_Pipeline_Total"].sum()) == 221


def test_bundled_only_sheet_splits_pipeline_to_per_size_inventory():
    """Bundled-only PO sheet (XXL-3XL) fans out to per-size rows when no L/XL lines exist."""
    days = pd.date_range("2026-05-01", periods=20, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": ["1917YKBLUE-3XL"] * 10 + ["1917YKBLUE-XXL"] * 10,
            "TxnDate": list(days),
            "Transaction Type": ["Shipment"] * 20,
            "Quantity": [2] * 20,
            "Units_Effective": [2] * 20,
            "Source": ["Amazon"] * 20,
        }
    )
    inv = pd.DataFrame(
        {"OMS_SKU": ["1917YKBLUE-3XL", "1917YKBLUE-XXL"], "Total_Inventory": [23, 23]}
    )
    existing_po = pd.DataFrame(
        {
            "OMS_SKU": ["1917YKBLUE-XXL-3XL"],
            "PO_Pipeline_Total": [200],
            "Pending_Cutting": [196],
            "Balance_to_Dispatch": [4],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=135,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing_po,
    )
    assert po[po["OMS_SKU"] == "1917YKBLUE-XXL-3XL"].empty
    for sku in ("1917YKBLUE-3XL", "1917YKBLUE-XXL"):
        row = po.loc[po["OMS_SKU"] == sku].iloc[0]
        assert int(row["PO_Pipeline_Total"]) == 100
        assert int(row["Pending_Cutting"]) == 98
        assert int(row["Balance_to_Dispatch"]) == 2


def test_po_1917ykblue_pipeline_matches_sheet_with_sku_mapping():
    """sku_mapping must not collapse per-size Existing PO qty onto bundled listing rows."""
    import json
    from pathlib import Path

    mapping_path = Path(__file__).resolve().parents[1] / "backend" / "data" / "yash_sku_mapping_master.json"
    mapping = json.loads(mapping_path.read_text(encoding="utf-8"))

    existing_po = pd.DataFrame(
        {
            "OMS_SKU": [
                "1917YKBLUE-3XL",
                "1917YKBLUE-4XL",
                "1917YKBLUE-5XL",
                "1917YKBLUE-L",
                "1917YKBLUE-M",
                "1917YKBLUE-S",
                "1917YKBLUE-XL",
                "1917YKBLUE-XXL",
                "1917YKBLUE-4XL-5XL",
                "1917YKBLUE-L-XL",
                "1917YKBLUE-S-M",
                "1917YKBLUE-XXL-3XL",
            ],
            "PO_Pipeline_Total": [130, 170, 150, 120, 80, 150, 100, 100, 4, 4, 1, 4],
            "Balance_to_Dispatch": [130, 170, 150, 120, 80, 150, 100, 100, 4, 4, 1, 4],
        }
    )
    days = pd.date_range("2026-05-01", periods=20, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": ["1917YKBLUE-4XL-5XL"] * 20,
            "TxnDate": days,
            "Transaction Type": ["Shipment"] * 20,
            "Quantity": [2] * 20,
            "Units_Effective": [2] * 20,
            "Source": ["Amazon"] * 20,
        }
    )
    inv = pd.DataFrame(
        {
            "OMS_SKU": [
                "1917YKBLUE-4XL-5XL",
                "1917YKBLUE-L-XL",
                "1917YKBLUE-S-M",
                "1917YKBLUE-XXL-3XL",
            ],
            "Total_Inventory": [18, 56, 15, 20],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=135,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing_po,
        sku_mapping=mapping,
    )
    assert int(po.loc[po["OMS_SKU"] == "1917YKBLUE-4XL-5XL", "PO_Pipeline_Total"].iloc[0]) == 4
    assert int(po.loc[po["OMS_SKU"] == "1917YKBLUE-L-XL", "PO_Pipeline_Total"].iloc[0]) == 4
    assert int(po.loc[po["OMS_SKU"] == "1917YKBLUE-S-M", "PO_Pipeline_Total"].iloc[0]) == 1
    assert int(po.loc[po["OMS_SKU"] == "1917YKBLUE-XXL-3XL", "PO_Pipeline_Total"].iloc[0]) == 4
    assert int(po.loc[po["OMS_SKU"] == "1917YKBLUE-3XL", "PO_Pipeline_Total"].iloc[0]) == 130
    assert int(po.loc[po["OMS_SKU"] == "1917YKBLUE-4XL", "PO_Pipeline_Total"].iloc[0]) == 170
    assert int(po.loc[po["OMS_SKU"] == "1917YKBLUE-5XL", "PO_Pipeline_Total"].iloc[0]) == 150
    assert int(po.loc[po["OMS_SKU"] == "1917YKBLUE-L", "PO_Pipeline_Total"].iloc[0]) == 120
    assert float(po.loc[po["OMS_SKU"] == "1917YKBLUE-4XL", "Total_Inventory"].iloc[0]) == 0.0
    assert float(po.loc[po["OMS_SKU"] == "1917YKBLUE-5XL", "Total_Inventory"].iloc[0]) == 0.0
    assert float(po.loc[po["OMS_SKU"] == "1917YKBLUE-L", "Total_Inventory"].iloc[0]) == 0.0
    assert float(po.loc[po["OMS_SKU"] == "1917YKBLUE-4XL-5XL", "Total_Inventory"].iloc[0]) == 18.0
    yk = po[po["OMS_SKU"].astype(str).str.contains("1917YKBLUE")]
    assert len(yk) >= 12
    if "1917YKBLUE-XXXL" in set(yk["OMS_SKU"].astype(str)):
        assert int(po.loc[po["OMS_SKU"] == "1917YKBLUE-XXXL", "PO_Pipeline_Total"].iloc[0]) == 0


def test_po_4_jun_fixture_1917ykblue_sizes_stay_separate():
    """Regression: Po 4-Jun-26 — bundled listings and per-size rows keep sheet quantities."""
    from pathlib import Path

    fixture = Path(__file__).resolve().parent / "fixtures" / "Po_4-Jun-26.xlsx"
    if not fixture.is_file():
        return
    from backend.services.existing_po import parse_existing_po

    existing_po = parse_existing_po(fixture.read_bytes(), fixture.name)
    days = pd.date_range("2026-05-01", periods=20, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": ["1917YKBLUE-4XL-5XL"] * 20,
            "TxnDate": days,
            "Transaction Type": ["Shipment"] * 20,
            "Quantity": [2] * 20,
            "Units_Effective": [2] * 20,
            "Source": ["Amazon"] * 20,
        }
    )
    inv = pd.DataFrame(
        {
            "OMS_SKU": [
                "1917YKBLUE-4XL-5XL",
                "1917YKBLUE-L-XL",
                "1917YKBLUE-S-M",
                "1917YKBLUE-XXL-3XL",
            ],
            "Total_Inventory": [18, 56, 15, 20],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=135,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing_po,
    )
    assert int(po.loc[po["OMS_SKU"] == "1917YKBLUE-4XL-5XL", "PO_Pipeline_Total"].iloc[0]) == 4
    assert int(po.loc[po["OMS_SKU"] == "1917YKBLUE-L-XL", "PO_Pipeline_Total"].iloc[0]) == 4
    assert int(po.loc[po["OMS_SKU"] == "1917YKBLUE-S-M", "PO_Pipeline_Total"].iloc[0]) == 1
    assert int(po.loc[po["OMS_SKU"] == "1917YKBLUE-XXL-3XL", "PO_Pipeline_Total"].iloc[0]) == 4
    assert int(po.loc[po["OMS_SKU"] == "1917YKBLUE-3XL", "PO_Pipeline_Total"].iloc[0]) == 130
    assert int(po.loc[po["OMS_SKU"] == "1917YKBLUE-4XL", "PO_Pipeline_Total"].iloc[0]) == 170
    assert int(po.loc[po["OMS_SKU"] == "1917YKBLUE-5XL", "PO_Pipeline_Total"].iloc[0]) == 150
    # Bundled listing keeps inventory; per-size pipeline rows do not inherit bundled stock.
    assert float(po.loc[po["OMS_SKU"] == "1917YKBLUE-4XL", "Total_Inventory"].iloc[0]) == 0.0
    assert float(po.loc[po["OMS_SKU"] == "1917YKBLUE-L", "Total_Inventory"].iloc[0]) == 0.0
    assert float(po.loc[po["OMS_SKU"] == "1917YKBLUE-4XL-5XL", "Total_Inventory"].iloc[0]) == 18.0
    assert float(po.loc[po["OMS_SKU"] == "1917YKBLUE-L-XL", "Total_Inventory"].iloc[0]) == 56.0
    assert len(po[po["OMS_SKU"].astype(str).str.contains("1917YKBLUE")]) == 12


def test_bundled_inventory_does_not_fan_out_to_per_size_rows():
    """Per-size PO rows from the sheet must not inherit bundled listing inventory."""
    days = pd.date_range("2026-05-01", periods=20, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": ["1917YKBLUE-4XL-5XL"] * 20,
            "TxnDate": days,
            "Transaction Type": ["Shipment"] * 20,
            "Quantity": [2] * 20,
            "Units_Effective": [2] * 20,
            "Source": ["Amazon"] * 20,
        }
    )
    inv = pd.DataFrame(
        {"OMS_SKU": ["1917YKBLUE-4XL-5XL"], "Total_Inventory": [18]}
    )
    existing_po = pd.DataFrame(
        {
            "OMS_SKU": ["1917YKBLUE-4XL", "1917YKBLUE-5XL", "1917YKBLUE-4XL-5XL"],
            "PO_Pipeline_Total": [170, 150, 4],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=135,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing_po,
    )
    assert float(po.loc[po["OMS_SKU"] == "1917YKBLUE-4XL", "Total_Inventory"].iloc[0]) == 0.0
    assert float(po.loc[po["OMS_SKU"] == "1917YKBLUE-5XL", "Total_Inventory"].iloc[0]) == 0.0
    assert float(po.loc[po["OMS_SKU"] == "1917YKBLUE-4XL-5XL", "Total_Inventory"].iloc[0]) == 18.0
    assert int(po.loc[po["OMS_SKU"] == "1917YKBLUE-4XL", "PO_Pipeline_Total"].iloc[0]) == 170


def test_po_pipeline_ghost_row_inherits_sheet_lead_not_global_default():
    """Pipeline-only SKUs were merged before the status sheet; they kept global lead_time."""
    days = pd.date_range("2025-11-01", periods=30, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": ["GHOSTBASE"] * 30,
            "TxnDate": days,
            "Transaction Type": ["Shipment"] * 30,
            "Quantity": [2] * 30,
            "Units_Effective": [2] * 30,
            "Source": ["Amazon"] * 30,
        }
    )
    inv = pd.DataFrame({"OMS_SKU": ["GHOSTBASE"], "Total_Inventory": [100]})
    existing_po = pd.DataFrame({"OMS_SKU": ["PIPELINE-ONLY-SKU"], "PO_Pipeline_Total": [40]})
    sheet = pd.DataFrame(
        {
            "OMS_SKU": ["PIPELINE-ONLY-SKU"],
            "SKU_Sheet_Status": ["Open"],
            "Lead_Time_From_Sheet": [63.0],
            "SKU_Sheet_Closed": [False],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=1,
        target_days=90,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing_po,
        sku_status_df=sheet,
    )
    ghost_row = po[po["OMS_SKU"] == "PIPELINE-ONLY-SKU"]
    assert len(ghost_row) == 1
    assert int(ghost_row.iloc[0]["Lead_Time_Days"]) == 63


def test_sku_status_parse_detects_closed_and_lead():
    from backend.services.sku_status_lead import parse_sku_status_lead_dataframe

    raw = pd.DataFrame(
        {
            "SKU": ["ABC-L", "XYZ-M"],
            "Status": ["Open", "Closed SKU"],
            "Lead Time": [10, 22],
        }
    )
    out = parse_sku_status_lead_dataframe(raw, None)
    assert len(out) == 2
    r = out.set_index("OMS_SKU").loc["XYZ-M"]
    assert bool(r["SKU_Sheet_Closed"]) is True
    assert int(r["Lead_Time_From_Sheet"]) == 22


def test_closed_sku_sheet_zeros_po_qty():
    """Closed SKUs must not receive fresh PO recommendations when the status sheet marks them closed."""
    sales = _minimal_sales()
    inv = _minimal_inventory()
    sheet = pd.DataFrame(
        {
            "OMS_SKU": ["TEST-SKU-1"],
            "SKU_Sheet_Status": ["Closed SKU"],
            "SKU_Sheet_Closed": [True],
            "Lead_Time_From_Sheet": [45.0],
        }
    )
    po_sheet = calculate_po_base(sales, inv, 30, 7, 60, safety_pct=0.0, sku_status_df=sheet)
    po_none = calculate_po_base(sales, inv, 30, 7, 60, safety_pct=0.0, sku_status_df=None)
    assert int(po_sheet.iloc[0]["PO_Qty"]) == 0
    assert "closed" in str(po_sheet.iloc[0]["PO_Block_Reason"]).lower()
    assert bool(po_sheet.iloc[0]["SKU_Sheet_Closed"]) is True
    # Without a status sheet the same demand math may still suggest a PO.
    assert int(po_none.iloc[0]["PO_Qty"]) >= int(po_sheet.iloc[0]["PO_Qty"])


def test_size_variant_inherits_parent_row_sheet_status_and_close_hint():
    """Size-level inventory rows must inherit status from a style/parent row on the sheet."""
    days = pd.date_range("2025-11-01", periods=30, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": ["STY123-M"] * 30,
            "TxnDate": days,
            "Transaction Type": ["Shipment"] * 30,
            "Quantity": [2] * 30,
            "Units_Effective": [2] * 30,
            "Source": ["Amazon"] * 30,
        }
    )
    inv = pd.DataFrame({"OMS_SKU": ["STY123-M"], "Total_Inventory": [100]})
    sheet = pd.DataFrame(
        {
            "OMS_SKU": ["STY123"],
            "SKU_Sheet_Status": ["Closed Discontinued"],
            "SKU_Sheet_Closed": [True],
            "Lead_Time_From_Sheet": [30.0],
        }
    )
    po = calculate_po_base(
        sales,
        inv,
        30,
        45,
        90,
        demand_basis="Sold",
        safety_pct=0.0,
        sku_status_df=sheet,
        enforce_lead_time_release_gate=False,
    )
    row = po.iloc[0]
    assert "closed" in str(row["SKU_Sheet_Status"]).lower()
    assert bool(row["SKU_Sheet_Closed"]) is True
    assert "closed on sheet" in str(row["Suggest_Close_SKU"]).lower()


def test_digit_token_only_lead_does_not_enable_sheet_po_gate():
    """Digit-token fill can set ``Lead_Time_Days`` but must not satisfy the sheet PO gate."""
    sku = "FOOBAR4002BAZ-XL"
    days = pd.date_range("2025-11-01", periods=30, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": [sku] * 30,
            "TxnDate": days,
            "Transaction Type": ["Shipment"] * 30,
            "Quantity": [5] * 30,
            "Units_Effective": [5] * 30,
            "Source": ["Amazon"] * 30,
        }
    )
    inv = pd.DataFrame({"OMS_SKU": [sku], "Total_Inventory": [0]})
    # Sheet row shares the digit token ``4002`` but not parent/prefix alignment with ``sku``.
    sheet = pd.DataFrame(
        {
            "OMS_SKU": ["PREFIX-4002"],
            "SKU_Sheet_Status": ["Open"],
            "SKU_Sheet_Closed": [False],
            "Lead_Time_From_Sheet": [55.0],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=120,
        demand_basis="Sold",
        safety_pct=0.0,
        sku_status_df=sheet,
    )
    row = po.loc[po["OMS_SKU"] == sku].iloc[0]
    # Digit-only borrow is not an approved sheet lead — PO blocked and lead shown as 0.
    assert int(row["Lead_Time_Days"]) == 0
    assert not bool(row["Lead_Time_From_Status_Sheet"])
    assert int(row["PO_Qty"]) == 0
    assert "no lead time resolved" in str(row["PO_Block_Reason"]).lower()


def test_parse_sku_status_fuzzy_column_detects_status_without_header_name():
    from backend.services.sku_status_lead import parse_sku_status_lead_dataframe

    raw = pd.DataFrame(
        {
            "SKU": ["A-1", "B-1"],
            "LeadTime": [10, 12],
            "ColX": ["Open", "Closed SKU"],
        }
    )
    out = parse_sku_status_lead_dataframe(raw, None)
    assert bool(out.set_index("OMS_SKU").loc["B-1"]["SKU_Sheet_Closed"]) is True


def test_merge_po_optional_sheets_into_warm_cache():
    import backend.main as main_mod
    from backend.session import AppSession

    main_mod.clear_warm_cache()
    sess = AppSession()
    sess.daily_inventory_history_df = pd.DataFrame(
        {"OMS_SKU": ["Z-1"], "Date": [pd.Timestamp("2025-01-01")], "Qty": [3.0]}
    )
    sess.sku_status_lead_df = pd.DataFrame(
        {
            "OMS_SKU": ["Z-1"],
            "SKU_Sheet_Status": ["Open"],
            "SKU_Sheet_Closed": [False],
            "Lead_Time_From_Sheet": [10.0],
        }
    )
    main_mod.merge_po_optional_sheets_into_warm_cache(sess)
    assert not main_mod._warm_cache["daily_inventory_history_df"].empty
    assert not main_mod._warm_cache["sku_status_lead_df"].empty


def test_single_size_minimum_zeros_gross_when_only_one_variant_needs_stock():
    sales = _sales_two_sizes_same_parent()
    inv = pd.DataFrame(
        {
            "OMS_SKU": ["CUTPARENT-L", "CUTPARENT-XL"],
            "Total_Inventory": [0, 99_999],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=7,
        target_days=60,
        demand_basis="Sold",
        safety_pct=0.0,
        enforce_two_size_minimum=True,
    )
    row_l = po[po["OMS_SKU"] == "CUTPARENT-L"].iloc[0]
    assert int(row_l["Gross_PO_Qty"]) == 0
    reason = str(row_l["PO_Block_Reason"]).lower()
    assert "only 1 size" in reason and "alter" in reason


def test_single_size_minimum_zeros_final_po_after_pipeline_or_caps():
    sales = _sales_two_sizes_same_parent()
    inv = pd.DataFrame(
        {
            "OMS_SKU": ["CUTPARENT-L", "CUTPARENT-XL"],
            "Total_Inventory": [0, 0],
        }
    )
    # Both sizes have gross demand, but XL is fully covered in existing pipeline.
    existing = pd.DataFrame(
        {
            "OMS_SKU": ["CUTPARENT-L", "CUTPARENT-XL"],
            "PO_Pipeline_Total": [0, 2_000_000],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=7,
        target_days=60,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing,
        enforce_two_size_minimum=True,
    )
    row_l = po[po["OMS_SKU"] == "CUTPARENT-L"].iloc[0]
    row_xl = po[po["OMS_SKU"] == "CUTPARENT-XL"].iloc[0]
    assert int(row_l["Gross_PO_Qty"]) == 0
    assert int(row_l["PO_Qty"]) == 0
    assert int(row_xl["PO_Qty"]) == 0
    reason = str(row_l["PO_Block_Reason"]).lower()
    assert "only 1 size" in reason and "alter" in reason


def test_two_size_minimum_allows_multiple_sizes_with_trailing_color_tokens():
    sales = pd.DataFrame(
        {
            "Sku": ["PARENT-3XL-RED", "PARENT-4XL-RED"],
            "TxnDate": [pd.Timestamp("2025-01-30"), pd.Timestamp("2025-01-30")],
            "Quantity": [30, 25],
            "Units_Effective": [30, 25],
            "Transaction Type": ["Shipment", "Shipment"],
            "Source": ["Amazon", "Amazon"],
        }
    )
    inv = pd.DataFrame(
        {
            "OMS_SKU": ["PARENT-3XL-RED", "PARENT-4XL-RED"],
            "Total_Inventory": [0, 0],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=30,
        target_days=90,
        demand_basis="Sold",
        safety_pct=0.0,
        enforce_two_size_minimum=True,
    )
    assert int(po[po["OMS_SKU"] == "PARENT-3XL-RED"].iloc[0]["PO_Qty"]) > 0
    assert int(po[po["OMS_SKU"] == "PARENT-4XL-RED"].iloc[0]["PO_Qty"]) > 0


def test_sheet_lead_time_applied_per_row():
    sales = _minimal_sales()
    inv = _minimal_inventory()
    sheet = pd.DataFrame(
        {
            "OMS_SKU": ["TEST-SKU-1"],
            "SKU_Sheet_Status": ["Active"],
            "SKU_Sheet_Closed": [False],
            "Lead_Time_From_Sheet": [45.0],
        }
    )
    po = calculate_po_base(sales, inv, 30, 7, 60, safety_pct=0.0, sku_status_df=sheet)
    assert int(po.iloc[0]["Lead_Time_Days"]) == 45


def test_po_release_uses_target_cover_balance_days_formula():
    sales = _minimal_sales()
    inv = pd.DataFrame({"OMS_SKU": ["TEST-SKU-1"], "Total_Inventory": [20]})
    sheet = pd.DataFrame(
        {
            "OMS_SKU": ["TEST-SKU-1"],
            "SKU_Sheet_Status": ["Open"],
            "SKU_Sheet_Closed": [False],
            "Lead_Time_From_Sheet": [45.0],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=90,
        lead_time=30,
        target_days=90,
        demand_basis="Sold",
        safety_pct=0.0,
        sku_status_df=sheet,
        enforce_lead_time_release_gate=False,
    )
    row = po.iloc[0]
    ads = float(row["ADS"])
    inv = 20.0
    target_cover = 90.0
    expected_raw = max((ads * target_cover) - inv, 0.0)
    import math
    expected_po_target = int(math.ceil(expected_raw / 5.0) * 5.0)
    expected_po = expected_po_target
    expected_proj = round(inv / ads, 1) if ads > 0 else 999.0
    assert int(row["Gross_PO_Qty"]) == expected_po_target
    assert int(row["PO_Qty"]) == expected_po
    assert float(row["Projected_Running_Days"]) == expected_proj


def test_lead_time_release_gate_blocks_when_projected_cover_exceeds_lead():
    """No PO while projected (inv + pipeline) / ADS is above factory lead; disable gate for legacy target-only."""
    sales = _minimal_sales()  # ADS ≈ 2/day
    inv = pd.DataFrame({"OMS_SKU": ["TEST-SKU-1"], "Total_Inventory": [80]})
    sheet = pd.DataFrame(
        {
            "OMS_SKU": ["TEST-SKU-1"],
            "SKU_Sheet_Status": ["Open"],
            "SKU_Sheet_Closed": [False],
            "Lead_Time_From_Sheet": [30.0],
        }
    )
    common = dict(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=30,
        target_days=90,
        demand_basis="Sold",
        safety_pct=0.0,
        sku_status_df=sheet,
    )

    po_on = calculate_po_base(**common, enforce_lead_time_release_gate=True)
    po_off = calculate_po_base(**common, enforce_lead_time_release_gate=False)
    # 80 / 2 = 40d projected > 30d lead → gate on ⇒ no release.
    assert int(po_on.iloc[0]["PO_Qty"]) == 0
    assert "lead time" in str(po_on.iloc[0]["PO_Block_Reason"]).lower()
    assert int(po_off.iloc[0]["PO_Qty"]) > 0


def test_sheet_lead_window_blocks_po_when_projected_cover_gt_lead_days():
    """Operator rule: e.g. 49d projected cover vs 45d sheet lead → no PO yet (strict >)."""
    days = pd.date_range("2025-11-01", periods=30, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": ["1001YKBEIGE-XXL"] * 30,
            "TxnDate": days,
            "Transaction Type": ["Shipment"] * 30,
            "Quantity": [1] * 30,
            "Units_Effective": [1] * 30,
            "Source": ["Amazon"] * 30,
        }
    )
    inv = pd.DataFrame({"OMS_SKU": ["1001YKBEIGE-XXL"], "Total_Inventory": [49]})
    sheet = pd.DataFrame(
        {
            "OMS_SKU": ["1001YKBEIGE-XXL"],
            "SKU_Sheet_Status": ["Open"],
            "SKU_Sheet_Closed": [False],
            "Lead_Time_From_Sheet": [45.0],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=93,
        demand_basis="Sold",
        safety_pct=0.0,
        sku_status_df=sheet,
        enforce_lead_time_release_gate=True,
    )
    row = po.iloc[0]
    assert float(row["Projected_Running_Days"]) == 49.0
    assert int(row["Lead_Time_Days"]) == 45
    assert int(row["PO_Qty"]) == 0
    assert "lead time" in str(row["PO_Block_Reason"]).lower()


def test_below_target_cover_gets_po_when_projected_within_lead_window():
    """Lead gate: release only below factory lead; qty tops up toward post-PO target cover."""
    days = pd.date_range("2025-11-01", periods=30, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": ["1052YKGREEN-5XL"] * 30,
            "TxnDate": days,
            "Transaction Type": ["Shipment"] * 30,
            "Quantity": [1] * 30,
            "Units_Effective": [1] * 30,
            "Source": ["Amazon"] * 30,
        }
    )
    inv = pd.DataFrame({"OMS_SKU": ["1052YKGREEN-5XL"], "Total_Inventory": [0]})
    existing = pd.DataFrame(
        {"OMS_SKU": ["1052YKGREEN-5XL"], "PO_Pipeline_Total": [30]}
    )
    sheet = pd.DataFrame(
        {
            "OMS_SKU": ["1052YKGREEN-5XL"],
            "SKU_Sheet_Status": ["Open"],
            "SKU_Sheet_Closed": [False],
            "Lead_Time_From_Sheet": [45.0],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=90,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing,
        sku_status_df=sheet,
        enforce_lead_time_release_gate=True,
    )
    row = po.iloc[0]
    assert float(row["Projected_Running_Days"]) == 30.0
    assert int(row["Lead_Time_Days"]) == 45
    assert int(row["PO_Qty"]) == 60
    assert float(row["Post_PO_Cover_Days_Capped"]) == pytest.approx(90.0, abs=1.0)


def test_po_released_when_projected_below_entered_lead_not_sheet_lead():
    """4032DRSGREEN-S: entered lead 70d; projected ~67d → PO toward 180d target."""
    sku = "4032DRSGREEN-S"
    days = pd.date_range("2026-05-18", periods=30, freq="D")
    sales_rows = []
    for d in days[:8]:
        sales_rows.append(
            {
                "Sku": sku,
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 1,
                "Units_Effective": 1,
                "Source": "Amazon",
            }
        )
    sales = pd.DataFrame(sales_rows)
    inv = pd.DataFrame({"OMS_SKU": [sku], "Total_Inventory": [4]})
    existing = pd.DataFrame({"OMS_SKU": [sku], "PO_Pipeline_Total": [14]})
    sheet = pd.DataFrame(
        {
            "OMS_SKU": [sku],
            "SKU_Sheet_Status": ["Medium Selling"],
            "SKU_Sheet_Closed": [False],
            "Lead_Time_From_Sheet": [45.0],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=70,
        target_days=180,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing,
        sku_status_df=sheet,
        enforce_lead_time_release_gate=True,
        enforce_two_size_minimum=False,
    )
    row = po.iloc[0]
    assert float(row["Projected_Running_Days"]) < 70
    assert int(row["Lead_Time_Days"]) == 45
    assert int(row["PO_Qty"]) > 0


def test_no_po_when_projected_cover_exceeds_sheet_lead_with_zero_entered_lead():
    """lead_time=0 → gate uses sheet lead (45d); projected 68d > 45d → no PO."""
    sku = "1003YKMUSTARD-XXL"
    days = pd.date_range("2026-05-18", periods=30, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": [sku] * 23,
            "TxnDate": days[:23],
            "Transaction Type": ["Shipment"] * 23,
            "Quantity": [1] * 23,
            "Units_Effective": [1] * 23,
            "Source": ["Amazon"] * 23,
        }
    )
    inv = pd.DataFrame({"OMS_SKU": [sku], "Total_Inventory": [33]})
    existing = pd.DataFrame({"OMS_SKU": [sku], "PO_Pipeline_Total": [20]})
    sheet = pd.DataFrame(
        {
            "OMS_SKU": [sku],
            "SKU_Sheet_Status": ["High selling"],
            "SKU_Sheet_Closed": [False],
            "Lead_Time_From_Sheet": [45.0],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=0,
        target_days=180,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing,
        sku_status_df=sheet,
        enforce_lead_time_release_gate=True,
        enforce_two_size_minimum=False,
    )
    row = po.iloc[0]
    assert float(row["Projected_Running_Days"]) > 45
    assert int(row["Lead_Time_Days"]) == 45
    assert int(row["PO_Qty"]) == 0
    assert "sheet lead" in str(row["PO_Block_Reason"]).lower()


def test_po_released_when_projected_below_sheet_lead_with_zero_entered_lead():
    """lead_time=0 → sheet lead 70d; projected ~67d → PO toward 180d target."""
    sku = "4032DRSGREEN-S"
    days = pd.date_range("2026-05-18", periods=30, freq="D")
    sales_rows = []
    for d in days[:8]:
        sales_rows.append(
            {
                "Sku": sku,
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 1,
                "Units_Effective": 1,
                "Source": "Amazon",
            }
        )
    sales = pd.DataFrame(sales_rows)
    inv = pd.DataFrame({"OMS_SKU": [sku], "Total_Inventory": [4]})
    existing = pd.DataFrame({"OMS_SKU": [sku], "PO_Pipeline_Total": [14]})
    sheet = pd.DataFrame(
        {
            "OMS_SKU": [sku],
            "SKU_Sheet_Status": ["Medium Selling"],
            "SKU_Sheet_Closed": [False],
            "Lead_Time_From_Sheet": [70.0],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=0,
        target_days=180,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing,
        sku_status_df=sheet,
        enforce_lead_time_release_gate=True,
        enforce_two_size_minimum=False,
    )
    row = po.iloc[0]
    assert float(row["Projected_Running_Days"]) < 70
    assert int(row["Lead_Time_Days"]) == 70
    assert int(row["PO_Qty"]) > 0


def test_no_po_when_projected_cover_exceeds_entered_lead_time():
    """1003YKMUSTARD-XXL: projected 68d > entered lead 60d → no PO (sheet lead 45 ignored for gate)."""
    sku = "1003YKMUSTARD-XXL"
    days = pd.date_range("2026-05-18", periods=30, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": [sku] * 23,
            "TxnDate": days[:23],
            "Transaction Type": ["Shipment"] * 23,
            "Quantity": [1] * 23,
            "Units_Effective": [1] * 23,
            "Source": ["Amazon"] * 23,
        }
    )
    inv = pd.DataFrame({"OMS_SKU": [sku], "Total_Inventory": [33]})
    existing = pd.DataFrame({"OMS_SKU": [sku], "PO_Pipeline_Total": [20]})
    sheet = pd.DataFrame(
        {
            "OMS_SKU": [sku],
            "SKU_Sheet_Status": ["High selling"],
            "SKU_Sheet_Closed": [False],
            "Lead_Time_From_Sheet": [45.0],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=60,
        target_days=180,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing,
        sku_status_df=sheet,
        enforce_lead_time_release_gate=True,
        enforce_two_size_minimum=False,
    )
    row = po.iloc[0]
    assert float(row["Projected_Running_Days"]) > 60
    assert int(row["PO_Qty"]) == 0
    assert "lead time" in str(row["PO_Block_Reason"]).lower()


def test_pipeline_skus_all_get_po_when_projected_below_entered_lead():
    """Every size with low shelf stock gets PO — not a single-SKU exception."""
    skus = ["4032DRSGREEN-S", "4032DRSGREEN-M", "4032DRSGREEN-XXL"]
    inv_vals = [4, 0, 8]
    pipe_vals = [14, 10, 10]
    days = pd.date_range("2026-05-18", periods=30, freq="D")
    sales_rows = []
    for sku in skus:
        for d in days[:6]:
            sales_rows.append(
                {
                    "Sku": sku,
                    "TxnDate": d,
                    "Transaction Type": "Shipment",
                    "Quantity": 1,
                    "Units_Effective": 1,
                    "Source": "Amazon",
                }
            )
    sales = pd.DataFrame(sales_rows)
    inv = pd.DataFrame({"OMS_SKU": skus, "Total_Inventory": inv_vals})
    existing = pd.DataFrame({"OMS_SKU": skus, "PO_Pipeline_Total": pipe_vals})
    sheet = pd.DataFrame(
        {
            "OMS_SKU": skus,
            "SKU_Sheet_Status": ["Medium Selling"] * 3,
            "SKU_Sheet_Closed": [False] * 3,
            "Lead_Time_From_Sheet": [45.0] * 3,
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=95,
        target_days=180,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing,
        sku_status_df=sheet,
        enforce_lead_time_release_gate=True,
        enforce_two_size_minimum=False,
    )
    rows = po.set_index("OMS_SKU").loc[skus]
    for sku in skus:
        r = rows.loc[sku]
        assert float(r["Projected_Running_Days"]) < 95, sku
        assert float(r["Projected_Running_Days"]) < 180, sku
        assert int(r["PO_Qty"]) > 0, sku


def test_lead_gate_qty_targets_post_po_cover_not_lead_only():
    """Proj ~18d, lead 60d → PO sized toward 180d post-PO cover (not ~65d lead-only top-up)."""
    sku = "LOWCOV-M"
    days = pd.date_range("2026-05-18", periods=30, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": [sku] * 30,
            "TxnDate": days,
            "Transaction Type": ["Shipment"] * 30,
            "Quantity": [1] * 30,
            "Units_Effective": [1] * 30,
            "Source": ["Amazon"] * 30,
        }
    )
    inv = pd.DataFrame({"OMS_SKU": [sku], "Total_Inventory": [0]})
    existing = pd.DataFrame({"OMS_SKU": [sku], "PO_Pipeline_Total": [18]})
    sheet = pd.DataFrame(
        {
            "OMS_SKU": [sku],
            "SKU_Sheet_Status": ["Open"],
            "SKU_Sheet_Closed": [False],
            "Lead_Time_From_Sheet": [60.0],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=60,
        target_days=180,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing,
        sku_status_df=sheet,
        enforce_lead_time_release_gate=True,
        enforce_two_size_minimum=False,
    )
    row = po.iloc[0]
    assert float(row["Projected_Running_Days"]) == pytest.approx(18.0, abs=1.0)
    assert int(row["PO_Qty"]) > 50
    assert float(row["Post_PO_Cover_Days_Capped"]) >= 170.0


def test_multi_size_parent_lifts_siblings_to_post_po_target_with_lead_gate():
    """1317-style: when ≥2 sizes already need PO, lift siblings below target cover in one pass."""
    parent = "1317YKBLUE"
    sizes = {
        f"{parent}-L": (16, 7, 0.457, 50.3),
        f"{parent}-7XL": (18, 1, 0.375, 50.7),
        f"{parent}-XS": (11, 0, 0.207, 53.1),
        f"{parent}-M": (35, 2, 0.391, 94.6),
        f"{parent}-8XL": (13, 2, 0.235, 63.8),
    }
    days = pd.date_range("2026-03-01", periods=45, freq="D")
    sales_rows = []
    for sku, (_inv, _pipe, ads, _proj) in sizes.items():
        qty = max(1, int(round(ads * 30)))
        for d in days[:qty]:
            sales_rows.append(
                {
                    "Sku": sku,
                    "TxnDate": d,
                    "Transaction Type": "Shipment",
                    "Quantity": 1,
                    "Units_Effective": 1,
                    "Source": "Amazon",
                }
            )
    sales = pd.DataFrame(sales_rows)
    inv = pd.DataFrame(
        [{"OMS_SKU": sku, "Total_Inventory": v[0]} for sku, v in sizes.items()]
    )
    existing = pd.DataFrame(
        [{"OMS_SKU": sku, "PO_Pipeline_Total": v[1]} for sku, v in sizes.items()]
    )
    sheet = pd.DataFrame(
        {
            "OMS_SKU": list(sizes.keys()),
            "SKU_Sheet_Status": ["High selling"] * len(sizes),
            "SKU_Sheet_Closed": [False] * len(sizes),
            "Lead_Time_From_Sheet": [45.0] * len(sizes),
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        existing_po_df=existing,
        period_days=45,
        lead_time=60,
        target_days=180,
        grace_days=0,
        demand_basis="Sold",
        safety_pct=0.0,
        sku_status_df=sheet,
        enforce_lead_time_release_gate=True,
        enforce_two_size_minimum=True,
    )
    by_sku = {str(r["OMS_SKU"]): r for _, r in po.iterrows()}
    assert int(by_sku[f"{parent}-L"]["PO_Qty"]) > 0
    assert int(by_sku[f"{parent}-7XL"]["PO_Qty"]) > 0
    assert int(by_sku[f"{parent}-M"]["PO_Qty"]) > 0
    assert int(by_sku[f"{parent}-8XL"]["PO_Qty"]) > 0
    assert float(by_sku[f"{parent}-M"]["Post_PO_Cover_Days_Capped"]) >= 170.0


def test_po_release_not_blocked_just_because_projected_cover_exceeds_lead_time():
    """With ``enforce_lead_time_release_gate=False``, target-only mode still tops up
    when projected cover is below target even if it already exceeds lead time."""
    sales = _minimal_sales()  # ADS ≈ 2/day
    inv = pd.DataFrame({"OMS_SKU": ["TEST-SKU-1"], "Total_Inventory": [80]})
    sheet = pd.DataFrame(
        {
            "OMS_SKU": ["TEST-SKU-1"],
            "SKU_Sheet_Status": ["Open"],
            "SKU_Sheet_Closed": [False],
            "Lead_Time_From_Sheet": [30.0],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=30,
        target_days=90,
        demand_basis="Sold",
        safety_pct=0.0,
        sku_status_df=sheet,
        enforce_lead_time_release_gate=False,
    )
    row = po.iloc[0]
    # Projected cover (≈ 40d) exceeds Lead_Time (30d) but is below target (90d).
    assert float(row["Projected_Running_Days"]) >= float(row["Lead_Time_Days"])
    assert int(row["Gross_PO_Qty"]) > 0
    assert int(row["PO_Qty"]) > 0


def test_po_zero_when_projected_cover_already_meets_target_cover():
    """When projected_days >= target_cover, the formula self-zeroes PO (no
    separate gate needed)."""
    sales = _minimal_sales()  # ADS ≈ 2/day
    # 200 units of stock @ ADS 2 ⇒ 100d projected ≥ 90d target ⇒ no PO needed.
    inv = pd.DataFrame({"OMS_SKU": ["TEST-SKU-1"], "Total_Inventory": [200]})
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=30,
        target_days=90,
        demand_basis="Sold",
        safety_pct=0.0,
    )
    row = po.iloc[0]
    assert int(row["Gross_PO_Qty"]) == 0
    assert int(row["PO_Qty"]) == 0


def test_sheet_without_positive_lead_blocks_po_when_status_sheet_loaded():
    """If a SKU status row has no resolvable lead, do not recommend PO using only the global default."""
    sales = _minimal_sales()
    inv = _minimal_inventory()
    sheet = pd.DataFrame(
        {
            "OMS_SKU": ["TEST-SKU-1"],
            "SKU_Sheet_Status": ["Open"],
            "SKU_Sheet_Closed": [False],
            "Lead_Time_From_Sheet": [float("nan")],
        }
    )
    po_sheet = calculate_po_base(sales, inv, 30, 7, 60, safety_pct=0.0, sku_status_df=sheet)
    po_none = calculate_po_base(
        sales, inv, 30, 7, 60, safety_pct=0.0, sku_status_df=None,
        enforce_lead_time_release_gate=False,
    )
    assert int(po_sheet.iloc[0]["PO_Qty"]) == 0
    assert "lead" in str(po_sheet.iloc[0]["PO_Block_Reason"]).lower()
    assert int(po_none.iloc[0]["PO_Qty"]) > 0
    assert float(po_sheet.iloc[0]["ADS"]) == float(po_none.iloc[0]["ADS"])


def test_sheet_lead_applies_with_inventory_sku_spacing_and_case():
    """Inventory / sheet SKUs are canonicalized so per-SKU lead matches (not stuck on global default)."""
    sales = _minimal_sales()
    inv = pd.DataFrame({"OMS_SKU": ["  test-sku-1\t"], "Total_Inventory": [50]})
    sheet = pd.DataFrame(
        {
            "OMS_SKU": ["TEST-SKU-1"],
            "SKU_Sheet_Status": ["Open"],
            "SKU_Sheet_Closed": [False],
            "Lead_Time_From_Sheet": [88.0],
        }
    )
    po = calculate_po_base(sales, inv, 30, 7, 60, safety_pct=0.0, sku_status_df=sheet)
    assert int(po.iloc[0]["Lead_Time_Days"]) == 88


def test_sheet_lead_rollup_to_parent_when_group_by_parent():
    """Lead sheet lists size-level SKUs; parent-level PO rows pick up max positive lead per parent."""
    days = pd.date_range("2025-11-01", periods=30, freq="D")
    rows = []
    for d in days:
        rows.append(
            {
                "Sku": "MYSTYLE-L",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 2,
                "Units_Effective": 2,
                "Source": "Amz",
            }
        )
    sales = pd.DataFrame(rows)
    inv = pd.DataFrame({"OMS_SKU": ["MYSTYLE"], "Total_Inventory": [50]})
    sheet = pd.DataFrame(
        {
            "OMS_SKU": ["MYSTYLE-L", "MYSTYLE-XL"],
            "SKU_Sheet_Status": ["Open", "Open"],
            "SKU_Sheet_Closed": [False, False],
            "Lead_Time_From_Sheet": [15.0, 25.0],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=7,
        target_days=60,
        safety_pct=0.0,
        sku_status_df=sheet,
        group_by_parent=True,
    )
    row = po[po["OMS_SKU"] == "MYSTYLE"].iloc[0]
    assert int(row["Lead_Time_Days"]) == 25
    # Sales were uploaded as ``MYSTYLE-L``; parent inventory row must inherit rolled-up units.
    assert int(row["Sold_Units"]) == 60  # 30 days × 2 units/day, all in-period shipments


def test_days_left_uses_total_inventory_when_column_present():
    """Days cover reflects Total_Inventory when OMS warehouse column is zero but FBA/total stock exists."""
    sales = _minimal_sales()
    inv = pd.DataFrame(
        {
            "OMS_SKU": ["TEST-SKU-1"],
            "Total_Inventory": [60],
            "OMS_Inventory": [0],
        }
    )
    po = calculate_po_base(sales, inv, 30, 7, 60, safety_pct=0.0)
    assert float(po.iloc[0]["ADS"]) > 0
    assert float(po.iloc[0]["Days_Left"]) > 0


def test_sheet_lead_increases_gross_vs_shorter_global_lead():
    """With lead gate on, a longer sheet lead increases PO when cover is below that lead."""
    sales = _minimal_sales()
    inv = _minimal_inventory()
    short = calculate_po_base(sales, inv, 30, 7, 60, safety_pct=0.0, sku_status_df=None)
    sheet = pd.DataFrame(
        {
            "OMS_SKU": ["TEST-SKU-1"],
            "SKU_Sheet_Status": ["Active"],
            "SKU_Sheet_Closed": [False],
            "Lead_Time_From_Sheet": [90.0],
        }
    )
    long_lead = calculate_po_base(sales, inv, 30, 7, 60, safety_pct=0.0, sku_status_df=sheet)
    assert int(long_lead.iloc[0]["Lead_Time_Days"]) == 90
    assert int(long_lead.iloc[0]["Gross_PO_Qty"]) > int(short.iloc[0]["Gross_PO_Qty"])
    assert float(long_lead.iloc[0]["ADS"]) == float(short.iloc[0]["ADS"])


def test_get_parent_sku_preserves_two_part_numeric_style_code():
    """``AK-1394`` style IDs must not collapse to ``AK`` (digit-only segment is not a size)."""
    assert get_parent_sku("AK-1394") == "AK-1394"
    assert get_parent_sku("AK-1394BROWN-L") == "AK-1394BROWN"
    assert get_parent_sku("1657-M") == "1657"


def test_sheet_lead_from_style_code_applies_to_color_size_variant():
    """Lead sheet lists ``1394``; inventory ``AK-1394BROWN-L`` inherits that lead."""
    days = pd.date_range("2025-11-01", periods=30, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": ["AK-1394BROWN-L"] * 30,
            "TxnDate": days,
            "Transaction Type": ["Shipment"] * 30,
            "Quantity": [1] * 30,
            "Units_Effective": [1] * 30,
            "Source": ["Amazon"] * 30,
        }
    )
    inv = pd.DataFrame({"OMS_SKU": ["AK-1394BROWN-L"], "Total_Inventory": [21]})
    sheet = pd.DataFrame(
        {
            "OMS_SKU": ["1394"],
            "SKU_Sheet_Status": ["Open"],
            "SKU_Sheet_Closed": [False],
            "Lead_Time_From_Sheet": [55.0],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=30,
        target_days=60,
        demand_basis="Sold",
        safety_pct=0.0,
        sku_status_df=sheet,
    )
    row = po.iloc[0]
    assert int(row["Lead_Time_Days"]) == 55
    ads = float(row["ADS"])
    assert ads > 0
    expected_days_left = round(float(row["Total_Inventory"]) / ads, 1)
    expected_projected = round((float(row["Total_Inventory"]) + float(row["PO_Pipeline_Total"])) / ads, 1)
    assert float(row["Days_Left"]) == pytest.approx(expected_days_left, abs=0.05)
    assert float(row["Projected_Running_Days"]) == pytest.approx(expected_projected, abs=0.05)


def test_post_po_target_days_accepts_180_and_above():
    """PO engine must accept editable Post-PO running targets beyond the old 150d UI cap."""
    sales = _minimal_sales()
    inv = pd.DataFrame({"OMS_SKU": ["TEST-SKU-1"], "Total_Inventory": [5]})
    po = calculate_po_base(
        sales, inv, period_days=30, lead_time=7, target_days=180,
        safety_pct=0.0, enforce_lead_time_release_gate=False,
    )
    row = po.iloc[0]
    assert float(row["ADS"]) > 0
    assert int(row["PO_Qty"]) > 0
    assert float(row["Post_PO_Cover_Days_Capped"]) >= 150.0


def test_projected_running_days_uses_inventory_plus_pipeline_over_ads():
    """Projected days should include Total_Inventory + Pipeline (before new PO release)."""
    sales = _minimal_sales()
    inv = pd.DataFrame(
        {
            "OMS_SKU": ["TEST-SKU-1"],
            "Total_Inventory": [40],
            "OMS_Inventory": [10],
        }
    )
    existing = pd.DataFrame({"OMS_SKU": ["TEST-SKU-1"], "PO_Pipeline_Total": [20]})
    po = calculate_po_base(sales, inv, 30, 7, 60, safety_pct=0.0, existing_po_df=existing)
    row = po.iloc[0]
    ads = float(row["ADS"])
    assert ads > 0
    expected = round((40 + 20) / ads, 1)
    expected_days_left = round(40 / ads, 1)
    assert float(row["Projected_Running_Days"]) == pytest.approx(expected, abs=0.05)
    assert float(row["Days_Left"]) == pytest.approx(expected_days_left, abs=0.05)
    # Uses Total_Inventory (40) in numerator, not OMS_Inventory (10) alone.
    assert float(row["Projected_Running_Days"]) != pytest.approx(round((10 + 20) / ads, 1), abs=0.05)


def test_ads_equals_sales_over_effective_days():
    """Steady daily seller across full period: final ADS tracks sold ÷ Eff_Days."""
    rows = []
    for d in pd.date_range("2025-11-01", periods=30, freq="D"):
        rows.append(
            {
                "Sku": "ADS-SKU-1",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 3,
                "Units_Effective": 3,
                "Source": "Amazon",
            }
        )
    sales = pd.DataFrame(rows)
    inv = pd.DataFrame({"OMS_SKU": ["ADS-SKU-1"], "Total_Inventory": [20]})
    po = calculate_po_base(sales, inv, 30, 7, 60, safety_pct=0.0, demand_basis="Sold")
    r = po.iloc[0]
    assert int(r["Eff_Days"]) == 30
    assert float(r["Sold_Units"]) == pytest.approx(90.0, abs=0.01)
    assert float(r["ADS"]) == pytest.approx(3.0, abs=0.02)
    assert float(r["Days_Left"]) == pytest.approx(round(20 / 3.0, 1), abs=0.05)


def test_variant_mode_parent_inventory_sku_falls_back_to_child_sales_for_ads():
    """
    Some inventory files contain parent-like OMS_SKU (no size suffix) even in variant mode.
    When exact sales key is missing but child sizes sold, PO should inherit parent rollup.
    """
    rows = []
    for d in pd.date_range("2025-11-01", periods=30, freq="D"):
        rows.append(
            {
                "Sku": "1007YKBLACK-XL",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 2,
                "Units_Effective": 2,
                "Source": "Amazon",
            }
        )
    sales = pd.DataFrame(rows)
    inv = pd.DataFrame({"OMS_SKU": ["1007YKBLACK"], "Total_Inventory": [40]})
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=7,
        target_days=60,
        demand_basis="Sold",
        safety_pct=0.0,
        group_by_parent=False,
    )
    row = po.iloc[0]
    assert int(row["Sold_Units"]) == 60
    assert float(row["ADS"]) == pytest.approx(2.0, abs=0.02)
    assert float(row["Flat30_ADS"]) > 0
    assert float(row["LY_ADS"]) >= 0


def test_variant_mode_parent_fallback_handles_non_delimited_size_suffix():
    """Sales SKU like ``1007YKBLACKXXL`` should roll up to parent ``1007YKBLACK``."""
    rows = []
    for d in pd.date_range("2025-11-01", periods=30, freq="D"):
        rows.append(
            {
                "Sku": "1007YKBLACKXXL",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 1,
                "Units_Effective": 1,
                "Source": "Amazon",
            }
        )
    sales = pd.DataFrame(rows)
    inv = pd.DataFrame({"OMS_SKU": ["1007YKBLACK"], "Total_Inventory": [100]})
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=7,
        target_days=60,
        demand_basis="Sold",
        safety_pct=0.0,
        group_by_parent=False,
    )
    row = po.iloc[0]
    assert int(row["Sold_Units"]) == 30
    assert float(row["ADS"]) == pytest.approx(1.0, abs=0.02)
    assert int(row["Eff_Days"]) == 30


def test_variant_mode_parent_token_uses_full_style_rollup_even_with_exact_row():
    """
    If inventory row is parent-style (1007YKBLACK), it should include exact row sales
    plus child size sales, not just the exact-key line.
    """
    rows = []
    # One exact parent-key sale
    rows.append(
        {
            "Sku": "1007YKBLACK",
            "TxnDate": pd.Timestamp("2025-11-01"),
            "Transaction Type": "Shipment",
            "Quantity": 1,
            "Units_Effective": 1,
            "Source": "Amazon",
        }
    )
    # Child-size sales for same style
    for d in pd.date_range("2025-11-02", periods=29, freq="D"):
        rows.append(
            {
                "Sku": "1007YKBLACK-5XL",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 1,
                "Units_Effective": 1,
                "Source": "Amazon",
            }
        )
    sales = pd.DataFrame(rows)
    inv = pd.DataFrame({"OMS_SKU": ["1007YKBLACK"], "Total_Inventory": [447]})
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=90,
        lead_time=30,
        target_days=210,
        demand_basis="Sold",
        safety_pct=0.0,
        group_by_parent=False,
    )
    row = po.iloc[0]
    assert int(row["Sold_Units"]) == 30
    assert float(row["ADS"]) == pytest.approx(1.0, abs=0.02)


def test_ship_units_150d_shows_broader_context_than_period_sold_units():
    """Sold_Units stays period-based; Ship_Units_150d reflects broader shipment context."""
    rows = []
    # Older shipments outside 30d window but inside 150d window.
    for d in pd.date_range("2025-07-01", periods=20, freq="D"):
        rows.append(
            {
                "Sku": "WIDE-SKU-1",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 1,
                "Units_Effective": 1,
                "Source": "Amazon",
            }
        )
    # Recent 30d window shipments.
    for d in pd.date_range("2025-11-01", periods=10, freq="D"):
        rows.append(
            {
                "Sku": "WIDE-SKU-1",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 1,
                "Units_Effective": 1,
                "Source": "Amazon",
            }
        )
    sales = pd.DataFrame(rows)
    inv = pd.DataFrame({"OMS_SKU": ["WIDE-SKU-1"], "Total_Inventory": [50]})
    po = calculate_po_base(sales, inv, 30, 7, 60, safety_pct=0.0, group_by_parent=False)
    row = po.iloc[0]
    assert int(row["Sold_Units"]) == 10
    assert int(row["Ship_Units_150d"]) == 30


def test_ads_caps_sparse_six_units_in_thirty_day_window():
    """6 sold in 30d with short Eff_Days must not exceed sold÷30 (0.2), not 6÷6=1.0."""
    rows = []
    # One unit on each of 6 days spread across ~4 weeks (sparse intermittent pattern).
    for d in pd.to_datetime(["2026-05-01", "2026-05-05", "2026-05-10", "2026-05-18", "2026-05-22", "2026-05-28"]):
        rows.append(
            {
                "Sku": "1050YKBLUE-L",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 1,
                "Units_Effective": 1,
                "Source": "Amazon",
            }
        )
    sales = pd.DataFrame(rows)
    inv = pd.DataFrame({"OMS_SKU": ["1050YKBLUE-L"], "Total_Inventory": [11]})
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=60,
        target_days=180,
        safety_pct=0.0,
        use_ly_fallback=False,
        demand_basis="Sold",
    )
    row = po.iloc[0]
    assert int(row["Sold_Units"]) == 6
    assert float(row["Recent_ADS"]) == pytest.approx(1.0, abs=0.05)  # raw sold÷eff before cap
    assert float(row["ADS"]) == pytest.approx(0.2, abs=0.02)  # capped to 6÷30


def test_ads_falls_back_to_ly_when_recent_is_zero():
    """If recent sales are zero but LY window has demand, ADS should not stay zero."""
    rows = []
    # Anchor max_date with another SKU in current window.
    for d in pd.date_range("2025-11-01", periods=5, freq="D"):
        rows.append(
            {
                "Sku": "ANCHOR-SKU-1",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 1,
                "Units_Effective": 1,
                "Source": "Amazon",
            }
        )
    # Target SKU: sales only in LY-aligned window (not recent).
    for d in pd.date_range("2024-11-01", periods=30, freq="D"):
        rows.append(
            {
                "Sku": "LY-ONLY-SKU-1",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 2,
                "Units_Effective": 2,
                "Source": "Amazon",
            }
        )
    sales = pd.DataFrame(rows)
    inv = pd.DataFrame({"OMS_SKU": ["LY-ONLY-SKU-1", "ANCHOR-SKU-1"], "Total_Inventory": [40, 10]})
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=7,
        target_days=60,
        demand_basis="Sold",
        safety_pct=0.0,
        group_by_parent=False,
    )
    row = po[po["OMS_SKU"] == "LY-ONLY-SKU-1"].iloc[0]
    assert float(row["Recent_ADS"]) == 0.0
    assert float(row["LY_ADS"]) > 0.0
    assert float(row["ADS"]) > 0.0


def test_recent_window_ignores_single_future_outlier_date():
    """
    One future-dated row from any SKU must not shift max_date and blank recent ADS for all.
    """
    rows = []
    # Target SKU has healthy recent shipments.
    for d in pd.date_range(pd.Timestamp.now().normalize() - pd.Timedelta(days=9), periods=10, freq="D"):
        rows.append(
            {
                "Sku": "DATE-ANCHOR-SKU-1",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 2,
                "Units_Effective": 2,
                "Source": "Amazon",
            }
        )
    # Outlier future row for another SKU (bad source date).
    rows.append(
        {
            "Sku": "BAD-FUTURE-SKU",
            "TxnDate": pd.Timestamp("2099-01-01"),
            "Transaction Type": "Shipment",
            "Quantity": 1,
            "Units_Effective": 1,
            "Source": "Amazon",
        }
    )
    sales = pd.DataFrame(rows)
    inv = pd.DataFrame(
        {
            "OMS_SKU": ["DATE-ANCHOR-SKU-1", "BAD-FUTURE-SKU"],
            "Total_Inventory": [50, 10],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=7,
        target_days=60,
        demand_basis="Sold",
        safety_pct=0.0,
        group_by_parent=False,
    )
    r = po[po["OMS_SKU"] == "DATE-ANCHOR-SKU-1"].iloc[0]
    assert int(r["Sold_Units"]) == 20
    assert float(r["Recent_ADS"]) > 0


def test_po_qty_is_target_cover_balance_days_based():
    """Released PO follows target-cover balance-days formula."""
    rows = []
    for d in pd.date_range("2025-11-01", periods=30, freq="D"):
        rows.append(
            {
                "Sku": "LEAD-CAP-SKU-1",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 2,
                "Units_Effective": 2,
                "Source": "Amazon",
            }
        )
    sales = pd.DataFrame(rows)
    inv = pd.DataFrame({"OMS_SKU": ["LEAD-CAP-SKU-1"], "Total_Inventory": [10]})
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=210,
        demand_basis="Sold",
        safety_pct=0.0,
        group_by_parent=False,
        enforce_lead_time_release_gate=False,
    )
    row = po.iloc[0]
    assert float(row["ADS"]) > 0
    ads = float(row["ADS"])
    expected_raw = max((ads * 210.0) - 10.0, 0.0)
    import math
    expected_po_target = int(math.ceil(expected_raw / 5.0) * 5.0)
    expected_po = expected_po_target
    expected_proj = round(10.0 / ads, 1)
    assert int(row["PO_Qty"]) == expected_po
    assert float(row["Projected_Running_Days"]) == expected_proj


# ── Daily Inventory History parser + PO override ─────────────────────────────


def _wide_inv_history_workbook():
    """Build a 2-sheet wide-format Excel matching the production export layout."""
    import io
    import pandas as pd

    dates = pd.date_range("2025-12-08", periods=30, freq="D")
    cols = ["Total Inv.", "Total"] + list(range(len(dates)))

    oms_rows = [
        ["Item SkuCode", "Item"] + [d for d in dates],
        ["INV-VAR-1", "INV-PARENT"] + [(0 if i < 10 else 5) for i in range(len(dates))],
        ["INV-VAR-2", "INV-PARENT"] + [3 for _ in dates],
    ]
    oms_df = pd.DataFrame(oms_rows, columns=cols)

    amz_rows = [
        ["Item SkuCode", "Item SkuCode"] + [d for d in dates],
        ["INV-PLVAR-1", "INV-PLPARENT"] + [0 for _ in dates],
    ]
    amz_df = pd.DataFrame(amz_rows, columns=["Total Inv.", "SKU"] + list(range(len(dates))))

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        oms_df.to_excel(w, sheet_name="OMS", index=False)
        amz_df.to_excel(w, sheet_name="Amazon Inventory", index=False)
    buf.seek(0)
    return buf


def test_daily_inventory_date_map_uses_row1_when_row0_missing_trailing_column():
    """New 'today' columns often have the date stamp in row 1, not row 0."""
    from backend.services.daily_inventory_history import _parse_one_sheet

    df = pd.DataFrame(
        [
            ["Item SkuCode", "Item", 100, 110, ""],
            ["", "", "2026-05-17", "2026-05-18", "2026-05-19"],
            ["SKU-A", "PARENT", 5, 6, 7],
        ]
    )
    tall = _parse_one_sheet(df, {})
    assert not tall.empty
    by_date = tall.groupby("Date")["Qty"].sum()
    assert float(by_date[pd.Timestamp("2026-05-19")]) == 7.0


def test_merge_inventory_history_keeps_newest_snapshot_day():
    from backend.services.daily_inventory_history import merge_inventory_history

    old = pd.DataFrame(
        {
            "OMS_SKU": ["SKU-A", "SKU-A"],
            "Date": pd.to_datetime(["2026-05-17", "2026-05-18"]),
            "Qty": [10.0, 11.0],
        }
    )
    new = pd.DataFrame(
        {
            "OMS_SKU": ["SKU-A"],
            "Date": pd.to_datetime(["2026-05-19"]),
            "Qty": [20.0],
        }
    )
    merged = merge_inventory_history(old, new)
    assert len(merged) == 3
    row = merged[merged["Date"] == pd.Timestamp("2026-05-19")].iloc[0]
    assert int(row["Qty"]) == 20


def test_daily_inventory_history_parser_picks_variant_columns():
    from backend.services.daily_inventory_history import parse_daily_inventory_history_upload

    buf = _wide_inv_history_workbook()
    df = parse_daily_inventory_history_upload(buf, "Daily Inventory History.xlsx")
    assert not df.empty
    skus = set(df["OMS_SKU"].unique())
    assert "INV-VAR-1" in skus
    assert "INV-VAR-2" in skus
    assert df["Qty"].min() >= 0


def test_effective_days_from_history_counts_in_stock_days():
    from backend.services.daily_inventory_history import (
        effective_days_from_history,
        parse_daily_inventory_history_upload,
    )

    buf = _wide_inv_history_workbook()
    df = parse_daily_inventory_history_upload(buf, "x.xlsx")
    end = pd.Timestamp(df["Date"].max()).normalize()
    start = end - pd.Timedelta(days=29)
    eff = effective_days_from_history(df, start, end)
    by_sku = {r["OMS_SKU"]: int(r["Eff_Days_Inventory"]) for _, r in eff.iterrows()}
    assert by_sku["INV-VAR-1"] == 20  # 30-day window, OOS first 10 days
    assert by_sku["INV-VAR-2"] == 30
    assert by_sku["INV-PLVAR-1"] == 0


def test_po_uses_inventory_history_eff_days_to_lift_ads():
    """ADS denominator should drop to in-stock days, lifting Recent_ADS."""
    sales_dates = pd.date_range("2025-12-18", periods=20, freq="D")
    sales = pd.DataFrame(
        [
            {
                "Sku": "INV-OVR",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 2,
                "Units_Effective": 2,
                "Source": "Amazon",
            }
            for d in sales_dates
        ]
    )
    inv = pd.DataFrame({"OMS_SKU": ["INV-OVR"], "Total_Inventory": [10]})

    po_plain = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=90,
        demand_basis="Sold",
        safety_pct=0.0,
        group_by_parent=False,
    )
    plain = po_plain.iloc[0]
    assert int(plain["Eff_Days"]) == 20

    inv_hist = pd.DataFrame(
        {
            "OMS_SKU": ["INV-OVR"] * 30,
            "Date": pd.date_range("2025-12-08", periods=30, freq="D"),
            "Qty": [0] * 10 + [5] * 20,
        }
    )
    po_hist = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=90,
        demand_basis="Sold",
        safety_pct=0.0,
        group_by_parent=False,
        inventory_history_df=inv_hist,
    )
    hist = po_hist.iloc[0]
    assert int(hist["Eff_Days_Inventory"]) == 20
    assert int(hist["Eff_Days"]) == 20

    inv_hist_short = pd.DataFrame(
        {
            "OMS_SKU": ["INV-OVR"] * 30,
            "Date": pd.date_range("2025-12-08", periods=30, freq="D"),
            "Qty": [0] * 25 + [5] * 5,
        }
    )
    po_short = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=90,
        demand_basis="Sold",
        safety_pct=0.0,
        group_by_parent=False,
        inventory_history_df=inv_hist_short,
    )
    short = po_short.iloc[0]
    assert int(short["Eff_Days_Inventory"]) == 5
    assert int(short["Eff_Days"]) == 5
    assert float(short["Recent_ADS"]) > float(plain["Recent_ADS"])


def test_sparse_sales_use_active_eff_days_not_scaled_inventory_span():
    """4032DRSGREEN-style: sparse sales over few days must not be diluted by long in-stock span."""
    sale_days = pd.to_datetime(["2026-05-20", "2026-05-22", "2026-05-25", "2026-05-28"])
    sales_rows = []
    for d in sale_days:
        sales_rows.append(
            {
                "Sku": "4032DRSGREEN-L",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 3,
                "Units_Effective": 3,
                "Source": "Amazon",
            }
        )
    sales = pd.DataFrame(sales_rows)
    inv = pd.DataFrame({"OMS_SKU": ["4032DRSGREEN-L"], "Total_Inventory": [1]})
    existing_po = pd.DataFrame(
        {
            "OMS_SKU": ["4032DRSGREEN-L"],
            "PO_Pipeline_Total": [28],
            "Pending_Cutting": [21],
            "Balance_to_Dispatch": [7],
        }
    )
    sheet = pd.DataFrame(
        {
            "OMS_SKU": ["4032DRSGREEN-L"],
            "SKU_Sheet_Status": ["Medium Selling"],
            "SKU_Sheet_Closed": [False],
            "Lead_Time_From_Sheet": [45.0],
        }
    )
    # 12 in-stock days inside a 14-day snapshot window → scale would push Eff_Days to ~26.
    hist_dates = pd.date_range("2026-05-15", periods=14, freq="D")
    inv_hist = pd.DataFrame(
        {
            "OMS_SKU": ["4032DRSGREEN-L"] * 14,
            "Date": hist_dates,
            "Qty": [2 if i < 12 else 0 for i in range(14)],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=90,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing_po,
        sku_status_df=sheet,
        inventory_history_df=inv_hist,
        enforce_lead_time_release_gate=False,
    )
    row = po.loc[po["OMS_SKU"] == "4032DRSGREEN-L"].iloc[0]
    assert int(row["Sold_Units"]) == 12
    assert int(row["Eff_Days_Inventory"]) == 12
    assert int(row["Eff_Days"]) <= 9, (
        f"Expected active sales span (~9d), not scaled inventory span (~26d); got {row['Eff_Days']}"
    )
    assert float(row["ADS"]) >= 0.4  # period cap: 12 sold ÷ 30d
    assert int(row["PO_Qty"]) > 0
    assert float(row["Projected_Running_Days"]) < 95

    po_gated = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=95,
        target_days=135,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing_po,
        sku_status_df=sheet,
        inventory_history_df=inv_hist,
        enforce_lead_time_release_gate=True,
    )
    gated = po_gated.loc[po_gated["OMS_SKU"] == "4032DRSGREEN-L"].iloc[0]
    assert int(gated["PO_Qty"]) > 0
    assert float(gated["Projected_Running_Days"]) < 95
    assert "lead time" not in str(gated["PO_Block_Reason"]).lower()


def test_intermittent_sales_use_distinct_txn_days_not_calendar_span():
    """When few sale days are spread across a wide calendar, ADS must use distinct txn days."""
    sale_days = pd.to_datetime(
        ["2026-05-01", "2026-05-06", "2026-05-11", "2026-05-16", "2026-05-21", "2026-05-26"]
    )
    sales = pd.DataFrame(
        {
            "Sku": ["4032DRSGREEN-L"] * 6,
            "TxnDate": sale_days,
            "Transaction Type": ["Shipment"] * 6,
            "Quantity": [1] * 6,
            "Units_Effective": [1] * 6,
            "Source": ["Amazon"] * 6,
        }
    )
    inv = pd.DataFrame({"OMS_SKU": ["4032DRSGREEN-L"], "Total_Inventory": [1]})
    existing_po = pd.DataFrame(
        {
            "OMS_SKU": ["4032DRSGREEN-L"],
            "PO_Pipeline_Total": [28],
            "Pending_Cutting": [21],
            "Balance_to_Dispatch": [7],
        }
    )
    hist_dates = pd.date_range("2026-05-01", periods=30, freq="D")
    inv_hist = pd.DataFrame(
        {
            "OMS_SKU": ["4032DRSGREEN-L"] * 30,
            "Date": hist_dates,
            "Qty": [2] * 30,
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=150,
        target_days=200,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing_po,
        inventory_history_df=inv_hist,
        enforce_lead_time_release_gate=True,
    )
    row = po.loc[po["OMS_SKU"] == "4032DRSGREEN-L"].iloc[0]
    assert int(row["Sold_Units"]) == 6
    assert int(row["Eff_Days"]) <= 6, (
        f"26-day calendar span must collapse to 6 distinct sale days; got {row['Eff_Days']}"
    )
    assert float(row["ADS"]) >= 0.2  # period cap: 6 sold ÷ 30d
    assert int(row["PO_Qty"]) > 0
    assert float(row["Projected_Running_Days"]) < 150


def test_bursty_sales_keep_calendar_eff_days_not_distinct_only():
    """Bursty demand must not collapse Eff_Days to distinct txn days (inflates ADS / PO)."""
    sale_days = pd.to_datetime(
        ["2026-05-01", "2026-05-10", "2026-05-20", "2026-05-30"]
    )
    sales_rows = []
    for d in sale_days:
        for _ in range(5):
            sales_rows.append(
                {
                    "Sku": "BURST-SKU-M",
                    "TxnDate": d,
                    "Transaction Type": "Shipment",
                    "Quantity": 1,
                    "Units_Effective": 1,
                    "Source": "Amazon",
                }
            )
    sales = pd.DataFrame(sales_rows)
    inv = pd.DataFrame({"OMS_SKU": ["BURST-SKU-M"], "Total_Inventory": [80]})
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=135,
        demand_basis="Sold",
        safety_pct=0.0,
    )
    row = po.loc[po["OMS_SKU"] == "BURST-SKU-M"].iloc[0]
    assert int(row["Sold_Units"]) == 20
    assert int(row["Eff_Days"]) >= 29, (
        f"Expected ~30d calendar span for bursty seller, not 4 distinct days; got {row['Eff_Days']}"
    )
    assert float(row["ADS"]) < 1.5
    assert int(row["PO_Qty"]) < 200


def test_low_volume_sales_keep_diluted_eff_days_not_distinct_inflation():
    """≤5 sold in the ADS window must not collapse to Eff_Days=1–5 (PO explosion)."""
    sale_days = pd.to_datetime(["2026-05-01", "2026-05-15"])
    sales = pd.DataFrame(
        {
            "Sku": ["LOWVOL-SKU-M", "LOWVOL-SKU-M"],
            "TxnDate": sale_days,
            "Transaction Type": ["Shipment", "Shipment"],
            "Quantity": [2, 3],
            "Units_Effective": [2, 3],
            "Source": ["Amazon", "Amazon"],
        }
    )
    inv = pd.DataFrame({"OMS_SKU": ["LOWVOL-SKU-M"], "Total_Inventory": [10]})
    hist_dates = pd.date_range("2026-05-01", periods=30, freq="D")
    inv_hist = pd.DataFrame(
        {
            "OMS_SKU": ["LOWVOL-SKU-M"] * 30,
            "Date": hist_dates,
            "Qty": [5] * 30,
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=135,
        demand_basis="Sold",
        safety_pct=0.0,
        inventory_history_df=inv_hist,
        enforce_lead_time_release_gate=False,
    )
    row = po.loc[po["OMS_SKU"] == "LOWVOL-SKU-M"].iloc[0]
    assert int(row["Sold_Units"]) == 5
    assert int(row["Eff_Days"]) >= 15, (
        f"Low-volume SKU should keep calendar/period Eff_Days, not {row['Eff_Days']}"
    )
    assert float(row["ADS"]) < 0.5
    assert int(row["PO_Qty"]) < 150


def test_inventory_matrix_does_not_shorten_eff_days_for_light_non_sparse_sellers():
    """Wide matrix must not crush Eff_Days to 1–3 on ≤7 sold / calendar sellers."""
    sale_days = pd.to_datetime(["2026-05-05", "2026-05-20"])
    sales = pd.DataFrame(
        {
            "Sku": ["MATRIX-LIGHT-M", "MATRIX-LIGHT-M"],
            "TxnDate": sale_days,
            "Transaction Type": ["Shipment", "Shipment"],
            "Quantity": [3, 3],
            "Units_Effective": [3, 3],
            "Source": ["Amazon", "Amazon"],
        }
    )
    inv = pd.DataFrame({"OMS_SKU": ["MATRIX-LIGHT-M"], "Total_Inventory": [8]})
    hist_dates = pd.date_range("2026-05-01", periods=30, freq="D")
    inv_hist = pd.DataFrame(
        {
            "OMS_SKU": ["MATRIX-LIGHT-M"] * 30,
            "Date": hist_dates,
            "Qty": [5 if i in (4, 19) else 0 for i in range(30)],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=180,
        demand_basis="Sold",
        safety_pct=0.0,
        inventory_history_df=inv_hist,
        enforce_lead_time_release_gate=False,
    )
    row = po.loc[po["OMS_SKU"] == "MATRIX-LIGHT-M"].iloc[0]
    assert int(row["Sold_Units"]) == 6
    assert int(row["Eff_Days"]) >= 15
    assert float(row["ADS"]) < 0.5
    assert int(row["PO_Qty"]) < 200


def test_4032_drsgreen_family_gets_po_with_pipeline_and_inventory_history():
    """Regression: parent 4032DRSGREEN must not be zeroed when warehouse stock is low but pipeline exists."""
    base_sales = []
    for sku, units, days in (
        ("4032DRSGREEN-L", 3, ["2026-05-20", "2026-05-22", "2026-05-25", "2026-05-28"]),
        ("4032DRSGREEN-M", 2, ["2026-05-18", "2026-05-24", "2026-05-27"]),
        ("4032DRSGREEN-XL", 4, ["2026-05-19", "2026-05-21", "2026-05-26", "2026-05-29"]),
    ):
        for d in days:
            for _ in range(units):
                base_sales.append(
                    {
                        "Sku": sku,
                        "TxnDate": pd.Timestamp(d),
                        "Transaction Type": "Shipment",
                        "Quantity": 1,
                        "Units_Effective": 1,
                        "Source": "Amazon",
                    }
                )
    sales = pd.DataFrame(base_sales)
    inv = pd.DataFrame(
        {
            "OMS_SKU": ["4032DRSGREEN-L", "4032DRSGREEN-M", "4032DRSGREEN-XL"],
            "Total_Inventory": [1, 1, 0],
        }
    )
    existing_po = pd.DataFrame(
        {
            "OMS_SKU": ["4032DRSGREEN-L", "4032DRSGREEN-M", "4032DRSGREEN-XL"],
            "PO_Pipeline_Total": [28, 31, 1],
            "Pending_Cutting": [21, 21, 0],
            "Balance_to_Dispatch": [7, 10, 1],
        }
    )
    sheet = pd.DataFrame(
        {
            "OMS_SKU": ["4032DRSGREEN-L", "4032DRSGREEN-M", "4032DRSGREEN-XL"],
            "SKU_Sheet_Status": ["Medium Selling"] * 3,
            "SKU_Sheet_Closed": [False] * 3,
            "Lead_Time_From_Sheet": [45.0] * 3,
        }
    )
    hist_dates = pd.date_range("2026-05-15", periods=14, freq="D")
    inv_hist_rows = []
    for sku in ("4032DRSGREEN-L", "4032DRSGREEN-M", "4032DRSGREEN-XL"):
        for i, d in enumerate(hist_dates):
            inv_hist_rows.append(
                {"OMS_SKU": sku, "Date": d, "Qty": 2 if i < 12 else 0}
            )
    inv_hist = pd.DataFrame(inv_hist_rows)
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=95,
        target_days=135,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing_po,
        sku_status_df=sheet,
        inventory_history_df=inv_hist,
        enforce_lead_time_release_gate=True,
        enforce_two_size_minimum=True,
    )
    rows = po[po["OMS_SKU"].str.startswith("4032DRSGREEN-")].set_index("OMS_SKU")
    assert int(rows.loc["4032DRSGREEN-L", "PO_Qty"]) > 0
    assert int(rows.loc["4032DRSGREEN-XL", "PO_Qty"]) > 0
    # M already has ~160d projected cover from pipeline — no fresh PO required.
    assert float(rows.loc["4032DRSGREEN-M", "Projected_Running_Days"]) > 90
    assert float(rows.loc["4032DRSGREEN-L", "Projected_Running_Days"]) < 95


def test_oos_size_gets_po_when_sibling_still_selling():
    """1361YKBLUE-XL exhausted: impute ADS from XXL so the size still gets a PO."""
    days_out = pd.date_range("2026-05-05", periods=20, freq="D")
    sales = pd.DataFrame(
        [
            {
                "Sku": "1361YKBLUE-XXL",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 4,
                "Units_Effective": 4,
                "Source": "Amazon",
            }
            for d in days_out
        ]
    )
    inv = pd.DataFrame(
        {
            "OMS_SKU": ["1361YKBLUE-XL", "1361YKBLUE-XXL"],
            "Total_Inventory": [0, 55],
        }
    )
    inv_hist = pd.DataFrame(
        [
            {
                "OMS_SKU": "1361YKBLUE-XL",
                "Date": d,
                "Qty": 5 if d < pd.Timestamp("2026-05-10") else 0,
            }
            for d in pd.date_range("2026-04-10", periods=30, freq="D")
        ]
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=135,
        demand_basis="Net",
        safety_pct=0.0,
        inventory_history_df=inv_hist,
    )
    xl = po.loc[po["OMS_SKU"] == "1361YKBLUE-XL"].iloc[0]
    xxl = po.loc[po["OMS_SKU"] == "1361YKBLUE-XXL"].iloc[0]
    assert int(xl["Net_Units"]) == 0
    assert int(xxl["Net_Units"]) > 0
    assert int(xl["Eff_Days_Inventory"]) > 0
    assert float(xl["Eff_Days"]) > 0
    assert float(xl["Recent_ADS"]) > 0
    assert float(xl["ADS"]) > 0
    assert int(xl["Gross_PO_Qty"]) > 0
    assert int(xl["PO_Qty"]) > 0


def test_pipeline_only_oos_size_gets_inventory_eff_days_before_ghost_injection():
    """XL in existing PO but missing from inventory must still get Eff_Days from history."""
    dates = list(pd.date_range("2026-05-01", periods=30, freq="D"))
    inv_hist = pd.DataFrame(
        [
            {"OMS_SKU": "1361YKBLUE-XL", "Date": d, "Qty": 5 if i < 8 else 0}
            for i, d in enumerate(dates)
        ]
    )
    sales = pd.DataFrame(
        [
            {
                "Sku": "1361YKBLUE-XXL",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 4,
                "Units_Effective": 4,
                "Source": "Amazon",
            }
            for d in pd.date_range("2026-05-15", periods=15, freq="D")
        ]
    )
    inv = pd.DataFrame({"OMS_SKU": ["1361YKBLUE-XXL"], "Total_Inventory": [55]})
    existing_po = pd.DataFrame(
        {
            "OMS_SKU": ["1361YKBLUE-XL", "1361YKBLUE-XXL"],
            "PO_Pipeline_Total": [19, 5],
            "Balance_to_Dispatch": [19, 5],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=135,
        demand_basis="Net",
        safety_pct=0.0,
        existing_po_df=existing_po,
        inventory_history_df=inv_hist,
    )
    xl = po.loc[po["OMS_SKU"] == "1361YKBLUE-XL"].iloc[0]
    assert int(xl["Eff_Days_Inventory"]) == 8
    assert int(xl["Eff_Days"]) == 8
    assert float(xl["ADS"]) > 0
    assert int(xl["Gross_PO_Qty"]) > 0


def test_inventory_history_xl_xl_key_joins_to_canonical_xl():
    """Daily history rows keyed as 1361YKBLUE-XL-XL must attach to 1361YKBLUE-XL."""
    dates = list(pd.date_range("2026-05-01", periods=30, freq="D"))
    inv_hist = pd.DataFrame(
        [
            {"OMS_SKU": "1361YKBLUE-XL-XL", "Date": d, "Qty": 5 if i < 8 else 0}
            for i, d in enumerate(dates)
        ]
    )
    sales = pd.DataFrame(
        [
            {
                "Sku": "1361YKBLUE-XXL",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 4,
                "Units_Effective": 4,
                "Source": "Amazon",
            }
            for d in pd.date_range("2026-05-15", periods=15, freq="D")
        ]
    )
    inv = pd.DataFrame(
        {"OMS_SKU": ["1361YKBLUE-XL", "1361YKBLUE-XXL"], "Total_Inventory": [0, 55]}
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=135,
        demand_basis="Net",
        safety_pct=0.0,
        inventory_history_df=inv_hist,
    )
    xl = po.loc[po["OMS_SKU"] == "1361YKBLUE-XL"].iloc[0]
    assert int(xl["Eff_Days_Inventory"]) == 8
    assert int(xl["Eff_Days"]) == 8
    assert float(xl["ADS"]) > 0


def test_urgent_all_sizes_ghost_rows_use_canonical_skus_not_mapping_keys():
    """Part B must not emit 1361YKBLUE-XL-XL when mapping keys are duplicate-suffix IDs."""
    import json
    from pathlib import Path

    from backend.services.po_engine import calculate_po_base

    mapping_path = Path(__file__).resolve().parents[1] / "backend/data/yash_sku_mapping_master.json"
    sku_mapping = json.loads(mapping_path.read_text())
    days = pd.date_range("2026-05-01", periods=20, freq="D")
    sales = pd.DataFrame(
        [
            {
                "Sku": "1361YKBLUE-XXL",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 4,
                "Units_Effective": 4,
                "Source": "Amazon",
            }
            for d in days
        ]
    )
    inv = pd.DataFrame({"OMS_SKU": ["1361YKBLUE-XXL"], "Total_Inventory": [55]})
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=135,
        demand_basis="Net",
        safety_pct=0.0,
        sku_mapping=sku_mapping,
        urgent_all_sizes_days=45,
    )
    yk = po[po["OMS_SKU"].astype(str).str.contains("1361YKBLUE", na=False)]
    skus = set(yk["OMS_SKU"].astype(str))
    assert "1361YKBLUE-L-L" not in skus
    assert "1361YKBLUE-XL-XL" not in skus
    assert "1361YKBLUE-XXL-XXL" not in skus
    assert "1361YKBLUE-L" in skus
    assert "1361YKBLUE-XL" in skus
    assert "1361YKBLUE-XXL" in skus
    xxl = yk.loc[yk["OMS_SKU"] == "1361YKBLUE-XXL"].iloc[0]
    assert int(xxl["Net_Units"]) > 0


def test_seasonal_window_includes_two_forward_months_for_june_run():
    """June PO run uses Jun+Jul+Aug prior-year history (captures July–August peak)."""
    from backend.services.po_engine import _seasonal_adjacent_months_ads

    sku = "PEAK-M"
    rows = []
    for d in pd.date_range("2025-08-01", periods=31, freq="D"):
        rows.append(
            {
                "Sku": sku,
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 3,
                "Units_Effective": 3,
            }
        )
    sales = pd.DataFrame(rows)
    out = _seasonal_adjacent_months_ads(
        sales, pd.Timestamp("2026-06-15"), False, "Sold", months_forward=2
    )
    assert not out.empty
    assert float(out.loc[out["OMS_SKU"] == sku, "Seasonal_Month_ADS"].iloc[0]) > 0.5


def test_ly_and_seasonal_lift_ads_when_recent_window_is_weak():
    """Weak current month but strong same season last year → ADS uses seasonal/LY uplift."""
    sku = "UPLIFT-M"
    frames = []
    for d in pd.date_range("2026-05-18", periods=30, freq="D"):
        if d.day % 6 == 0:
            frames.append(
                {
                    "Sku": sku,
                    "TxnDate": d,
                    "Transaction Type": "Shipment",
                    "Quantity": 1,
                    "Units_Effective": 1,
                }
            )
    for d in pd.date_range("2025-07-01", periods=62, freq="D"):
        frames.append(
            {
                "Sku": sku,
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 2,
                "Units_Effective": 2,
            }
        )
    sales = pd.concat([pd.DataFrame(frames)], ignore_index=True)
    inv = pd.DataFrame({"OMS_SKU": [sku], "Total_Inventory": [10]})
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=60,
        target_days=180,
        demand_basis="Sold",
        safety_pct=0.0,
        use_seasonality=True,
        use_ly_fallback=True,
    )
    row = po.iloc[0]
    recent = float(row["Recent_ADS"])
    seasonal = float(row["Seasonal_Month_ADS"])
    ads = float(row["ADS"])
    assert recent < 0.3
    assert seasonal > recent
    assert ads >= seasonal


def test_high_cover_style_no_po_when_projected_exceeds_lead_gate():
    """Excel rule: no PO when projected cover already exceeds factory lead (e.g. 1023YKPBLUE)."""
    parent = "1023YKPBLUE"
    sizes = ["6XL", "M", "XS", "L", "XL"]
    skus = [f"{parent}-{sz}" for sz in sizes]
    days = pd.date_range("2026-05-01", periods=30, freq="D")
    sales_frames = []
    for sku in skus:
        sales_frames.append(
            pd.DataFrame(
                {
                    "Sku": [sku] * len(days),
                    "TxnDate": days,
                    "Transaction Type": ["Shipment"] * len(days),
                    "Quantity": [1] * len(days),
                    "Units_Effective": [1] * len(days),
                    "Source": ["Amazon"] * len(days),
                }
            )
        )
    sales = pd.concat(sales_frames, ignore_index=True)
    # ADS ≈ 1/day; inv + pipeline ⇒ projected 150–175d (above 60d lead).
    inv = pd.DataFrame(
        {
            "OMS_SKU": skus,
            "Total_Inventory": [120, 110, 100, 150, 140],
        }
    )
    existing = pd.DataFrame(
        {
            "OMS_SKU": skus,
            "PO_Pipeline_Total": [50, 65, 75, 80, 70],
        }
    )
    sheet = pd.DataFrame(
        {
            "OMS_SKU": skus,
            "SKU_Sheet_Status": ["High selling"] * len(skus),
            "SKU_Sheet_Closed": [False] * len(skus),
            "Lead_Time_From_Sheet": [60.0] * len(skus),
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=60,
        target_days=180,
        demand_basis="Sold",
        safety_pct=0.0,
        existing_po_df=existing,
        sku_status_df=sheet,
        enforce_lead_time_release_gate=True,
        enforce_two_size_minimum=True,
    )
    style = po[po["OMS_SKU"].astype(str).str.startswith(parent)]
    assert len(style) == len(sizes)
    assert all(float(r["Projected_Running_Days"]) > 60 for _, r in style.iterrows())
    assert int(style["PO_Qty"].sum()) == 0


def test_unbundled_per_size_rows_get_proportional_sales_not_bundled_eff_days():
    """Individual sizes get proportional sales from bundled listing fan-out, not inherited Eff_Days.

    1917YKBLUE-L-XL has 40 net units sold (20 days × 2). L and XL each get 20 via fan-out.
    Per-size Eff_Days stays 0 for XL (no direct sales key, no inventory history for XL).
    The bundled listing row keeps its full Eff_Days from inventory history.
    """
    existing_po = pd.DataFrame(
        {
            "OMS_SKU": ["1917YKBLUE-L", "1917YKBLUE-XL", "1917YKBLUE-L-XL"],
            "PO_Pipeline_Total": [120, 100, 4],
            "Balance_to_Dispatch": [120, 100, 4],
        }
    )
    days = pd.date_range("2026-05-01", periods=20, freq="D")
    sales = pd.DataFrame(
        {
            "Sku": ["1917YKBLUE-L-XL"] * 20,
            "TxnDate": days,
            "Transaction Type": ["Shipment"] * 20,
            "Quantity": [2] * 20,
            "Units_Effective": [2] * 20,
            "Source": ["Amazon"] * 20,
        }
    )
    inv = pd.DataFrame({"OMS_SKU": ["1917YKBLUE-L-XL"], "Total_Inventory": [56]})
    inv_hist = pd.DataFrame(
        {
            "OMS_SKU": ["1917YKBLUE-L-XL"] * 25 + ["1917YKBLUE-L"] * 25,
            "Date": list(pd.date_range("2026-05-01", periods=25, freq="D")) * 2,
            "Qty": [5] * 50,
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=135,
        demand_basis="Net",
        safety_pct=0.0,
        existing_po_df=existing_po,
        inventory_history_df=inv_hist,
    )
    bundled = po.loc[po["OMS_SKU"] == "1917YKBLUE-L-XL"].iloc[0]
    assert int(bundled["Net_Units"]) > 0
    assert int(bundled["Eff_Days"]) == 30
    l_row = po.loc[po["OMS_SKU"] == "1917YKBLUE-L"].iloc[0]
    xl_row = po.loc[po["OMS_SKU"] == "1917YKBLUE-XL"].iloc[0]
    # Individual sizes receive proportional sales from bundled-listing fan-out (20 each from 40 total).
    assert int(l_row["Net_Units"]) > 0, "L should inherit proportional Net_Units from L-XL bundle"
    assert int(xl_row["Net_Units"]) > 0, "XL should inherit proportional Net_Units from L-XL bundle"
    # XL has no direct sales key or inventory history → Eff_Days stays 0 (no diluted denominator).
    assert float(xl_row["Eff_Days"]) == 0.0
    assert float(xl_row["Recent_ADS"]) == 0.0


def test_eff_days_zero_when_no_sales_despite_inventory_history():
    """Per-size rows with zero demand must not show active or extrapolated Eff_Days."""
    inv_hist = pd.DataFrame(
        {
            "OMS_SKU": ["ZERO-SALES"] * 25,
            "Date": pd.date_range("2025-12-08", periods=25, freq="D"),
            "Qty": [5] * 25,
        }
    )
    sales = pd.DataFrame(
        [
            {
                "Sku": "OTHER-SKU",
                "TxnDate": pd.Timestamp("2025-12-20"),
                "Transaction Type": "Shipment",
                "Quantity": 1,
                "Units_Effective": 1,
                "Source": "Amazon",
            }
        ]
    )
    inv = pd.DataFrame(
        {
            "OMS_SKU": ["ZERO-SALES", "OTHER-SKU"],
            "Total_Inventory": [0, 5],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=90,
        demand_basis="Net",
        safety_pct=0.0,
        group_by_parent=False,
        inventory_history_df=inv_hist,
    )
    row = po.loc[po["OMS_SKU"] == "ZERO-SALES"].iloc[0]
    assert int(row["Net_Units"]) == 0
    assert int(row["Eff_Days_Inventory"]) == 25
    assert int(row["Eff_Days"]) == 0
    assert float(row["Recent_ADS"]) == 0.0


def test_in_stock_sku_shows_inventory_eff_days_without_ads_window_sales():
    """SKUs with on-hand stock + inventory history must show Eff_Days even when ADS-window sold=0."""
    inv_hist = pd.DataFrame(
        {
            "OMS_SKU": ["STALE-SELLER"] * 25,
            "Date": pd.date_range("2026-01-03", periods=25, freq="D"),
            "Qty": [5] * 25,
        }
    )
    sales = pd.DataFrame(
        [
            {
                "Sku": "STALE-SELLER",
                "TxnDate": pd.Timestamp("2025-10-15"),
                "Transaction Type": "Shipment",
                "Quantity": 12,
                "Units_Effective": 12,
                "Source": "Amazon",
            },
            {
                "Sku": "RECENT-OTHER",
                "TxnDate": pd.Timestamp("2026-02-01"),
                "Transaction Type": "Shipment",
                "Quantity": 1,
                "Units_Effective": 1,
                "Source": "Amazon",
            },
        ]
    )
    inv = pd.DataFrame(
        {
            "OMS_SKU": ["STALE-SELLER", "RECENT-OTHER"],
            "Total_Inventory": [18, 1],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=90,
        demand_basis="Sold",
        safety_pct=0.0,
        group_by_parent=False,
        inventory_history_df=inv_hist,
    )
    row = po.loc[po["OMS_SKU"] == "STALE-SELLER"].iloc[0]
    assert int(row["Sold_Units"]) == 0
    assert int(row["Ship_Units_150d"]) == 12
    assert int(row["Eff_Days_Inventory"]) == 25
    assert int(row["Eff_Days"]) == 0
    assert float(row["Recent_ADS"]) == 0.0


def test_eff_days_zero_when_no_sales_in_ads_window_despite_ship150_span():
    """1037DPT19WHITE-style: 0 sold in 30d must not show 150d span as Eff_Days."""
    end = pd.Timestamp("2026-06-17")
    old_dates = pd.date_range(end - pd.Timedelta(days=120), periods=7, freq="7D")
    sales = pd.DataFrame(
        [
            {
                "Sku": "1037DPT19WHITE-4XL",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 1,
                "Units_Effective": 1,
                "Source": "Amazon",
            }
            for d in old_dates
        ]
        + [
            {
                "Sku": "OTHER",
                "TxnDate": end,
                "Transaction Type": "Shipment",
                "Quantity": 1,
                "Units_Effective": 1,
                "Source": "Amazon",
            }
        ]
    )
    inv = pd.DataFrame(
        {"OMS_SKU": ["1037DPT19WHITE-4XL", "OTHER"], "Total_Inventory": [4, 1]}
    )
    hist_dates = pd.date_range(end - pd.Timedelta(days=13), periods=14, freq="D")
    inv_hist = pd.DataFrame(
        {
            "OMS_SKU": ["1037DPT19WHITE-4XL"] * 14,
            "Date": hist_dates,
            "Qty": [1] * 14,
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=180,
        demand_basis="Sold",
        safety_pct=0.0,
        planning_date=str(end.date()),
        inventory_history_df=inv_hist,
    )
    row = po.loc[po["OMS_SKU"] == "1037DPT19WHITE-4XL"].iloc[0]
    assert int(row["Sold_Units"]) == 0
    assert int(row["Ship_Units_150d"]) == 7
    assert int(row["Eff_Days"]) == 0
    assert int(row["Eff_Days_Inventory"]) == 14
    assert float(row["Recent_ADS"]) == 0.0


def test_in_stock_ship150_keeps_eff_days_after_existing_po_unbundle():
    """30d-quiet SKU with stock + 150d shipments must keep Eff_Days after pipeline merge."""
    end = pd.Timestamp("2026-06-08")
    old_dates = pd.date_range(end - pd.Timedelta(days=55), periods=12, freq="D")
    sales = pd.DataFrame(
        [
            {
                "Sku": "QUIET-STOCK",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 8,
                "Units_Effective": 8,
                "Source": "Amazon",
            }
            for d in old_dates
        ]
        + [
            {
                "Sku": "OTHER",
                "TxnDate": end,
                "Transaction Type": "Shipment",
                "Quantity": 1,
                "Units_Effective": 1,
                "Source": "Amazon",
            }
        ]
    )
    inv = pd.DataFrame({"OMS_SKU": ["QUIET-STOCK", "OTHER"], "Total_Inventory": [4, 1]})
    existing_po = pd.DataFrame(
        {
            "OMS_SKU": ["QUIET-STOCK"],
            "PO_Pipeline_Total": [180],
            "Pending_Cutting": [180],
            "Balance_to_Dispatch": [0],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=40,
        target_days=135,
        demand_basis="Sold",
        existing_po_df=existing_po,
    )
    row = po.loc[po["OMS_SKU"] == "QUIET-STOCK"].iloc[0]
    assert int(row["Sold_Units"]) == 0
    assert int(row["Ship_Units_150d"]) == 96
    assert int(row["Total_Inventory"]) == 4
    assert int(row["Eff_Days"]) == 0


def test_bundled_inventory_pipeline_stays_on_band_sales_fan_out_to_per_size():
    """1917-style: dispatch stays on listing bands; individual sizes get proportional sales.

    Pipeline (Balance_to_Dispatch, Pending_Cutting) must stay on the bundled listing rows.
    Sales data fans out to individual sizes so they have ADS / historical data for PO math —
    even when both the bundled listing AND individual rows are in inventory.
    """
    existing_po = pd.DataFrame(
        {
            "OMS_SKU": [
                "1917YKBLUE-L-XL",
                "1917YKBLUE-S-M",
                "1917YKBLUE-XXL-3XL",
            ],
            "PO_Pipeline_Total": [324, 324, 194],
            "Pending_Cutting": [320, 320, 190],
            "Balance_to_Dispatch": [4, 4, 4],
        }
    )
    inv = pd.DataFrame(
        {
            "OMS_SKU": [
                "1917YKBLUE-L-XL",
                "1917YKBLUE-S-M",
                "1917YKBLUE-XXL-3XL",
                "1917YKBLUE-M",
                "1917YKBLUE-S",
                "1917YKBLUE-XL",
                "1917YKBLUE-XXL",
            ],
            "Total_Inventory": [46, 46, 46, 10, 10, 10, 10],
        }
    )
    sales = pd.DataFrame(
        [
            {
                "Sku": "1917YKBLUE-L-XL",
                "TxnDate": pd.Timestamp("2026-06-01"),
                "Transaction Type": "Shipment",
                "Quantity": 12,
                "Units_Effective": 12,
                "Source": "Amazon",
            },
            {
                "Sku": "1917YKBLUE-S-M",
                "TxnDate": pd.Timestamp("2026-06-02"),
                "Transaction Type": "Shipment",
                "Quantity": 16,
                "Units_Effective": 16,
                "Source": "Amazon",
            },
        ]
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=90,
        lead_time=45,
        target_days=135,
        demand_basis="Sold",
        existing_po_df=existing_po,
    )
    lxl = po.loc[po["OMS_SKU"] == "1917YKBLUE-L-XL"].iloc[0]
    sm = po.loc[po["OMS_SKU"] == "1917YKBLUE-S-M"].iloc[0]
    # Pipeline stays on bundled listing rows — not fanned out to individual sizes.
    assert int(lxl["Pending_Cutting"]) == 320
    assert int(sm["Pending_Cutting"]) == 320
    assert int(lxl["Balance_to_Dispatch"]) == 4
    assert int(lxl["Sold_Units"]) == 12
    m_row = po.loc[po["OMS_SKU"] == "1917YKBLUE-M"].iloc[0]
    s_row = po.loc[po["OMS_SKU"] == "1917YKBLUE-S"].iloc[0]
    # Individual sizes get proportional sales from bundled-listing fan-out (S-M=16 → M=8, S=8).
    assert int(m_row["Sold_Units"]) > 0, "M should get proportional Sold_Units from S-M bundle fan-out"
    assert int(s_row["Sold_Units"]) > 0, "S should get proportional Sold_Units from S-M bundle fan-out"
    # Pipeline must NOT flow to individual sizes (only to the bundled listing rows).
    assert int(m_row["Balance_to_Dispatch"]) == 0
    assert int(s_row["Balance_to_Dispatch"]) == 0


def test_bundled_only_sheet_expands_when_inventory_is_per_size_only():
    """Fan out pipeline to per-size only when inventory has no bundled listing row."""
    existing_po = pd.DataFrame(
        {
            "OMS_SKU": ["1917YKBLUE-L-XL"],
            "PO_Pipeline_Total": [324],
            "Pending_Cutting": [320],
            "Balance_to_Dispatch": [4],
        }
    )
    inv = pd.DataFrame(
        {"OMS_SKU": ["1917YKBLUE-L", "1917YKBLUE-XL"], "Total_Inventory": [23, 23]}
    )
    po = calculate_po_base(
        sales_df=pd.DataFrame(
            {
                "Sku": ["OTHER"],
                "TxnDate": [pd.Timestamp("2026-06-01")],
                "Transaction Type": ["Shipment"],
                "Quantity": [1],
                "Units_Effective": [1],
                "Source": ["Amazon"],
            }
        ),
        inv_df=inv,
        period_days=90,
        lead_time=45,
        target_days=135,
        demand_basis="Sold",
        existing_po_df=existing_po,
    )
    l_row = po.loc[po["OMS_SKU"] == "1917YKBLUE-L"].iloc[0]
    xl_row = po.loc[po["OMS_SKU"] == "1917YKBLUE-XL"].iloc[0]
    assert int(l_row["Pending_Cutting"]) == 160
    assert int(xl_row["Pending_Cutting"]) == 160
    assert int(l_row["Balance_to_Dispatch"]) == 2
    assert int(xl_row["Balance_to_Dispatch"]) == 2


def test_calculate_quarterly_history_fans_out_bundled_listing_sales():
    sales = pd.DataFrame(
        [
            {
                "Sku": "1917YKBLUE-L-XL",
                "TxnDate": pd.Timestamp("2026-05-15"),
                "Transaction Type": "Shipment",
                "Quantity": 10,
            },
            {
                "Sku": "1917YKBLUE-L-XL",
                "TxnDate": pd.Timestamp("2024-08-10"),
                "Transaction Type": "Shipment",
                "Quantity": 20,
            },
        ]
    )
    pivot = calculate_quarterly_history(
        sales_df=sales,
        mtr_df=None,
        sku_mapping=None,
        group_by_parent=False,
        n_quarters=8,
    )
    l = pivot.loc[pivot["OMS_SKU"] == "1917YKBLUE-L"].iloc[0]
    xl = pivot.loc[pivot["OMS_SKU"] == "1917YKBLUE-XL"].iloc[0]
    assert int(l.get("Apr-Jun 2026", 0)) == 5
    assert int(xl.get("Apr-Jun 2026", 0)) == 5
    assert int(l.get("Jul-Sep 2024", 0)) == 10
    assert int(xl.get("Jul-Sep 2024", 0)) == 10


def test_calculate_quarterly_history_fans_out_to_individual_sizes_even_when_inventory_lists_band():
    """Historical sales must always fan out to individual sizes so they get ADS / quarterly data.

    Even when inventory has a bundled listing (1917YKBLUE-L-XL), the individual size rows
    (L, XL) need proportional sales history for PO calculation.  The retain_bundled_listing_skus
    parameter is accepted for backward compatibility but no longer suppresses fan-out.
    """
    sales = pd.DataFrame(
        [
            {
                "Sku": "1917YKBLUE-L-XL",
                "TxnDate": pd.Timestamp("2026-05-15"),
                "Transaction Type": "Shipment",
                "Quantity": 10,
            },
        ]
    )
    pivot = calculate_quarterly_history(
        sales_df=sales,
        mtr_df=None,
        sku_mapping=None,
        group_by_parent=False,
        n_quarters=8,
        retain_bundled_listing_skus={"1917YKBLUE-L-XL"},  # retained for API compat, not used
    )
    band = pivot.loc[pivot["OMS_SKU"] == "1917YKBLUE-L-XL"].iloc[0]
    assert int(band.get("Apr-Jun 2026", 0)) == 10
    # Individual sizes must also appear with proportional sales (5 each from 10 band units).
    skus = set(pivot["OMS_SKU"].astype(str))
    assert "1917YKBLUE-L" in skus, "Individual L size must get historical sales from bundled listing"
    assert "1917YKBLUE-XL" in skus, "Individual XL size must get historical sales from bundled listing"
    l_row = pivot.loc[pivot["OMS_SKU"] == "1917YKBLUE-L"].iloc[0]
    xl_row = pivot.loc[pivot["OMS_SKU"] == "1917YKBLUE-XL"].iloc[0]
    assert int(l_row.get("Apr-Jun 2026", 0)) == 5
    assert int(xl_row.get("Apr-Jun 2026", 0)) == 5


def test_ship150_span_fallback_when_no_inventory_history():
    """SKUs missing from daily inventory history still show Eff_Days from 150d shipments."""
    old_dates = pd.date_range("2025-11-01", periods=10, freq="D")
    sales = pd.DataFrame(
        [
            {
                "Sku": "NO-HIST",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 2,
                "Units_Effective": 2,
                "Source": "Amazon",
            }
            for d in old_dates
        ]
        + [
            {
                "Sku": "RECENT-OTHER",
                "TxnDate": pd.Timestamp("2026-02-01"),
                "Transaction Type": "Shipment",
                "Quantity": 1,
                "Units_Effective": 1,
                "Source": "Amazon",
            }
        ]
    )
    inv = pd.DataFrame({"OMS_SKU": ["NO-HIST", "RECENT-OTHER"], "Total_Inventory": [8, 1]})
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=90,
        demand_basis="Sold",
        safety_pct=0.0,
        group_by_parent=False,
    )
    row = po.loc[po["OMS_SKU"] == "NO-HIST"].iloc[0]
    assert int(row["Sold_Units"]) == 0
    assert int(row["Ship_Units_150d"]) == 20
    assert int(row["Eff_Days"]) == 0
    """Real-world bug: daily-inventory sheet had only 24 of 30 snapshot dates.

    Every SKU in-stock for all 24 snapshot days was getting Eff_Days = 24 (the
    sheet's day count), which dropped ADS uniformly. The engine must extrapolate
    in-stock rate over the full ADS window so SKUs in stock 24/24 snapshot days
    land at Eff_Days ≈ 30, not 24.
    """
    sales_dates = pd.date_range("2025-12-18", periods=20, freq="D")
    sales = pd.DataFrame(
        [
            {
                "Sku": "INV-COV",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 2,
                "Units_Effective": 2,
                "Source": "Amazon",
            }
            for d in sales_dates
        ]
    )
    inv = pd.DataFrame({"OMS_SKU": ["INV-COV"], "Total_Inventory": [10]})

    sheet_end = pd.Timestamp("2026-01-06")
    inv_hist_partial = pd.DataFrame(
        {
            "OMS_SKU": ["INV-COV"] * 24,
            "Date": pd.date_range(end=sheet_end, periods=24, freq="D"),
            "Qty": [5] * 24,
        }
    )

    po_partial = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=90,
        demand_basis="Sold",
        safety_pct=0.0,
        group_by_parent=False,
        inventory_history_df=inv_hist_partial,
    )
    row = po_partial.iloc[0]
    assert int(row["Eff_Days_Inventory"]) == 24
    assert int(row["Inv_Coverage_Days"]) == 24
    assert int(row["Eff_Days"]) == 30, (
        f"24 in-stock / 24 covered days should extrapolate to full 30-day window; got {row['Eff_Days']}"
    )

    inv_hist_half = pd.DataFrame(
        {
            "OMS_SKU": ["INV-COV"] * 24,
            "Date": pd.date_range(end=sheet_end, periods=24, freq="D"),
            "Qty": [5] * 12 + [0] * 12,
        }
    )
    po_half = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=90,
        demand_basis="Sold",
        safety_pct=0.0,
        group_by_parent=False,
        inventory_history_df=inv_hist_half,
    )
    half = po_half.iloc[0]
    assert int(half["Eff_Days_Inventory"]) == 12
    assert int(half["Eff_Days"]) == 15, (
        f"12 in-stock / 24 covered should scale to 15 (= 50% of 30); got {half['Eff_Days']}"
    )


def test_blank_inventory_cells_are_missing_not_oos():
    """Real-world bug: warehouse skips snapshots on Sundays. Blank cells were
    being read as Qty=0 → counted as OOS days, dragging Eff_Days down by 2.

    Blanks must be treated as "no snapshot taken" and dropped, so the global
    sheet coverage (28 days when 2 Sundays are missing) drives the scaling.
    """
    import io
    from openpyxl import Workbook
    from backend.services.daily_inventory_history import (
        coverage_days_within,
        effective_days_from_history,
        parse_daily_inventory_history_upload,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "OMS"
    dates = pd.date_range("2026-04-13", "2026-05-12", freq="D")
    sundays = {pd.Timestamp("2026-05-03"), pd.Timestamp("2026-05-10")}
    # Header row 1: column-total placeholder values
    ws.cell(row=1, column=1, value="Total Inv.")
    ws.cell(row=1, column=2, value="Total")
    for i, _ in enumerate(dates, start=3):
        ws.cell(row=1, column=i, value=100)
    # Header row 2: SKU columns + dates
    ws.cell(row=2, column=1, value="Item SkuCode")
    ws.cell(row=2, column=2, value="Item")
    for i, d in enumerate(dates, start=3):
        ws.cell(row=2, column=i, value=d.to_pydatetime())
    # SKU row — non-blank every weekday, blank every Sunday in window.
    ws.cell(row=3, column=1, value="SUN-SKIP-L")
    ws.cell(row=3, column=2, value="SUN-SKIP")
    for i, d in enumerate(dates, start=3):
        if pd.Timestamp(d) in sundays:
            continue  # blank cell — no snapshot
        ws.cell(row=3, column=i, value=58)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    df = parse_daily_inventory_history_upload(buf, "blanks.xlsx")
    sub = df[df["OMS_SKU"] == "SUN-SKIP-L"]
    assert len(sub) == 28, f"Sundays must be dropped, not zero-filled (got {len(sub)})"

    start = pd.Timestamp("2026-04-13")
    end = pd.Timestamp("2026-05-12")
    assert coverage_days_within(df, start, end) == 28
    eff = effective_days_from_history(df, start, end)
    assert int(eff.iloc[0]["Eff_Days_Inventory"]) == 28

    sales = pd.DataFrame(
        [
            {
                "Sku": "SUN-SKIP-L",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 2,
                "Units_Effective": 2,
                "Source": "Amazon",
            }
            for d in pd.date_range("2026-04-13", "2026-05-12", freq="D")
        ]
    )
    inv = pd.DataFrame({"OMS_SKU": ["SUN-SKIP-L"], "Total_Inventory": [38]})
    out = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=90,
        demand_basis="Sold",
        safety_pct=0.0,
        group_by_parent=False,
        inventory_history_df=df,
    )
    row = out.iloc[0]
    assert int(row["Eff_Days_Inventory"]) == 28
    assert int(row["Inv_Coverage_Days"]) == 28
    assert int(row["Eff_Days"]) == 30, (
        f"28 in-stock / 28 covered should extrapolate to full 30-day window; "
        f"got Eff_Days={row['Eff_Days']}"
    )


def test_effective_days_counts_single_piece_as_in_stock():
    """A day with at least 1 piece on-hand counts toward Eff_Days. Only
    zero / negative on-hand is excluded."""
    from backend.services.daily_inventory_history import effective_days_from_history

    history = pd.DataFrame(
        {
            "OMS_SKU": ["SINGLEPIECE"] * 5,
            "Date": pd.date_range("2026-05-08", periods=5, freq="D"),
            "Qty": [10, 2, 1, 1, 5],
        }
    )
    out = effective_days_from_history(
        history,
        cutoff_start=pd.Timestamp("2026-05-08"),
        cutoff_end=pd.Timestamp("2026-05-12"),
    )
    # All 5 days are >= 1 → all in-stock.
    assert int(out["Eff_Days_Inventory"].iloc[0]) == 5


def test_effective_days_default_threshold_is_one():
    """Default min_qty is the IN_STOCK_MIN_QTY constant (= 1)."""
    from backend.services.daily_inventory_history import (
        IN_STOCK_MIN_QTY,
        effective_days_from_history,
    )

    assert IN_STOCK_MIN_QTY == 1.0
    history = pd.DataFrame(
        {
            "OMS_SKU": ["T"] * 3,
            "Date": pd.date_range("2026-05-10", periods=3, freq="D"),
            "Qty": [2, 1, 0],
        }
    )
    out = effective_days_from_history(
        history,
        cutoff_start=pd.Timestamp("2026-05-10"),
        cutoff_end=pd.Timestamp("2026-05-12"),
    )
    assert int(out["Eff_Days_Inventory"].iloc[0]) == 2  # day 1 (qty 2) + day 2 (qty 1)


def test_two_size_minimum_emits_actionable_recommendation():
    """When only one size in a parent has demand, block the PO and surface
    a recommendation to alter the SKU rather than a raw 'single size' tag."""
    sales = pd.DataFrame(
        {
            "Sku": ["SHIRT-L"] * 30,
            "TxnDate": pd.date_range("2026-04-01", periods=30, freq="D"),
            "Quantity": [3] * 30,
            "Units_Effective": [3] * 30,
            "Transaction Type": ["Shipment"] * 30,
            "Source": ["Amazon"] * 30,
        }
    )
    inv = pd.DataFrame(
        {
            "OMS_SKU": ["SHIRT-L", "SHIRT-XL", "SHIRT-M"],
            "Total_Inventory": [0, 9_999, 9_999],
        }
    )
    po = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=7,
        target_days=60,
        demand_basis="Sold",
        safety_pct=0.0,
        enforce_two_size_minimum=True,
    )
    row_l = po[po["OMS_SKU"] == "SHIRT-L"].iloc[0]
    assert int(row_l["PO_Qty"]) == 0
    reason = str(row_l["PO_Block_Reason"]).lower()
    assert "only 1 size" in reason
    assert "alter" in reason
    # User-friendly wording — not a developer-style "single size only" tag.
    assert "single size only" not in reason


def test_extend_history_rolls_forward_using_sales_units_effective():
    """Baseline inventory + sales activity → derived snapshots after sheet_max."""
    from backend.services.daily_inventory_history import extend_history_with_sales

    baseline = pd.DataFrame(
        {
            "OMS_SKU": ["AUTO-A", "AUTO-A", "AUTO-A", "AUTO-B"],
            "Date": pd.to_datetime(
                ["2026-05-08", "2026-05-09", "2026-05-10", "2026-05-10"]
            ),
            "Qty": [50, 48, 45, 100],
        }
    )
    sales = pd.DataFrame(
        [
            # 2026-05-11: 3 shipped of A, 0 of B
            *(
                {"Sku": "AUTO-A", "TxnDate": pd.Timestamp("2026-05-11"),
                 "Transaction Type": "Shipment", "Quantity": 1, "Units_Effective": 1}
                for _ in range(3)
            ),
            # 2026-05-12: 2 shipped of A, 5 refunded of A (net -3)
            *(
                {"Sku": "AUTO-A", "TxnDate": pd.Timestamp("2026-05-12"),
                 "Transaction Type": "Shipment", "Quantity": 1, "Units_Effective": 1}
                for _ in range(2)
            ),
            *(
                {"Sku": "AUTO-A", "TxnDate": pd.Timestamp("2026-05-12"),
                 "Transaction Type": "Refund", "Quantity": 1, "Units_Effective": -1}
                for _ in range(5)
            ),
            # 2026-05-12: 4 shipped of B
            *(
                {"Sku": "AUTO-B", "TxnDate": pd.Timestamp("2026-05-12"),
                 "Transaction Type": "Shipment", "Quantity": 1, "Units_Effective": 1}
                for _ in range(4)
            ),
        ]
    )

    out = extend_history_with_sales(baseline, sales, cap_date=pd.Timestamp("2026-05-12"))
    out["Date"] = pd.to_datetime(out["Date"])
    a = out[out["OMS_SKU"] == "AUTO-A"].sort_values("Date")
    b = out[out["OMS_SKU"] == "AUTO-B"].sort_values("Date")

    a_qty = dict(zip(a["Date"].dt.strftime("%Y-%m-%d"), a["Qty"]))
    b_qty = dict(zip(b["Date"].dt.strftime("%Y-%m-%d"), b["Qty"]))

    # Uploaded rows untouched
    assert a_qty["2026-05-10"] == 45
    assert b_qty["2026-05-10"] == 100
    # 2026-05-11: A had 3 shipped → 45 - 3 = 42; B had no activity → carry-forward 100
    assert a_qty["2026-05-11"] == 42
    assert b_qty["2026-05-11"] == 100
    # 2026-05-12: A net = 2-5 = -3 (refunds) → 42 - (-3) = 45; B 100 - 4 = 96
    assert a_qty["2026-05-12"] == 45
    assert b_qty["2026-05-12"] == 96

    # Derived rows are tagged
    derived = a[a["Date"] >= pd.Timestamp("2026-05-11")]
    assert (derived["Source"] == "derived").all()


def test_extend_history_floors_at_zero_when_over_sold():
    """Stock can't go negative — derived snapshot floors at 0."""
    from backend.services.daily_inventory_history import extend_history_with_sales

    baseline = pd.DataFrame(
        {"OMS_SKU": ["TINY"], "Date": [pd.Timestamp("2026-05-10")], "Qty": [3]}
    )
    sales = pd.DataFrame(
        [
            {"Sku": "TINY", "TxnDate": pd.Timestamp("2026-05-11"),
             "Transaction Type": "Shipment", "Quantity": 1, "Units_Effective": 1}
            for _ in range(10)
        ]
    )
    out = extend_history_with_sales(baseline, sales, cap_date=pd.Timestamp("2026-05-11"))
    qty = float(out[(out["OMS_SKU"] == "TINY") & (out["Date"] == pd.Timestamp("2026-05-11"))]["Qty"].iloc[0])
    assert qty == 0.0


def test_extend_history_no_sales_does_not_fill_calendar_gaps():
    """Without sales rows after the baseline, do not invent daily snapshots —
    that previously inflated Eff_Days when today's sales were not uploaded."""
    from backend.services.daily_inventory_history import extend_history_with_sales

    baseline = pd.DataFrame(
        {"OMS_SKU": ["IDLE"], "Date": [pd.Timestamp("2026-05-10")], "Qty": [12]}
    )
    out = extend_history_with_sales(
        baseline,
        sales_df=pd.DataFrame(columns=["Sku", "TxnDate", "Units_Effective"]),
        cap_date=pd.Timestamp("2026-05-13"),
    )
    derived = out[out.get("Source", "") == "derived"] if "Source" in out.columns else out.iloc[0:0]
    assert len(derived) == 0


def test_extend_history_skips_skus_without_baseline():
    """Sales for SKUs with no baseline snapshot are ignored — we can't infer
    a starting on-hand from sales alone."""
    from backend.services.daily_inventory_history import extend_history_with_sales

    baseline = pd.DataFrame(
        {"OMS_SKU": ["KNOWN"], "Date": [pd.Timestamp("2026-05-10")], "Qty": [10]}
    )
    sales = pd.DataFrame(
        [
            {"Sku": "UNKNOWN", "TxnDate": pd.Timestamp("2026-05-11"),
             "Transaction Type": "Shipment", "Quantity": 1, "Units_Effective": 1},
            {"Sku": "KNOWN", "TxnDate": pd.Timestamp("2026-05-11"),
             "Transaction Type": "Shipment", "Quantity": 1, "Units_Effective": 1},
        ]
    )
    out = extend_history_with_sales(baseline, sales, cap_date=pd.Timestamp("2026-05-11"))
    assert "UNKNOWN" not in set(out["OMS_SKU"].unique())
    assert int(out[(out["OMS_SKU"] == "KNOWN") & (out["Date"] == pd.Timestamp("2026-05-11"))]["Qty"].iloc[0]) == 9


def test_po_auto_extends_inventory_history_so_user_uploads_baseline_once():
    """End-to-end: baseline inventory ends a week before sales max date.

    User intent: upload Daily Inventory History once → going forward the
    engine auto-derives day-by-day inventory from sales so Eff_Days stays
    accurate without re-uploading.
    """
    # Baseline: 10-day window of full stock, ending May 5.
    baseline_dates = pd.date_range("2026-04-26", "2026-05-05", freq="D")
    inv_hist = pd.DataFrame(
        {
            "OMS_SKU": ["AUTO-EFF"] * len(baseline_dates),
            "Date": baseline_dates,
            "Qty": [40] * len(baseline_dates),
        }
    )
    # Sales: 5 ships per day from May 6 to May 12 → stock will run dry on
    # May 13 (40 - 5*8 = 0). For 30-day window ending May 12, in-stock days
    # = 10 (uploaded) + 7 (derived but still positive) = 17.
    sales_dates = pd.date_range("2026-05-06", "2026-05-12", freq="D")
    sales = pd.DataFrame(
        [
            {"Sku": "AUTO-EFF", "TxnDate": d,
             "Transaction Type": "Shipment", "Quantity": 5, "Units_Effective": 5,
             "Source": "Amazon"}
            for d in sales_dates
            for _ in range(1)  # one row per day, qty=5
        ]
    )
    # also add a non-stock-changing reference sale on the same range to be
    # the "max_date" anchor.
    inv = pd.DataFrame({"OMS_SKU": ["AUTO-EFF"], "Total_Inventory": [5]})

    out = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=90,
        demand_basis="Sold",
        safety_pct=0.0,
        group_by_parent=False,
        inventory_history_df=inv_hist,
    )
    row = out.iloc[0]
    # Coverage spans baseline (10d) + derived (7d) = 17 days within the window.
    # The other 13 days have no data (sheet started May 5 - 10d = Apr 26), so
    # they're not covered. coverage_days_within counts unique dates → 17.
    assert int(row["Inv_Coverage_Days"]) == 17
    # Stock stays > 0 throughout the 17 covered days, so in_stock = 17.
    assert int(row["Eff_Days_Inventory"]) == 17
    # Active sales span is 7 days; scaled inventory window must not dilute ADS below that.
    assert int(row["Eff_Days"]) == 7
    assert float(row["ADS"]) == 5.0


def test_po_inv_window_anchors_at_latest_data_not_stale_sales_max():
    """User intent: 'today is May 12, eff days must be calc'd for the days before May 12.'

    If sales is stale but the daily inventory sheet is fresh, the engine must
    anchor the inv-effective-days window at the *latest* date in the data
    (max(sales_max, inv_sheet_max)) — not at the older sales_max. Otherwise the
    most recent snapshot days would silently fall outside the window.
    """
    stale_sales_max = pd.Timestamp("2026-04-30")
    sales = pd.DataFrame(
        [
            {
                "Sku": "WIN-ANCHOR",
                "TxnDate": d,
                "Transaction Type": "Shipment",
                "Quantity": 2,
                "Units_Effective": 2,
                "Source": "Amazon",
            }
            for d in pd.date_range(end=stale_sales_max, periods=20, freq="D")
        ]
    )
    inv = pd.DataFrame({"OMS_SKU": ["WIN-ANCHOR"], "Total_Inventory": [10]})

    # Inventory sheet runs 12 days past sales_max: 2026-04-30 .. 2026-05-12.
    fresh_inv_max = pd.Timestamp("2026-05-12")
    inv_hist = pd.DataFrame(
        {
            "OMS_SKU": ["WIN-ANCHOR"] * 30,
            "Date": pd.date_range(end=fresh_inv_max, periods=30, freq="D"),
            "Qty": [5] * 30,
        }
    )

    out = calculate_po_base(
        sales_df=sales,
        inv_df=inv,
        period_days=30,
        lead_time=45,
        target_days=90,
        demand_basis="Sold",
        safety_pct=0.0,
        group_by_parent=False,
        inventory_history_df=inv_hist,
    )
    row = out.iloc[0]
    assert int(row["Eff_Days_Inventory"]) == 30, (
        "Sheet should anchor at 2026-05-12 (latest snapshot), counting all 30 in-stock days"
    )
    assert int(row["Inv_Coverage_Days"]) == 30
    assert int(row["Eff_Days"]) == 30


def test_raise_ledger_view_date_outside_planning_window():
    """Raises older than lookback must still show on the Raise date column."""
    from backend.services.po_raise_ledger import aggregate_raise_ledger_for_po

    ledger = pd.DataFrame(
        {
            "OMS_SKU": ["SKU-OLD"],
            "Raised_Qty": [99],
            "Raised_Date": [pd.Timestamp("2026-04-01")],
        }
    )
    agg = aggregate_raise_ledger_for_po(
        ledger,
        {},
        pd.Timestamp("2026-05-18"),
        lookback_days=14,
        raise_view_date="2026-04-01",
    )
    assert not agg.empty
    row = agg[agg["OMS_SKU"] == "SKU-OLD"].iloc[0]
    assert int(row["PO_Raised_On_View_Date"]) == 99
    assert int(row["PO_Confirmed_Raise_Pipeline"]) == 0


def test_raise_ledger_dedupes_duplicate_sku_day_rows_for_last_raised():
    """Repeated imports must not inflate last raised (e.g. 91×15 → 1365)."""
    from backend.services.po_raise_ledger import aggregate_raise_ledger_for_po

    dupes = pd.DataFrame(
        {
            "OMS_SKU": ["1037DPT19WHITE-7XL"] * 91,
            "Raised_Qty": [15] * 91,
            "Raised_Date": [pd.Timestamp("2026-05-16")] * 91,
        }
    )
    agg = aggregate_raise_ledger_for_po(
        dupes,
        {},
        pd.Timestamp("2026-05-19"),
        lookback_days=14,
    )
    row = agg[agg["OMS_SKU"] == "1037DPT19WHITE-7XL"].iloc[0]
    assert int(row["PO_Last_Raised_Qty"]) == 15
    assert int(row["PO_Confirmed_Raise_Pipeline"]) == 15


def test_raise_ledger_last_raised_ignores_parent_when_sizes_exist():
    from backend.services.po_raise_ledger import aggregate_raise_ledger_for_po

    ledger = pd.DataFrame(
        {
            "OMS_SKU": ["1037DPT19WHITE", "1037DPT19WHITE-7XL", "1037DPT19WHITE-XL"],
            "Raised_Qty": [1365, 15, 55],
            "Raised_Date": [pd.Timestamp("2026-05-16")] * 3,
        }
    )
    agg = aggregate_raise_ledger_for_po(
        ledger,
        {},
        pd.Timestamp("2026-05-19"),
        lookback_days=14,
    )
    r7 = agg[agg["OMS_SKU"] == "1037DPT19WHITE-7XL"].iloc[0]
    rx = agg[agg["OMS_SKU"] == "1037DPT19WHITE-XL"].iloc[0]
    assert int(r7["PO_Last_Raised_Qty"]) == 15
    assert int(rx["PO_Last_Raised_Qty"]) == 55
    assert "1037DPT19WHITE" not in set(agg["OMS_SKU"].astype(str))


def test_raise_ledger_last_raised_matches_saturday_fixture_import():
    from pathlib import Path

    from backend.services.po_raise_import import parse_ledger_upload_bytes
    from backend.services.po_raise_ledger import aggregate_raise_ledger_for_po

    fixture = Path(__file__).resolve().parent / "fixtures" / "po_recommendation_16-5-26.csv"
    if not fixture.is_file():
        pytest.skip("fixture missing")
    accum, err = parse_ledger_upload_bytes(fixture.read_bytes(), fixture.name)
    assert err is None
    ledger = pd.DataFrame(
        {
            "OMS_SKU": list(accum.keys()),
            "Raised_Qty": list(accum.values()),
            "Raised_Date": [pd.Timestamp("2026-05-16")] * len(accum),
        }
    )
    agg = aggregate_raise_ledger_for_po(
        ledger,
        {},
        pd.Timestamp("2026-05-19"),
        lookback_days=14,
        raise_view_date="2026-05-16",
    )
    for sku, expected in (
        ("1037DPT19WHITE-7XL", 15),
        ("1037DPT19WHITE-8XL", 25),
        ("1037DPT19WHITE-L", 20),
        ("1037DPT19WHITE-XL", 55),
    ):
        row = agg[agg["OMS_SKU"] == sku].iloc[0]
        assert int(row["PO_Last_Raised_Qty"]) == expected
        assert row["PO_Last_Raised_Date"] == "2026-05-16"
        assert int(row["PO_Raised_On_View_Date"]) == expected


def test_raise_ledger_last_and_view_date_columns():
    from backend.services.po_raise_ledger import aggregate_raise_ledger_for_po

    ledger = pd.DataFrame(
        {
            "OMS_SKU": ["SKU-A", "SKU-A", "SKU-B"],
            "Raised_Qty": [100, 50, 20],
            "Raised_Date": [
                pd.Timestamp("2026-05-14"),
                pd.Timestamp("2026-05-16"),
                pd.Timestamp("2026-05-16"),
            ],
        }
    )
    agg = aggregate_raise_ledger_for_po(
        ledger,
        {},
        pd.Timestamp("2026-05-18"),
        lookback_days=14,
        raise_view_date="2026-05-16",
    )
    a = agg[agg["OMS_SKU"] == "SKU-A"].iloc[0]
    b = agg[agg["OMS_SKU"] == "SKU-B"].iloc[0]
    assert int(a["PO_Last_Raised_Qty"]) == 50
    assert a["PO_Last_Raised_Date"] == "2026-05-16"
    assert int(a["PO_Raised_On_View_Date"]) == 50
    assert int(b["PO_Raised_On_View_Date"]) == 20
    assert int(a["PO_Raised_Yesterday"]) == 0


def test_raise_ledger_parent_total_not_applied_when_sizes_exist():
    """Parent-level raise row must not fill sizes that have no row when other sizes do."""
    from backend.services.po_engine import calculate_po_base

    sales = _minimal_sales()
    inv = pd.DataFrame(
        {
            "OMS_SKU": ["STYLE-A-S", "STYLE-A-M", "STYLE-A-3XL"],
            "Total_Inventory": [10, 10, 10],
        }
    )
    ledger = pd.DataFrame(
        {
            "OMS_SKU": ["STYLE-A", "STYLE-A-M"],
            "Raised_Qty": [2100, 350],
            "Raised_Date": [pd.Timestamp("2026-05-15")] * 2,
        }
    )
    po = calculate_po_base(
        sales,
        inv,
        30,
        30,
        90,
        safety_pct=0.0,
        planning_date="2026-05-16",
        po_raise_ledger_df=ledger,
        raise_ledger_lookback_days=14,
    )
    by_sku = po.set_index("OMS_SKU")
    assert int(by_sku.loc["STYLE-A-M", "PO_Confirmed_Raise_Pipeline"]) == 350
    assert int(by_sku.loc["STYLE-A-3XL", "PO_Confirmed_Raise_Pipeline"]) == 0
    assert int(by_sku.loc["STYLE-A-S", "PO_Confirmed_Raise_Pipeline"]) == 0


def test_raise_ledger_per_size_not_parent_broadcast():
    """Each size gets its own raised qty; parent total must not copy to every variant."""
    from backend.services.po_engine import calculate_po_base

    sales = _minimal_sales()
    inv = pd.DataFrame(
        {
            "OMS_SKU": ["STYLE-A-S", "STYLE-A-M", "STYLE-A-L"],
            "Total_Inventory": [10, 10, 10],
        }
    )
    ledger = pd.DataFrame(
        {
            "OMS_SKU": ["STYLE-A-S", "STYLE-A-M", "STYLE-A-L"],
            "Raised_Qty": [5, 15, 20],
            "Raised_Date": [pd.Timestamp("2026-05-15")] * 3,
        }
    )
    po = calculate_po_base(
        sales,
        inv,
        30,
        30,
        90,
        safety_pct=0.0,
        planning_date="2026-05-16",
        po_raise_ledger_df=ledger,
        raise_ledger_lookback_days=14,
        raise_view_date="2026-05-15",
    )
    by_sku = po.set_index("OMS_SKU")
    assert int(by_sku.loc["STYLE-A-S", "PO_Raised_On_View_Date"]) == 5
    assert int(by_sku.loc["STYLE-A-M", "PO_Raised_On_View_Date"]) == 15
    assert int(by_sku.loc["STYLE-A-L", "PO_Raised_On_View_Date"]) == 20
    assert int(by_sku.loc["STYLE-A-S", "PO_Confirmed_Raise_Pipeline"]) == 5
    assert int(by_sku.loc["STYLE-A-M", "PO_Confirmed_Raise_Pipeline"]) == 15
    assert int(by_sku.loc["STYLE-A-L", "PO_Confirmed_Raise_Pipeline"]) == 20


def test_po_raise_ledger_feeds_effective_pipeline_and_drops_repeat_po():
    """Confirmed raises (yesterday) add to effective pipeline so the next run
    does not re-recommend the same full PO for the same SKU."""
    sales = _minimal_sales()
    inv = pd.DataFrame({"OMS_SKU": ["TEST-SKU-1"], "Total_Inventory": [50]})
    planning = "2026-01-15"
    po_base = calculate_po_base(
        sales,
        inv,
        30,
        30,
        90,
        safety_pct=0.0,
        planning_date=planning,
    )
    assert not po_base.empty
    qty_first = int(po_base.iloc[0]["PO_Qty"])
    assert qty_first > 0
    assert int(po_base.iloc[0].get("PO_Confirmed_Raise_Pipeline", 0)) == 0

    ledger = pd.DataFrame(
        {
            "OMS_SKU": ["TEST-SKU-1"],
            "Raised_Qty": [qty_first],
            "Raised_Date": [pd.Timestamp("2026-01-14")],
        }
    )
    po2 = calculate_po_base(
        sales,
        inv,
        30,
        30,
        90,
        safety_pct=0.0,
        planning_date=planning,
        po_raise_ledger_df=ledger,
        raise_ledger_lookback_days=14,
    )
    row = po2.iloc[0]
    assert int(row["PO_Raised_Yesterday"]) == qty_first
    assert int(row["PO_Confirmed_Raise_Pipeline"]) == qty_first
    assert int(row["PO_Pipeline_Effective"]) == int(row["PO_Pipeline_Total"]) + qty_first
    assert int(row["PO_Qty"]) < qty_first


def test_parse_ledger_excel_export():
    import io

    from backend.services.po_raise_import import parse_ledger_dataframe, parse_ledger_upload_bytes

    df = pd.DataFrame(
        {
            "OMS_SKU": ["SKU-A", "SKU-B"],
            "PO_Qty": [10, 25],
            "Gross_PO_Qty": [10, 25],
        }
    )
    accum, err = parse_ledger_dataframe(df)
    assert err is None
    assert accum == {"SKU-A": 10, "SKU-B": 25}

    buf = io.BytesIO()
    df.to_excel(buf, index=False)
    accum2, err2 = parse_ledger_upload_bytes(buf.getvalue(), "po_recommendation.xlsx")
    assert err2 is None
    assert accum2 == accum


def test_po_raise_archive_auto_import_on_calculate_day(tmp_path, monkeypatch):
    """Archived export for a recent day is pulled into the ledger automatically."""
    import backend.services.po_raise_archive as arch

    monkeypatch.setattr(arch, "_ARCHIVE_DIR", str(tmp_path))
    arch._resolved_archive_root = None
    sess = type("S", (), {})()
    sess.sku_mapping = {}
    sess.po_raise_ledger_df = pd.DataFrame(columns=["OMS_SKU", "Raised_Qty", "Raised_Date"])
    sess._quarterly_cache = {}

    csv = "OMS_SKU,PO_Qty\nTEST-SKU-1,25\n"
    yday = pd.Timestamp("2026-05-14")
    arch.save_archive("aaaaaaaa-bbbb-cccc-dddd-eeeeeeeeeeee", yday, csv.encode())

    out = arch.try_auto_import_recent_ledgers(
        sess,
        "bbbbbbbb-cccc-dddd-eeee-ffffffffffff",
        "2026-05-15",
        group_by_parent=False,
        lookback_days=14,
    )
    assert out and out.get("ok")
    assert out.get("auto") is True
    assert int(sess.po_raise_ledger_df["Raised_Qty"].sum()) == 25

    again = arch.try_auto_import_recent_ledgers(
        sess,
        "bbbbbbbb-cccc-dddd-eeee-ffffffffffff",
        "2026-05-15",
        group_by_parent=False,
        lookback_days=14,
    )
    assert again is None


def test_summarize_raise_ledger_for_dashboard():
    from backend.services.po_raise_ledger import summarize_raise_ledger_for_dashboard

    ledger = pd.DataFrame(
        {
            "OMS_SKU": ["SKU-A", "SKU-B", "SKU-A"],
            "Raised_Qty": [10, 5, 3],
            "Raised_Date": [
                pd.Timestamp("2026-05-14"),
                pd.Timestamp("2026-05-14"),
                pd.Timestamp("2026-05-13"),
            ],
        }
    )
    out = summarize_raise_ledger_for_dashboard(
        ledger, lookback_days=30, planning_date="2026-05-15", max_skus_per_day=100
    )
    assert out["ledger_loaded"] is True
    assert out["total_units"] == 18
    assert out["total_skus"] == 2
    daily = {d["raised_date"]: d for d in out["daily_totals"]}
    assert daily["2026-05-14"]["sku_count"] == 2
    assert daily["2026-05-14"]["total_units"] == 15
    assert daily["2026-05-13"]["total_units"] == 3
    assert "2026-05-14" in out["by_day"]
    assert len(out["by_day"]["2026-05-14"]) == 2
    assert out["active_by_sku"][0]["oms_sku"] == "SKU-A"
    assert out["active_by_sku"][0]["qty"] == 13


def test_merge_po_optional_sheets_includes_raise_ledger():
    """Raise ledger must land in warm cache so new sessions inherit yesterday's raises."""
    import backend.main as main_mod
    from backend.session import AppSession

    main_mod._warm_cache = {}
    sess = AppSession()
    sess.po_raise_ledger_df = pd.DataFrame(
        {
            "OMS_SKU": ["TEST-SKU-1"],
            "Raised_Qty": [40],
            "Raised_Date": [pd.Timestamp("2026-05-14")],
        }
    )
    main_mod.merge_po_optional_sheets_into_warm_cache(sess)
    cached = main_mod._warm_cache.get("po_raise_ledger_df")
    assert cached is not None and not cached.empty
    assert int(cached.iloc[0]["Raised_Qty"]) == 40

