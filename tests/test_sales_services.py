"""Unit tests: sales aggregation, Amazon MTR dedup, platform summary."""

import pandas as pd

from backend.services.mtr import dedup_amazon_mtr_dataframe
from backend.services.helpers import map_to_oms_sku, sku_recognized_in_master
from backend.services.sales import (
    build_sales_df,
    canonical_sales_sku,
    canonical_sales_sku_series,
    filter_sales_for_export,
    get_anomalies,
    get_daily_dsr_report,
    get_platform_summary,
    get_sales_summary,
    list_sku_mapping_gaps,
)


def test_canonical_sales_sku_series_matches_scalar():
    s = pd.Series(["1023PLYKBLUE-3XL", "", None, "NAN"])
    out = canonical_sales_sku_series(s)
    assert out.iloc[0] == canonical_sales_sku("1023PLYKBLUE-3XL")
    assert out.iloc[1] == ""
    assert out.iloc[3] == ""


def test_get_daily_dsr_report_segments_and_others():
    df = pd.DataFrame(
        {
            "TxnDate": pd.to_datetime(["2025-04-09"] * 4),
            "Transaction Type": ["Shipment", "Shipment", "Refund", "Shipment"],
            "Quantity": [10, 5, 2, 3],
            "Source": ["Amazon", "Amazon", "Amazon", "Shopify"],
            "DSR_Segment": ["YG", "Akiko", "YG", "All"],
            "OrderId": ["a", "b", "c", "d"],
            "Sku": ["x", "y", "x", "z"],
            "Units_Effective": [10, 5, -2, 3],
        }
    )
    r = get_daily_dsr_report(df, "2025-04-09")
    assert r["subtotal"]["sales"] == 18
    assert r["subtotal"]["returns"] == 2
    amazon = next(s for s in r["sections"] if s["platform"] == "Amazon")
    assert amazon["section_sales"] == 15
    assert amazon["section_returns"] == 2
    others = next(s for s in r["sections"] if s["platform"] == "Others")
    assert others["rows"][0]["sales"] == 3
    assert others["rows"][0]["segment"] == "Shopify"


def test_get_sales_summary_includes_ist_same_calendar_day():
    """tz-aware +05:30 timestamps must not fall out of an IST single-day window."""
    df = pd.DataFrame(
        {
            "TxnDate": pd.to_datetime(["2026-04-08 00:30:00+05:30"]),
            "Transaction Type": ["Shipment"],
            "Quantity": [5],
            "Units_Effective": [5],
            "Sku": ["X"],
            "Source": ["Myntra"],
            "OrderId": ["1"],
        }
    )
    s = get_sales_summary(df, months=0, start_date="2026-04-08", end_date="2026-04-08")
    assert s["total_units"] == 5


def test_get_sales_summary_shipments_and_refunds():
    df = pd.DataFrame(
        {
            "TxnDate": pd.to_datetime(["2025-06-01", "2025-06-02", "2025-06-03"]),
            "Transaction Type": ["Shipment", "Shipment", "Refund"],
            "Quantity": [10, 5, 2],
            "Sku": ["A", "B", "A"],
            "Source": ["Amazon", "Amazon", "Amazon"],
            "Units_Effective": [10, 5, -2],
            "OrderId": [None, None, None],
        }
    )
    s = get_sales_summary(df, months=0)
    assert s["total_units"] == 15
    assert s["total_returns"] == 2
    assert s["net_units"] == 13
    assert abs(s["return_rate"] - (2 / 15 * 100)) < 0.15


def test_get_anomalies_return_spike_uses_unified_sales_and_date_window():
    """Raw platform frames can imply a huge return rate; unified + date filter must match cards."""
    raw_fk = pd.DataFrame(
        {
            "Date": pd.to_datetime(["2020-01-01", "2020-01-02"]),
            "TxnType": ["Shipment", "Refund"],
            "Quantity": [10.0, 50.0],
        }
    )
    sales = pd.DataFrame(
        {
            "TxnDate": pd.to_datetime(["2026-04-10", "2026-04-10"]),
            "Transaction Type": ["Shipment", "Refund"],
            "Quantity": [100, 3],
            "Source": ["Flipkart", "Flipkart"],
            "Sku": ["A", "A"],
            "Units_Effective": [100, -3],
            "OrderId": ["x", "y"],
        }
    )
    alerts = get_anomalies(
        pd.DataFrame(),
        pd.DataFrame(),
        pd.DataFrame(),
        raw_fk,
        snapdeal_df=pd.DataFrame(),
        inventory_df=pd.DataFrame(),
        sales_df=sales,
        start_date="2026-04-10",
        end_date="2026-04-10",
    )
    spike = [a for a in alerts if a.get("type") == "return_spike"]
    assert spike == []


def test_dedup_keeps_refund_sharing_order_with_invoice_shipment():
    d = pd.DataFrame(
        {
            "Invoice_Number": ["INV1", ""],
            "Order_Id": ["O1", "O1"],
            "SKU": ["SKU1", "SKU1"],
            "Transaction_Type": ["Shipment", "Refund"],
            "Quantity": [3.0, 1.0],
            "Date": pd.to_datetime(["2025-01-15", "2025-01-20"]),
        }
    )
    out = dedup_amazon_mtr_dataframe(d)
    assert len(out) == 2
    assert int(out[out["Transaction_Type"] == "Refund"]["Quantity"].sum()) == 1


def test_build_sales_df_amazon_with_mapping():
    mtr = pd.DataFrame(
        {
            "Date": pd.to_datetime(["2025-04-01"]),
            "SKU": ["1001YK-RED-L"],
            "Transaction_Type": ["Shipment"],
            "Quantity": [4.0],
            "Order_Id": ["OID-1"],
            "Invoice_Number": [""],
        }
    )
    mapping = {"1001YK-RED-L": "1001YK-RED-L"}
    merged = build_sales_df(
        mtr_df=mtr,
        myntra_df=pd.DataFrame(),
        meesho_df=pd.DataFrame(),
        flipkart_df=pd.DataFrame(),
        sku_mapping=mapping,
    )
    assert not merged.empty
    assert (merged["Source"] == "Amazon").all()
    assert merged["Quantity"].sum() == 4


def test_build_sales_df_amazon_with_empty_mapping_includes_rows():
    """Without a SKU map file, Amazon MTR must still land in unified sales (PL-normalised SKU)."""
    mtr = pd.DataFrame(
        {
            "Date": pd.to_datetime(["2025-04-01"]),
            "SKU": ["1023PLYKBLUE-3XL"],
            "Transaction_Type": ["Shipment"],
            "Quantity": [2.0],
            "Order_Id": ["O-99"],
            "Invoice_Number": [""],
        }
    )
    merged = build_sales_df(
        mtr_df=mtr,
        myntra_df=pd.DataFrame(),
        meesho_df=pd.DataFrame(),
        flipkart_df=pd.DataFrame(),
        sku_mapping={},
    )
    assert not merged.empty
    assert (merged["Source"] == "Amazon").all()
    assert merged["Quantity"].sum() == 2
    assert "1023YKBLUE-3XL" in merged["Sku"].astype(str).values or merged["Sku"].astype(str).str.contains("1023").any()


def test_filter_sales_for_export_date_and_source():
    df = pd.DataFrame(
        {
            "TxnDate": pd.to_datetime(["2025-06-01", "2025-06-15", "2025-07-01"]),
            "Transaction Type": ["Shipment", "Refund", "Shipment"],
            "Quantity": [10, 2, 5],
            "Sku": ["A", "A", "B"],
            "Source": ["Amazon", "Amazon", "Myntra"],
            "Units_Effective": [10, -2, 5],
        }
    )
    out = filter_sales_for_export(
        df, start_date="2025-06-01", end_date="2025-06-30", sources=["Amazon"]
    )
    assert len(out) == 2
    assert set(out["Source"].unique()) == {"Amazon"}


def test_sku_recognized_in_master_keys_and_values():
    m = {"AMZ1": "OMSVAL", "FK": "OTHER"}
    assert sku_recognized_in_master("AMZ1", m)
    assert sku_recognized_in_master("OMSVAL", m)
    assert not sku_recognized_in_master("STRANGER", m)


def test_sku_recognized_oms_value_pl_yk_mirror():
    """Master OMS cell 1012PLYK* must match sales token 1012YK* in gap detection."""
    m = {"MKEY": "1012PLYKCPINK-3XL"}
    assert sku_recognized_in_master("1012YKCPINK-3XL", m)


def test_sku_recognized_hyphen_spacing_and_dot_before_size():
    m = {"1057KDBLUE-7-8": "OMS-A", "1556YKNGREEN-3XL": "OMS-B"}
    assert sku_recognized_in_master("1057KDBLUE -7-8", m)
    assert sku_recognized_in_master("1556YKNGREEN.-3XL", m)


def test_map_glued_myntra_size_hyphens():
    m = {"1061YKBLUE-4XL-BLUE": "OMS-G"}
    assert map_to_oms_sku("1061YKBLUE4XLBLUE", m) == "OMS-G"
    assert sku_recognized_in_master("1061YKBLUE4XLBLUE", m)


def test_map_glued_multi_trimmed_variant():
    m = {"1378YKMULTI-3XL": "OMS-M"}
    assert map_to_oms_sku("1378YKMULTI3XLMULTI", m) == "OMS-M"
    assert sku_recognized_in_master("1378YKMULTI3XLMULTI", m)


def test_mustrad_maps_to_mustard_master():
    m = {"165YK-251MUSTARD-3XL": "OMS-T"}
    assert map_to_oms_sku("165YK-251MUSTRAD-3XL", m) == "OMS-T"


def test_map_hyphen_between_style_id_and_yk_block():
    """Unified / channel exports sometimes insert 165-YK251… vs 165YK251…"""
    m = {"165YK251MUSTRAD": "OMS-165"}
    assert map_to_oms_sku("165-YK251MUSTRAD", m) == "OMS-165"
    assert map_to_oms_sku("165YK251MUSTRAD", m) == "OMS-165"


def test_map_y_hyphen_k_digit_listing_typo():
    m = {"165YK251MUSTRAD": "OMS-165"}
    assert map_to_oms_sku("165Y-K251MUSTRAD", m) == "OMS-165"


def test_map_plak_to_plyk_then_pl_strip():
    m = {"165YK251MUSTRAD": "OMS-165"}
    assert map_to_oms_sku("165PLAK251MUSTRAD", m) == "OMS-165"


def test_bare_digits_match_embedded_yrn_key():
    m = {"YARYKASS100506552": "OMS-YRN"}
    assert map_to_oms_sku("100506552", m) == "OMS-YRN"
    assert sku_recognized_in_master("100506552", m)


def test_myntra_prefixed_vendor_code_maps_via_embedded_style_id():
    """PPMP / exports often use DSBY… / vendor-prefixed SKUs; OMS master keys carry the numeric tail."""
    m = {"STYLE131185528": "5009YKNGREY-8XL"}
    assert map_to_oms_sku("DSBYDRSS131185528", m) == "5009YKNGREY-8XL"
    assert sku_recognized_in_master("DSBYDRSS131185528", m)


def test_yrn_decimal_form_matches_integer_map_key():
    """Myntra YRN in Excel is often 100672680.0 in sales while the map key is 100672680."""
    m = {"100672680": "1001YK-XL"}
    assert sku_recognized_in_master("100672680.0", m)
    assert sku_recognized_in_master("100672680", m)
    gaps = list_sku_mapping_gaps(
        pd.DataFrame({"Sku": ["100672680.0", "999999999"]}),
        m,
        limit=20,
    )
    assert "100672680.0" not in gaps and "100672680" not in gaps
    assert "999999999" in gaps


def test_list_sku_mapping_gaps():
    sales = pd.DataFrame({"Sku": ["OMSVAL", "STRANGER", "MEESHO_TOTAL", ""]})
    m = {"AMZ1": "OMSVAL", "FK99": "OTHER"}
    gaps = list_sku_mapping_gaps(sales, m)
    assert "STRANGER" in gaps
    assert "OMSVAL" not in gaps
    assert "OTHER" not in gaps
    assert "MEESHO_TOTAL" not in gaps


def test_meesho_detects_listing_sku_header_and_combines_size():
    from backend.services.meesho import _combine_meesho_sku_size, _meesho_sku_base_series

    df = pd.DataFrame({"listing sku": ["1158YKGREEN"], "size": ["XL"]})
    base = _meesho_sku_base_series(df)
    assert base.iloc[0] == "1158YKGREEN"
    assert _combine_meesho_sku_size(base, df["size"]).iloc[0] == "1158YKGREEN-XL"


def test_meesho_kids_size_band_hyphen():
    from backend.services.meesho import _combine_meesho_sku_size

    base = pd.Series(["1158YKMUSTARD"])
    size = pd.Series(["7-8"])
    assert _combine_meesho_sku_size(base, size).iloc[0] == "1158YKMUSTARD-7-8"


def test_meesho_order_export_xlsx_coalesces_listing_sku_column():
    from io import BytesIO

    import pandas as pd

    from backend.services.meesho import parse_meesho_order_export_xlsx

    df = pd.DataFrame(
        {
            "TxnDate": [pd.Timestamp("2024-12-16")],
            "Sku": ["MEESHO_TOTAL"],
            "Listing Sku": ["1592YKBLUE-5XL"],
            "Transaction Type": ["Shipment"],
            "Quantity": [1],
            "Source": ["Meesho"],
            "OrderId": ["100056595506869312_1"],
        }
    )
    buf = BytesIO()
    df.to_excel(buf, index=False)
    out, msg = parse_meesho_order_export_xlsx(buf.getvalue())
    assert msg == "OK"
    assert out["SKU"].iloc[0] == "1592YKBLUE-5XL"


def test_meesho_order_export_xlsx_coalesces_sku1():
    from io import BytesIO

    import pandas as pd

    from backend.services.meesho import parse_meesho_order_export_xlsx

    df = pd.DataFrame(
        {
            "TxnDate": [pd.Timestamp("2024-12-16")],
            "Sku": ["MEESHO_TOTAL"],
            "OMS_Sku": [None],
            "Transaction Type": ["Shipment"],
            "Quantity": [1],
            "Units_Effective": [1],
            "Source": ["Meesho"],
            "OrderId": ["100056595506869312_1"],
            "Sku.1": ["1592YKBLUE-5XL"],
        }
    )
    buf = BytesIO()
    df.to_excel(buf, index=False)
    out, msg = parse_meesho_order_export_xlsx(buf.getvalue())
    assert msg == "OK"
    assert len(out) == 1
    assert out["SKU"].iloc[0] == "1592YKBLUE-5XL"


def test_meesho_csv_txn_types_ship_refund_cancel():
    """Returns/RTO are Refund; cancelled lines are Cancel (excluded from shipped KPI)."""
    from backend.services.meesho import parse_meesho_csv

    csv = (
        "order date,reason for credit entry,sku,quantity,sub order no\n"
        "2026-04-14,SHIPPED,A1,1,S1\n"
        "2026-04-14,READY_TO_SHIP,A2,1,S2\n"
        "2026-04-14,CANCELLED,A3,1,S3\n"
        "2026-04-14,RETURNED,A4,1,S4\n"
    )
    out, msg = parse_meesho_csv(csv.encode("utf-8"))
    assert msg == "OK"
    ship = out[out["TxnType"] == "Shipment"]["Quantity"].sum()
    ref = out[out["TxnType"] == "Refund"]["Quantity"].sum()
    can = out[out["TxnType"] == "Cancel"]["Quantity"].sum()
    assert float(ship) == 2.0
    assert float(ref) == 1.0
    assert float(can) == 1.0


def test_myntra_csv_maps_status_f_to_shipment():
    from backend.services.myntra import _parse_myntra_csv

    csv = (
        "created on,order status,seller sku code,myntra sku code\n"
        "2026-04-14 10:00:00,F,SK1,Y1\n"
        "2026-04-14 11:00:00,IC,SK2,Y2\n"
    )
    df, msg = _parse_myntra_csv(csv.encode("utf-8"), "t.csv", {})
    assert "OK" in msg
    assert df["TxnType"].tolist() == ["Shipment", "Cancel"]


def test_myntra_csv_prefers_dispatch_date_when_present():
    from backend.services.myntra import _parse_myntra_csv

    csv = (
        "created on,dispatch_date,order status,order line id,seller sku code,myntra sku code\n"
        "2026-03-29,2026-03-31,SHIPPED,L1,SK1,Y1\n"
    )
    df, msg = _parse_myntra_csv(csv.encode("utf-8"), "t.csv", {})
    assert "OK" in msg
    assert pd.Timestamp(df["Date"].iloc[0]).normalize() == pd.Timestamp("2026-03-31").normalize()


def test_myntra_csv_prefers_order_line_id_over_store_order_id():
    """Seller report column order puts store order id before line id; line id must win for dedup."""
    from backend.services.myntra import _parse_myntra_csv

    csv = (
        "created on,order status,store order id,order line id,seller sku code,myntra sku code\n"
        "2026-04-10 10:00:00,WP,PARENT1,LINEAA,SK1,Y1\n"
        "2026-04-10 10:00:00,WP,PARENT1,LINEBB,SK1,Y2\n"
    )
    df, msg = _parse_myntra_csv(csv.encode("utf-8"), "seller.csv", {})
    assert "OK" in msg
    assert df["OrderId"].tolist() == ["LINEAA", "LINEBB"]
    assert df["LineKey"].tolist() == ["LINEAA", "LINEBB"]


def test_myntra_merge_platform_data_idempotent_on_reupload():
    """Re-merging the same parsed frame must not double row counts (Tier-3 / cache paths)."""
    from backend.services.daily_store import merge_platform_data
    from backend.services.myntra import _parse_myntra_csv

    csv = (
        "created on,order status,store order id,order line id,seller sku code,myntra sku code\n"
        "2026-04-10 10:00:00,WP,PARENT1,L1,SK1,Y1\n"
        "2026-04-10 10:00:00,WP,PARENT1,L2,SK2,Y2\n"
    )
    df, _ = _parse_myntra_csv(csv.encode("utf-8"), "a.csv", {})
    m1 = merge_platform_data(pd.DataFrame(), df, "myntra")
    m2 = merge_platform_data(m1, df, "myntra")
    assert len(m1) == 2
    assert len(m2) == 2


def test_myntra_legacy_shadow_drops_parent_id_when_linekey_twin_exists():
    from backend.services.daily_store import _dedup_linekey_legacy_shadow
    import pandas as pd

    d = pd.DataFrame(
        {
            "Date": pd.to_datetime(["2026-04-10", "2026-04-10"]),
            "OMS_SKU": ["S1", "S1"],
            "TxnType": ["Shipment", "Shipment"],
            "Quantity": [1.0, 1.0],
            "RawStatus": ["WP", "WP"],
            "LineKey": ["", "L99"],
            "OrderId": ["PARENT", "L99"],
        }
    )
    out = _dedup_linekey_legacy_shadow(d)
    assert len(out) == 1
    assert out["OrderId"].iloc[0] == "L99"


def test_meesho_overlay_drops_synthetic_when_suborder_twin_same_fingerprint():
    from backend.services.daily_store import _dedup_meesho_cross_source_overlay
    import pandas as pd

    d = pd.DataFrame(
        {
            "Date": pd.to_datetime(["2024-04-10", "2024-04-10"]),
            "OMS_SKU": ["SKU1", "SKU1"],
            "TxnType": ["Shipment", "Shipment"],
            "Quantity": [2.0, 2.0],
            "LineKey": ["MEEEXP|20240410|SKU1|Shipment|2", "SUBORDER99"],
            "OrderId": ["MEEEXP|20240410|SKU1|Shipment|2", "SUBORDER99"],
            "MeeshoSubOrder": ["SUBORDER99", "SUBORDER99"],
        }
    )
    out = _dedup_meesho_cross_source_overlay(d)
    assert len(out) == 1
    assert out["LineKey"].iloc[0] == "SUBORDER99"


def test_meesho_overlay_prefers_tcs_over_export_when_all_synthetic():
    from backend.services.daily_store import _dedup_meesho_cross_source_overlay
    import pandas as pd

    d = pd.DataFrame(
        {
            "Date": pd.to_datetime(["2024-04-10", "2024-04-10"]),
            "OMS_SKU": ["SKU1", "SKU1"],
            "TxnType": ["Shipment", "Shipment"],
            "Quantity": [1.0, 1.0],
            "LineKey": ["MEEEXP|20240410|SKU1|Shipment|1", "MEETCS|20240410|SKU1|Shipment|1|199"],
            "OrderId": ["MEEEXP|20240410|SKU1|Shipment|1", "MEETCS|20240410|SKU1|Shipment|1|199"],
        }
    )
    out = _dedup_meesho_cross_source_overlay(d)
    assert len(out) == 1
    assert str(out["LineKey"].iloc[0]).startswith("MEETCS|")


def test_flipkart_overlay_drops_fkem_when_real_order_same_fingerprint():
    from backend.services.daily_store import _dedup_flipkart_cross_source_overlay
    import pandas as pd

    d = pd.DataFrame(
        {
            "Date": pd.to_datetime(["2024-04-10", "2024-04-10"]),
            "OMS_SKU": ["OMS1", "OMS1"],
            "TxnType": ["Shipment", "Shipment"],
            "Quantity": [3.0, 3.0],
            "LineKey": ["FKEM||LIST1|20240410|SHIP|3", "OD123456789"],
            "OrderId": ["FKEM||LIST1|20240410|SHIP|3", "OD123456789"],
        }
    )
    out = _dedup_flipkart_cross_source_overlay(d)
    assert len(out) == 1
    assert out["LineKey"].iloc[0] == "OD123456789"


def test_flipkart_overlay_prefers_order_export_over_earn_more_when_all_synthetic():
    from backend.services.daily_store import _dedup_flipkart_cross_source_overlay
    import pandas as pd

    d = pd.DataFrame(
        {
            "Date": pd.to_datetime(["2024-04-10", "2024-04-10"]),
            "OMS_SKU": ["OMS1", "OMS1"],
            "TxnType": ["Shipment", "Shipment"],
            "Quantity": [2.0, 2.0],
            "LineKey": ["FKEM||LIST1|20240410|SHIP|2", "PID1_LIST1_20240410"],
            "OrderId": ["FKEM||LIST1|20240410|SHIP|2", "PID1_LIST1_20240410"],
        }
    )
    out = _dedup_flipkart_cross_source_overlay(d)
    assert len(out) == 1
    assert "FKEM" not in str(out["LineKey"].iloc[0])


def test_merge_platform_data_runs_dedup_on_first_upload():
    """First merge used to skip _dedup_platform_df — overlays must run on single batch too."""
    from backend.services.daily_store import merge_platform_data
    import pandas as pd

    d = pd.DataFrame(
        {
            "Date": pd.to_datetime(["2024-04-10", "2024-04-10"]),
            "OMS_SKU": ["SKU1", "SKU1"],
            "TxnType": ["Shipment", "Shipment"],
            "Quantity": [2.0, 2.0],
            "LineKey": ["MEEEXP|20240410|SKU1|Shipment|2", "SUBORDER99"],
            "OrderId": ["MEEEXP|20240410|SKU1|Shipment|2", "SUBORDER99"],
            "MeeshoSubOrder": ["SUBORDER99", "SUBORDER99"],
        }
    )
    out = merge_platform_data(pd.DataFrame(), d, "meesho")
    assert len(out) == 1


def test_meesho_suborder_collapses_tcs_style_and_packet_linekey():
    """Same sub-order: TCS synthetic LineKey vs packet LineKey → one row."""
    from backend.services.daily_store import merge_platform_data
    import pandas as pd

    d = pd.DataFrame(
        {
            "Date": pd.to_datetime(["2026-04-10", "2026-04-10"]),
            "OMS_SKU": ["S1", "S1"],
            "TxnType": ["Shipment", "Shipment"],
            "Quantity": [1.0, 1.0],
            "LineKey": ["MEETCS|20260410|S1|Shipment|1|99", "PKT888"],
            "OrderId": ["MEETCS|20260410|S1|Shipment|1|99", "PKT888"],
            "MeeshoSubOrder": ["SUB777", "SUB777"],
            "RawStatus": ["Shipment", "DELIVERED"],
            "Invoice_Amount": [99.0, 99.0],
        }
    )
    out = merge_platform_data(pd.DataFrame(), d, "meesho")
    assert len(out) == 1
    assert out["LineKey"].iloc[0] == "PKT888"


def test_myntra_parent_order_shadow_drops_store_id_duplicate():
    from backend.services.daily_store import merge_platform_data
    import pandas as pd

    d = pd.DataFrame(
        {
            "Date": pd.to_datetime(["2026-04-10", "2026-04-10"]),
            "OMS_SKU": ["S1", "S1"],
            "TxnType": ["Shipment", "Shipment"],
            "Quantity": [1.0, 1.0],
            "LineKey": ["LINE99", "PARENT1"],
            "OrderId": ["LINE99", "PARENT1"],
            "ParentOrderId": ["PARENT1", "PARENT1"],
            "RawStatus": ["WP", "WP"],
        }
    )
    out = merge_platform_data(pd.DataFrame(), d, "myntra")
    assert len(out) == 1
    assert out["LineKey"].iloc[0] == "LINE99"


def test_myntra_parent_shadow_keeps_unrelated_orders_same_fingerprint():
    """Two different store orders, same SKU/qty/day — must not drop the parent-only row."""
    from backend.services.daily_store import merge_platform_data
    import pandas as pd

    d = pd.DataFrame(
        {
            "Date": pd.to_datetime(["2026-03-31", "2026-03-31"]),
            "OMS_SKU": ["S1", "S1"],
            "TxnType": ["Shipment", "Shipment"],
            "Quantity": [1.0, 1.0],
            "LineKey": ["LINE_A", "PARENT_B"],
            "OrderId": ["LINE_A", "PARENT_B"],
            "ParentOrderId": ["PARENT_A", "PARENT_B"],
            "RawStatus": ["WP", "WP"],
        }
    )
    out = merge_platform_data(pd.DataFrame(), d, "myntra")
    assert len(out) == 2


def test_flipkart_strong_dedup_keeps_two_skus_same_order_id():
    from backend.services.daily_store import merge_platform_data
    import pandas as pd

    d = pd.DataFrame(
        {
            "Date": pd.to_datetime(["2026-04-10", "2026-04-10"]),
            "OMS_SKU": ["SKU-A", "SKU-B"],
            "TxnType": ["Shipment", "Shipment"],
            "Quantity": [1.0, 2.0],
            "LineKey": ["ODSHARED", "ODSHARED"],
            "OrderId": ["ODSHARED", "ODSHARED"],
            "RawStatus": ["Sale", "Sale"],
        }
    )
    out = merge_platform_data(pd.DataFrame(), d, "flipkart")
    assert len(out) == 2


def test_build_sales_dedupes_duplicate_linekey_after_mapping():
    from backend.services.sales import build_sales_df
    import pandas as pd

    myn = pd.DataFrame(
        {
            "Date": pd.to_datetime(["2026-04-10", "2026-04-10"]),
            "OMS_SKU": ["LIST1", "LIST1"],
            "TxnType": ["Shipment", "Shipment"],
            "Quantity": [1.0, 1.0],
            "LineKey": ["LINE1", "LINE1"],
            "OrderId": ["LINE1", "LINE1"],
        }
    )
    sales = build_sales_df(
        pd.DataFrame(),
        myn,
        pd.DataFrame(),
        pd.DataFrame(),
        {"LIST1": "OMS1"},
    )
    m = (sales["Source"] == "Myntra") & (sales["Transaction Type"] == "Shipment")
    assert int(pd.to_numeric(sales.loc[m, "Quantity"], errors="coerce").sum()) == 1


def test_flipkart_earn_more_assigns_distinct_order_ids_for_dedup():
    """Same SKU/date/qty lines must not collapse in build_sales_df."""
    from io import BytesIO

    from openpyxl import Workbook

    from backend.services.flipkart import _parse_flipkart_earn_more
    from backend.services.sales import build_sales_df

    wb = Workbook()
    ws = wb.active
    ws.title = "earn_more_report"
    hdr = [
        "Order Date",
        "SKU ID",
        "Gross Units",
        "Final Sale Units",
        "Return Units",
        "Cancellation Units",
        "Final Sale Amount",
        "Return Amount",
        "Brand",
    ]
    ws.append(hdr)
    ws.append(["2026-04-10", "LIST1", 5, 3, 0, 0, 300, 0, "B1"])
    ws.append(["2026-04-10", "LIST1", 6, 4, 0, 0, 400, 0, "B1"])
    buf = BytesIO()
    wb.save(buf)
    fk = _parse_flipkart_earn_more(buf.getvalue(), "em.xlsx", {"LIST1": "OMS-1"})
    sales = build_sales_df(
        pd.DataFrame(),
        pd.DataFrame(),
        pd.DataFrame(),
        fk,
        {},
    )
    m = (sales["Source"] == "Flipkart") & (sales["Transaction Type"] == "Shipment")
    assert int(pd.to_numeric(sales.loc[m, "Quantity"], errors="coerce").sum()) == 7


def test_meesho_csv_skips_aggregate_sku_column_for_listing_sku():
    from backend.services.meesho import parse_meesho_csv

    csv = (
        "order date,sku,listing sku,quantity,customer state,sub order no\n"
        "2024-01-01,MEESHO_TOTAL,1158YKGREEN-XL,1,MH,SO1\n"
    )
    out, msg = parse_meesho_csv(csv.encode("utf-8"))
    assert msg == "OK"
    assert out["SKU"].iloc[0] == "1158YKGREEN-XL"


def test_meesho_nested_zip_finds_tcs_by_basename():
    """TCS files in subfolders (SomeFolder/tcs_sales_return.xlsx) must be detected."""
    import zipfile
    from io import BytesIO

    from backend.services.meesho import load_meesho_from_zip, meesho_to_sales_rows

    inner_df = pd.DataFrame(
        {
            "order_date": [pd.Timestamp("2025-12-31")],
            "sku": ["1158YKGREEN"],
            "size": ["XL"],
            "quantity": [1],
            "total_invoice_value": [100.0],
            "end_customer_state_new": ["MH"],
            "sub_order_num": ["SO1"],
        }
    )
    xb = BytesIO()
    inner_df.to_excel(xb, index=False)
    xlsx_bytes = xb.getvalue()

    inner_mem = BytesIO()
    with zipfile.ZipFile(inner_mem, "w", zipfile.ZIP_DEFLATED) as zin:
        zin.writestr("Reports/Dec/tcs_sales_return.xlsx", xlsx_bytes)

    outer_mem = BytesIO()
    with zipfile.ZipFile(outer_mem, "w", zipfile.ZIP_DEFLATED) as zout:
        zout.writestr("dec.zip", inner_mem.getvalue())

    combined, _n, _skip = load_meesho_from_zip(outer_mem.getvalue())
    assert not combined.empty
    assert combined["SKU"].iloc[0] == "1158YKGREEN-XL"
    sales = meesho_to_sales_rows(combined, None)
    assert sales["Sku"].iloc[0] != "MEESHO_TOTAL"


def test_meesho_flat_outer_zip_without_nested_zip():
    """Outer archive may contain TCS xlsx directly (no inner .zip)."""
    import zipfile
    from io import BytesIO

    from backend.services.meesho import load_meesho_from_zip, meesho_to_sales_rows

    inner_df = pd.DataFrame(
        {
            "order_date": [pd.Timestamp("2025-11-15")],
            "sku": ["999YKRED"],
            "size": ["M"],
            "quantity": [2],
            "total_invoice_value": [50.0],
            "end_customer_state_new": ["KA"],
            "sub_order_num": ["SO2"],
        }
    )
    xb = BytesIO()
    inner_df.to_excel(xb, index=False)
    outer_mem = BytesIO()
    with zipfile.ZipFile(outer_mem, "w", zipfile.ZIP_DEFLATED) as zout:
        zout.writestr("whatever/tcs_sales.xlsx", xb.getvalue())

    combined, _n, _sk = load_meesho_from_zip(outer_mem.getvalue())
    assert not combined.empty
    assert combined["SKU"].iloc[0] == "999YKRED-M"
    assert meesho_to_sales_rows(combined, None)["Sku"].iloc[0] != "MEESHO_TOTAL"


def test_meesho_to_sales_maps_and_oms_refresh_on_build():
    from backend.services.meesho import meesho_to_sales_rows

    mdf = pd.DataFrame(
        {
            "Date": pd.to_datetime(["2025-12-31"]),
            "TxnType": ["Refund"],
            "Quantity": [1.0],
            "SKU": ["1158YKGREEN-XL"],
            "OrderId": ["237653394323428672_1"],
        }
    )
    assert meesho_to_sales_rows(mdf, {"1158YKGREEN-XL": "1001OMS-XL"})["Sku"].iloc[0] == "1001OMS-XL"
    mdf2 = mdf.copy()
    m = {"1158YKGREEN-XL": "1001OMS-XL"}
    build_sales_df(
        mtr_df=pd.DataFrame(),
        myntra_df=pd.DataFrame(),
        meesho_df=mdf2,
        flipkart_df=pd.DataFrame(),
        sku_mapping=m,
    )
    assert mdf2["OMS_SKU"].iloc[0] == "1001OMS-XL"


def test_get_platform_summary_amazon_refunds():
    mtr = pd.DataFrame(
        {
            "Date": pd.to_datetime(["2025-04-01", "2025-04-02"]),
            "SKU": ["X", "X"],
            "Transaction_Type": ["Shipment", "Refund"],
            "Quantity": [100.0, 5.0],
        }
    )
    plat = get_platform_summary(mtr, pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), None)
    amz = next(p for p in plat if p["platform"] == "Amazon")
    assert amz["loaded"] is True
    assert amz["total_units"] == 100
    assert amz["total_returns"] == 5
    assert amz["return_rate"] == 5.0


def test_platform_summary_with_sales_df_matches_export_window():
    """Dashboard cards must use unified sales (same as CSV export), not raw platform frames."""
    mtr = pd.DataFrame(
        {
            "Date": pd.to_datetime(["2025-04-10", "2025-04-11"]),
            "SKU": ["A", "B"],
            "Transaction_Type": ["Shipment", "Shipment"],
            "Quantity": [3.0, 7.0],
            "Order_Id": ["O1", "O2"],
            "Invoice_Number": ["", ""],
        }
    )
    sales = build_sales_df(
        mtr_df=mtr,
        myntra_df=pd.DataFrame(),
        meesho_df=pd.DataFrame(),
        flipkart_df=pd.DataFrame(),
        sku_mapping={},
    )
    plat = get_platform_summary(
        mtr, pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), None,
        start_date="2025-04-01",
        end_date="2025-04-30",
        sales_df=sales,
    )
    amz = next(p for p in plat if p["platform"] == "Amazon")
    exp = filter_sales_for_export(
        sales, start_date="2025-04-01", end_date="2025-04-30", sources=["Amazon"]
    )
    exp_ship = exp[exp["Transaction Type"] == "Shipment"]["Quantity"].sum()
    assert amz["total_units"] == int(exp_ship)
    assert amz["total_units"] == 10


def test_map_to_oms_normalizes_yrn_float_string():
    m = {"47061570": "OMS-YRN"}
    assert map_to_oms_sku("47061570.0", m) == "OMS-YRN"
    assert map_to_oms_sku("4.706157e+07", m) == "OMS-YRN"


def test_flipkart_sales_report_maps_sku_id_when_sku_column_is_dash():
    from io import BytesIO

    from backend.services.flipkart import _parse_flipkart_xlsx

    m = {"FKLISTING99": "OMS-FK"}
    df = pd.DataFrame(
        {
            "Buyer Invoice Date": [pd.Timestamp("2025-01-15")],
            "Event Sub Type": ["Sale"],
            "Item Quantity": [1],
            "Buyer Invoice Amount": [100.0],
            "SKU": ["-"],
            "SKU ID": ["FKLISTING99"],
            "Order ID": ["OD334028117901509100"],
        }
    )
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="Sales Report", index=False)
    out = _parse_flipkart_xlsx(buf.getvalue(), "fk-Jan-2025.xlsx", m)
    assert not out.empty
    assert out["OMS_SKU"].iloc[0] == "OMS-FK"


def test_flipkart_parse_excel_order_dates_serial_and_iso():
    from datetime import date

    from backend.services.flipkart import _fk_parse_excel_order_dates

    s = pd.Series(["46126", "2026-04-01", 46113.0])
    out = _fk_parse_excel_order_dates(s)
    assert list(out.dt.date) == [date(2026, 4, 14), date(2026, 4, 1), date(2026, 4, 1)]


def test_flipkart_earn_more_prefers_final_sale_units_over_gross():
    """earn_more_report must not count cancelled demand in shipments (Final Sale vs Gross)."""
    from io import BytesIO

    from openpyxl import Workbook

    from backend.services.flipkart import _parse_flipkart_earn_more

    wb = Workbook()
    ws = wb.active
    ws.title = "earn_more_report"
    ws.append(
        [
            "Order Date",
            "SKU ID",
            "Gross Units",
            "Final Sale Units",
            "Return Units",
            "Cancellation Units",
            "Final Sale Amount",
            "Return Amount",
            "Brand",
        ]
    )
    ws.append(["2026-04-01", "LIST1", 10, 7, 0, 3, 700, 0, "B1"])
    buf = BytesIO()
    wb.save(buf)
    out = _parse_flipkart_earn_more(buf.getvalue(), "earn-Apr-2026.xlsx", {"LIST1": "OMS-1"})
    ship = out[out["TxnType"] == "Shipment"]
    assert len(ship) == 1
    assert float(ship["Quantity"].iloc[0]) == 7.0


def test_flipkart_earn_more_parses_numeric_excel_serial_order_date():
    """Order Date stored as Excel serial (common when re-saving XLSX) must not become NaT."""
    from io import BytesIO

    from openpyxl import Workbook

    from backend.services.flipkart import _parse_flipkart_earn_more

    wb = Workbook()
    ws = wb.active
    ws.title = "earn_more_report"
    ws.append(
        [
            "Order Date",
            "SKU ID",
            "Gross Units",
            "Final Sale Units",
            "Return Units",
            "Cancellation Units",
            "Final Sale Amount",
            "Return Amount",
            "Brand",
        ]
    )
    ws.append([46126, "LIST1", 10, 5, 0, 5, 500, 0, "B1"])
    buf = BytesIO()
    wb.save(buf)
    out = _parse_flipkart_earn_more(buf.getvalue(), "earn-Apr-2026.xlsx", {"LIST1": "OMS-1"})
    ship = out[out["TxnType"] == "Shipment"]
    assert len(ship) == 1
    assert ship["Date"].iloc[0].date().isoformat() == "2026-04-14"
