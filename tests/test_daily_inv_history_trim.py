"""Daily inventory history must be trimmed to recent days at upload time to prevent OOM."""

import pandas as pd
import pytest
from datetime import date, timedelta


def _make_history(n_skus: int, n_days: int) -> pd.DataFrame:
    """Build a synthetic tall inventory history DataFrame."""
    today = pd.Timestamp.now().normalize()
    dates = [today - pd.Timedelta(days=i) for i in range(n_days)]
    rows = []
    for sku_i in range(n_skus):
        sku = f"SKU-{sku_i:04d}"
        for d in dates:
            rows.append({"OMS_SKU": sku, "Date": d, "Qty": 10})
    return pd.DataFrame(rows)


def test_trim_keeps_only_recent_days():
    from backend.services.daily_inventory_upload_run import _trim_history_to_recent

    df = _make_history(n_skus=10, n_days=1000)
    assert len(df) == 10_000

    trimmed, note = _trim_history_to_recent(df, max_days=400)
    # Exactly 10 SKUs × 401 day-buckets (cutoff <= date <= max, inclusive)
    assert len(trimmed) <= 10 * 401
    assert len(trimmed) >= 10 * 399   # allow ±1 day for timestamp rounding
    assert "400" in note
    assert len(note) > 0


def test_trim_noop_when_already_short():
    from backend.services.daily_inventory_upload_run import _trim_history_to_recent

    df = _make_history(n_skus=5, n_days=30)
    trimmed, note = _trim_history_to_recent(df, max_days=400)
    assert len(trimmed) == len(df)
    assert note == ""


def test_trim_noop_when_disabled():
    from backend.services.daily_inventory_upload_run import _trim_history_to_recent

    df = _make_history(n_skus=5, n_days=1000)
    trimmed, note = _trim_history_to_recent(df, max_days=0)
    assert len(trimmed) == len(df)
    assert note == ""


def test_parse_trims_date_columns_before_melt():
    """Wide sheets with years of daily columns should only melt the recent window."""
    from backend.services.daily_inventory_history import parse_daily_inventory_history_dataframes

    n_skus = 100
    n_date_cols = 200  # far more than the 30-day retention window
    today = pd.Timestamp("2026-06-22")
    dates = [today - pd.Timedelta(days=i) for i in range(n_date_cols)]

    header_row = [""] * (2 + n_date_cols)
    header_row[0] = "Item SkuCode"
    header_row[1] = "Parent"
    for j, d in enumerate(dates):
        header_row[2 + j] = d.strftime("%Y-%m-%d")

    rows = [header_row]
    for i in range(n_skus):
        row = [f"SKU-{i:04d}", f"P-{i}"] + [10] * n_date_cols
        rows.append(row)

    wide = pd.DataFrame(rows)
    out = parse_daily_inventory_history_dataframes({"OMS": wide}, max_days=30)
    assert not out.empty
    unique_days = out["Date"].nunique()
    assert unique_days <= 30
    assert len(out) <= n_skus * 30


def test_full_replace_skips_merge(monkeypatch):
    """Re-uploading a full date window should not run merge_inventory_history."""
    import backend.services.daily_inventory_upload_run as mod
    import backend.services.daily_inventory_history as dih_mod
    from backend.session import AppSession

    monkeypatch.setattr(mod, "_MAX_HISTORY_DAYS", 30)
    today = pd.Timestamp.now().normalize()
    dates = [today - pd.Timedelta(days=i) for i in range(30)]
    incoming = pd.DataFrame(
        [{"OMS_SKU": "SKU-A", "Date": d, "Qty": 5.0} for d in dates]
        + [{"OMS_SKU": "SKU-B", "Date": d, "Qty": 3.0} for d in dates]
    )
    existing = incoming.copy()

    merge_calls: list[int] = []

    def _fake_merge(ex, inc):
        merge_calls.append(1)
        return dih_mod.merge_inventory_history(ex, inc)

    def _fake_parse(fp, filename, sku_mapping=None, **kwargs):
        return incoming

    monkeypatch.setattr(dih_mod, "parse_daily_inventory_history_upload", _fake_parse)
    monkeypatch.setattr(dih_mod, "merge_inventory_history", _fake_merge)

    sess = AppSession()
    sess.sku_mapping = {}
    sess.daily_inventory_history_df = existing
    result = mod.execute_daily_inventory_upload(sess, b"dummy", "test.xlsx")
    assert result["ok"] is True
    assert merge_calls == []


def test_column_usecols_trims_wide_xlsx():
    """Optimized xlsx read should only load SKU + recent date columns."""
    import io

    import openpyxl
    from backend.services.daily_inventory_history import (
        _column_usecols_for_inventory_sheet,
        parse_daily_inventory_history_upload,
    )

    n_date_cols = 80
    today = pd.Timestamp("2026-06-22")
    dates = [today - pd.Timedelta(days=i) for i in range(n_date_cols)]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OMS"
    ws.cell(1, 1, "Item SkuCode")
    ws.cell(1, 2, "Parent")
    for j, d in enumerate(dates):
        ws.cell(1, 3 + j, d.strftime("%Y-%m-%d"))
    for row_i in range(50):
        ws.cell(2 + row_i, 1, f"SKU-{row_i:04d}")
        ws.cell(2 + row_i, 2, f"P-{row_i}")
        for j in range(n_date_cols):
            ws.cell(2 + row_i, 3 + j, 10)
    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()

    usecols = _column_usecols_for_inventory_sheet(raw, "OMS", 30)
    assert usecols is not None
    assert len(usecols) <= 32  # SKU + up to 30 dates

    out = parse_daily_inventory_history_upload(io.BytesIO(raw), "test.xlsx", max_days=30)
    assert not out.empty
    assert out["Date"].nunique() <= 30
    assert len(out) <= 50 * 30


def test_execute_applies_trim(monkeypatch):
    """execute_daily_inventory_upload must trim the stored dataframe to _MAX_HISTORY_DAYS."""
    import backend.services.daily_inventory_upload_run as mod
    import backend.services.daily_inventory_history as dih_mod
    from backend.session import AppSession

    monkeypatch.setattr(mod, "_MAX_HISTORY_DAYS", 60)

    # Build a fake 5 SKUs × 500 days history (analogous to 9500 SKU × 3000 day problem).
    big_df = _make_history(n_skus=5, n_days=500)

    def _fake_parse(fp, filename, sku_mapping=None, **kwargs):
        return big_df

    # The function does a local import from .daily_inventory_history — patch that module.
    monkeypatch.setattr(dih_mod, "parse_daily_inventory_history_upload", _fake_parse)

    sess = AppSession()
    sess.sku_mapping = {}
    result = mod.execute_daily_inventory_upload(sess, b"dummy", "test.xlsx")

    assert result["ok"] is True
    # After trimming to 60 days: 5 SKUs × ~61 days ≈ 305 rows (not 2500)
    assert result["rows"] <= 5 * 62
    assert result["rows"] >= 5 * 59
    assert sess.daily_inventory_history_df is not None
    assert len(sess.daily_inventory_history_df) == result["rows"]
    assert "Kept last" in result["message"] or "trimmed" in result["message"].lower()
