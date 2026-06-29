"""
Parse Daily Inventory History (Excel) for PO effective-days calculation.

The export is wide-format. Two sheets are recognised:

    * ``OMS`` (or ``OMS Inventory``) – warehouse on-hand snapshots
    * ``Amazon Inventory``           – FBA snapshots (PL-prefixed SKUs)

The first column stores the SKU code (header text varies, e.g. ``Item SkuCode``).
The second column is the parent style. Subsequent columns are daily snapshots —
the column **header** is the daily total inventory (an integer), and the
**first data row** of each column carries the actual date stamp (``YYYY-MM-DD``).

Output is a tall DataFrame with three columns: ``OMS_SKU`` / ``Date`` / ``Qty``.
PO engine uses it to count only the days a SKU actually had stock on hand;
days with zero (or missing) inventory are excluded from the ADS denominator.
"""
from __future__ import annotations

import io
import os
import re
from datetime import date, datetime
from typing import BinaryIO, Callable, Optional
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd

_IST = ZoneInfo("Asia/Kolkata")
_DEFAULT_VIEW_DAYS = int(os.environ.get("DAILY_INV_VIEW_DAYS", "30"))


_TALL_COLS = ["OMS_SKU", "Date", "Qty"]
_STORE_COLS = ["OMS_SKU", "Date", "Qty", "Source"]
_SOURCE_RANK = {"snapshot": 3, "uploaded": 2, "derived": 1}


def _ensure_source_column(df: pd.DataFrame | None, default: str = "uploaded") -> pd.DataFrame:
    if df is None or getattr(df, "empty", True):
        return pd.DataFrame(columns=_STORE_COLS)
    out = df.copy()
    if "Source" not in out.columns:
        out["Source"] = default
    else:
        out["Source"] = out["Source"].astype(str).replace({"": default, "nan": default, "None": default})
    return out


def _coalesce_history_rows(combined: pd.DataFrame) -> pd.DataFrame:
    """One row per SKU-day — snapshot beats uploaded beats derived; then highest qty."""
    work = _ensure_source_column(combined, default="uploaded")
    work["Date"] = pd.to_datetime(work["Date"], errors="coerce").dt.normalize()
    work["Qty"] = pd.to_numeric(work["Qty"], errors="coerce")
    work = work.dropna(subset=["Date", "OMS_SKU", "Qty"])
    work = work[work["OMS_SKU"].astype(str).str.len() > 0]
    if work.empty:
        return pd.DataFrame(columns=_STORE_COLS)
    work["Qty"] = work["Qty"].astype(float).clip(lower=0.0)
    work["_rank"] = work["Source"].astype(str).str.strip().str.lower().map(_SOURCE_RANK).fillna(1)
    work = work.sort_values(
        ["OMS_SKU", "Date", "_rank", "Qty"],
        ascending=[True, True, False, False],
    )
    out = work.drop_duplicates(subset=["OMS_SKU", "Date"], keep="first").drop(columns=["_rank"])
    return drop_zero_derived_rows(out.reset_index(drop=True))


def record_inventory_snapshot_date(sess, snapshot_date: str) -> None:
    """Track authoritative daily snapshot columns appended to the history matrix."""
    d = str(snapshot_date or "").strip()[:10]
    if len(d) != 10:
        return
    cur = list(getattr(sess, "daily_inventory_history_snapshot_dates", None) or [])
    if d not in cur:
        cur.append(d)
        cur.sort()
        sess.daily_inventory_history_snapshot_dates = cur


def wide_matrix_upload_end_date(sess) -> pd.Timestamp | None:
    """Last date from the one-time wide Excel upload (not daily snapshot extensions)."""
    meta = read_daily_inventory_history_disk_meta() or {}
    for raw in (
        str(getattr(sess, "daily_inventory_history_wide_end_date", "") or "")[:10],
        str(meta.get("daily_inventory_history_wide_end_date") or "")[:10],
        inventory_sheet_end_date_from_filename(
            str(getattr(sess, "daily_inventory_history_filename", "") or "")
            or str(meta.get("daily_inventory_history_filename") or "")
        ),
    ):
        if len(raw) == 10:
            try:
                return pd.Timestamp(raw).normalize()
            except Exception:
                continue
    return None


def snapshot_dates_from_history(hist: pd.DataFrame | None) -> list[str]:
    if hist is None or getattr(hist, "empty", True) or "Source" not in hist.columns:
        return []
    work = _ensure_source_column(hist)
    mask = work["Source"].astype(str).str.lower() == "snapshot"
    if not bool(mask.any()):
        return []
    days = pd.to_datetime(work.loc[mask, "Date"], errors="coerce").dt.normalize().dropna()
    return sorted({str(pd.Timestamp(d).date()) for d in days.unique()})


def prune_non_snapshot_post_matrix_days(
    hist: pd.DataFrame | None,
    sess,
    *,
    extra_keep_dates: set[str] | None = None,
) -> pd.DataFrame:
    """Remove sales-derived filler after the wide matrix except known snapshot days."""
    if hist is None or getattr(hist, "empty", True):
        return hist if hist is not None else pd.DataFrame(columns=_STORE_COLS)
    wide_end = wide_matrix_upload_end_date(sess)
    if wide_end is None:
        return _ensure_source_column(hist)
    keep: set[str] = set(extra_keep_dates or set())
    keep.update(str(getattr(sess, "daily_inventory_history_snapshot_dates", None) or []))
    snap = str(getattr(sess, "inventory_snapshot_date", "") or "").strip()[:10]
    if len(snap) == 10:
        keep.add(snap)
    work = _ensure_source_column(hist)
    work["Date"] = pd.to_datetime(work["Date"], errors="coerce").dt.normalize()
    day_s = work["Date"].dt.strftime("%Y-%m-%d")
    is_snapshot = work["Source"].astype(str).str.lower() == "snapshot"
    keep_mask = (work["Date"] <= wide_end) | day_s.isin(keep) | is_snapshot
    return work.loc[keep_mask].reset_index(drop=True)

# Yash wide-matrix exports: ``28-5-26`` = 28 May 2026 (day-month-year).
_DMY_HEADER_RE = re.compile(r"^(\d{1,2})-(\d{1,2})-(\d{2,4})$")


def _safe_normalize(ts) -> pd.Timestamp | None:
    """Normalize a timestamp; return None for NaT / unparsable values."""
    if ts is None:
        return None
    try:
        t = pd.Timestamp(ts)
    except Exception:
        return None
    if pd.isna(t):
        return None
    return t.normalize()


def _parse_inventory_snapshot_date(value) -> pd.Timestamp | None:
    """Parse snapshot dates from wide-matrix headers (D-M-YY) and other exports."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, pd.Timestamp):
        return _safe_normalize(value)
    if isinstance(value, (datetime, date)):
        return _safe_normalize(value)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        fv = float(value)
        if 40_000 <= fv <= 60_000:
            try:
                ts = pd.to_datetime(fv, unit="D", origin="1899-12-30", errors="coerce")
                if ts is not None and pd.notna(ts) and 2015 <= int(ts.year) <= 2035:
                    return _safe_normalize(ts)
            except Exception:
                pass
        return None
    s = str(value).strip()
    if not s or s.lower() in {"nan", "none", "nat", "days"}:
        return None
    m = _DMY_HEADER_RE.match(s)
    if m:
        day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if year < 100:
            year += 2000
        try:
            ts = pd.Timestamp(year=year, month=month, day=day)
            if 2015 <= int(ts.year) <= 2035:
                return _safe_normalize(ts)
        except ValueError:
            return None
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        try:
            ts = pd.to_datetime(s[:10], errors="coerce")
            if ts is not None and pd.notna(ts) and 2015 <= int(ts.year) <= 2035:
                return _safe_normalize(ts)
        except Exception:
            return None
    try:
        ts = pd.to_datetime(s, errors="coerce")
        if ts is not None and pd.notna(ts) and 2015 <= int(ts.year) <= 2035:
            return pd.Timestamp(ts).normalize()
    except Exception:
        return None
    return None


def _norm(s) -> str:
    return "".join(ch for ch in str(s).strip().lower() if ch.isalnum())


def _looks_like_sku_header(value) -> bool:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return False
    n = _norm(value)
    return n in {
        "skucode",
        "itemskucode",
        "sku",
        "omssku",
        "item",
        "itemcode",
        "sellersku",
        "stylesku",
    }


def _is_date_value(v) -> bool:
    return _parse_inventory_snapshot_date(v) is not None


_VARIANT_LABELS = {"itemskucode", "skucode", "sellersku", "stylesku"}
_PARENT_LABELS = {"item", "itemcode", "parent", "parentsku", "parentstyle", "style", "sku", "omssku"}


def _detect_sku_column(df: pd.DataFrame) -> Optional[int]:
    """
    Return position of the SKU (variant) column.

    Real layout uses row 0 as a label row (``Item SkuCode``/``Item``); column
    headers are often noise (``Total Inv.``, ``SKU``). Treat row 0 as truth and
    bias toward variant-style markers; only fall back to column headers when
    row 0 has no recognisable label.
    """
    if df is None or df.empty or df.shape[1] == 0:
        return None

    if df.shape[0] > 0:
        first = list(df.iloc[0].values)
        variant_first = next(
            (i for i, v in enumerate(first) if _norm(v) in _VARIANT_LABELS),
            None,
        )
        if variant_first is not None:
            return variant_first
        any_first = next((i for i, v in enumerate(first) if _looks_like_sku_header(v)), None)
        if any_first is not None:
            return any_first

    cols = list(df.columns)
    variant_col = next((i for i, c in enumerate(cols) if _norm(c) in _VARIANT_LABELS), None)
    if variant_col is not None:
        return variant_col
    any_col = next((i for i, c in enumerate(cols) if _looks_like_sku_header(c)), None)
    if any_col is not None:
        return any_col
    return 0  # fallback to leftmost column


def _detect_parent_column(df: pd.DataFrame, sku_idx: int) -> Optional[int]:
    """Best-effort: parent column sits next to SKU column with header 'Item' / 'SKUCode'."""
    candidates: list[int] = []
    if df.shape[0] > 0:
        first = df.iloc[0]
        for i, val in enumerate(first.values):
            if i == sku_idx:
                continue
            if _norm(val) in _PARENT_LABELS:
                candidates.append(i)
    if not candidates:
        for i, col in enumerate(df.columns):
            if i == sku_idx:
                continue
            if _norm(col) in _PARENT_LABELS:
                candidates.append(i)
    return candidates[0] if candidates else None


def _row_date_hints(df: pd.DataFrame, row_idx: int, skip_cols: set[int]) -> dict[int, pd.Timestamp]:
    out: dict[int, pd.Timestamp] = {}
    if row_idx < 0 or row_idx >= len(df):
        return out
    row = df.iloc[row_idx]
    for i, val in enumerate(row.values):
        if i in skip_cols:
            continue
        if _is_date_value(val):
            ts = _parse_inventory_snapshot_date(val)
            if ts is not None:
                out[i] = ts
    return out


def _build_column_date_map(
    df: pd.DataFrame,
    sku_idx: int,
    parent_idx: Optional[int],
) -> tuple[dict[int, pd.Timestamp], int]:
    """Map column index → snapshot date; return (date_map, first_data_row_index).

    Yash exports put dates in row 0 for most columns; a newly added "today"
  column may have its date only in row 1. Union hints from the first few rows
    and column headers instead of picking only row 0 OR only headers.
    """
    skip = {sku_idx}
    if parent_idx is not None:
        skip.add(parent_idx)

    header_dates: dict[int, pd.Timestamp] = {}
    for i, col in enumerate(df.columns):
        if i in skip:
            continue
        if isinstance(col, (pd.Timestamp,)) or _is_date_value(col):
            ts = _parse_inventory_snapshot_date(col)
            if ts is not None:
                header_dates[i] = ts

    row_maps: list[dict[int, pd.Timestamp]] = []
    for ridx in range(min(4, len(df))):
        rm = _row_date_hints(df, ridx, skip)
        if rm:
            row_maps.append(rm)

    date_map: dict[int, pd.Timestamp] = {}
    if row_maps:
        primary = max(row_maps, key=len)
        date_map.update(primary)
        for rm in row_maps:
            for i, d in rm.items():
                date_map.setdefault(i, d)
    for i, d in header_dates.items():
        date_map.setdefault(i, d)

    header_row_count = 0
    if date_map:
        positions = sorted(date_map.keys())
        for ridx in range(min(4, len(df))):
            row = df.iloc[ridx]
            hits = 0
            for i in positions:
                if i >= len(row):
                    continue
                if _is_date_value(row.iloc[i]):
                    hits += 1
            if hits >= max(2, len(positions) // 3):
                header_row_count = ridx + 1

    return date_map, header_row_count


def _trim_date_map_to_window(
    date_map: dict[int, pd.Timestamp],
    max_days: int,
) -> dict[int, pd.Timestamp]:
    """Drop snapshot columns outside the last ``max_days`` calendar days (anchored on max date)."""
    if max_days <= 0 or not date_map:
        return date_map
    max_date = max(pd.Timestamp(d).normalize() for d in date_map.values())
    cutoff = max_date - pd.Timedelta(days=max_days - 1)
    return {
        i: d
        for i, d in date_map.items()
        if pd.Timestamp(d).normalize() >= cutoff
    }


def drop_zero_derived_rows(df: pd.DataFrame | None) -> pd.DataFrame:
    """Drop sales-derived zero rows for SKUs that never had uploaded on-hand."""
    if df is None or getattr(df, "empty", True):
        return pd.DataFrame(columns=_TALL_COLS)
    if "Source" not in df.columns:
        return df
    qty = pd.to_numeric(df["Qty"], errors="coerce").fillna(0.0)
    uploaded = df["Source"].astype(str) != "derived"
    stocked_skus = set(
        df.loc[uploaded & (qty > 0), "OMS_SKU"].astype(str).str.strip()
    )
    drop = (
        (df["Source"].astype(str) == "derived")
        & (qty <= 0)
        & (~df["OMS_SKU"].astype(str).isin(stocked_skus))
    )
    if not bool(drop.any()):
        return df
    return df.loc[~drop].reset_index(drop=True)


def merge_inventory_history(
    existing: Optional[pd.DataFrame],
    incoming: pd.DataFrame,
) -> pd.DataFrame:
    """Union SKU-day rows; snapshot/uploaded rows beat sales-derived duplicates."""
    if incoming is None or incoming.empty:
        return _ensure_source_column(existing) if existing is not None else pd.DataFrame(columns=_STORE_COLS)
    if existing is None or existing.empty:
        return _ensure_source_column(incoming, default="uploaded")
    combined = pd.concat(
        [_ensure_source_column(existing), _ensure_source_column(incoming, default="uploaded")],
        ignore_index=True,
    )
    return _coalesce_history_rows(combined)


def inventory_history_max_date(df: Optional[pd.DataFrame]) -> Optional[pd.Timestamp]:
    if df is None or df.empty or "Date" not in df.columns:
        return None
    mx = pd.to_datetime(df["Date"], errors="coerce").max()
    return pd.Timestamp(mx).normalize() if pd.notna(mx) else None


_FILENAME_END_RE = re.compile(
    r"(?:\bto\b|\bthrough\b|\buntil\b)\s+(\d{1,2})[\-\s]([A-Za-z]{3,9})[\-\s](\d{2,4})",
    re.I,
)


def inventory_sheet_end_date_from_filename(filename: str) -> str:
    """Best-effort sheet end date from operator filenames (e.g. '… To 25-Jun-26.xlsx')."""
    from datetime import datetime

    m = _FILENAME_END_RE.search(str(filename or ""))
    if not m:
        return ""
    day_s, mon_s, yr_s = m.groups()
    yr = int(yr_s)
    if yr < 100:
        yr += 2000
    for fmt in ("%d-%b-%Y", "%d-%B-%Y"):
        try:
            return datetime.strptime(f"{int(day_s)}-{mon_s}-{yr}", fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return ""


def promote_daily_inventory_matrix_max_date(sess, candidate: str | None) -> None:
    """Never downgrade the wide-matrix upload end date on the session."""
    cand = str(candidate or "").strip()[:10]
    if len(cand) != 10:
        return
    cur = str(getattr(sess, "daily_inventory_history_matrix_max_date", "") or "").strip()[:10]
    if not cur or cand > cur:
        sess.daily_inventory_history_matrix_max_date = cand


def recanonicalize_inventory_history_skus(
    df: pd.DataFrame | None,
    sku_mapping: dict | None,
) -> pd.DataFrame:
    """Re-key history rows with the current SKU mapping (same keys PO engine uses)."""
    if df is None or getattr(df, "empty", True):
        return pd.DataFrame(columns=_TALL_COLS)
    mapping = sku_mapping if isinstance(sku_mapping, dict) else {}
    if not mapping:
        return df
    from .po_engine import canonical_oms_key

    out = df.copy()
    out["OMS_SKU"] = out["OMS_SKU"].astype(str).map(lambda s: canonical_oms_key(s, mapping))
    out = out[out["OMS_SKU"].astype(str).str.len() > 0]
    if out.empty:
        return pd.DataFrame(columns=_TALL_COLS)
    out["Date"] = pd.to_datetime(out["Date"], errors="coerce").dt.normalize()
    out["Qty"] = pd.to_numeric(out["Qty"], errors="coerce")
    out = out.dropna(subset=["Date", "Qty", "OMS_SKU"])
    if out.empty:
        return pd.DataFrame(columns=_TALL_COLS)
    return (
        out.groupby(["OMS_SKU", "Date"], as_index=False)["Qty"]
        .max()
        .sort_values(["OMS_SKU", "Date"])
        .reset_index(drop=True)
    )


def today_ist_timestamp() -> pd.Timestamp:
    return pd.Timestamp.now(tz=_IST).normalize().tz_localize(None)


def inventory_history_view_end_date(
    df: pd.DataFrame | None,
    end_date: str | None = None,
) -> str:
    """End anchor for UI windows — use latest matrix date when it trails today."""
    try:
        end = pd.Timestamp(str(end_date or "")[:10]).normalize() if end_date else today_ist_timestamp()
    except Exception:
        end = today_ist_timestamp()
    if pd.isna(end):
        end = today_ist_timestamp()
    mx = inventory_history_max_date(df)
    if mx is not None and mx < end - pd.Timedelta(days=1):
        end = pd.Timestamp(mx).normalize()
    return str(end.date())


def disk_inventory_meta_looks_placeholder(meta: dict | None) -> bool:
    """True when on-disk meta is a tiny stub (e.g. 1 row / 1 SKU) rather than the real matrix."""
    if not isinstance(meta, dict) or not meta:
        return True
    rows = int(meta.get("daily_inventory_history_rows") or 0)
    skus = int(meta.get("daily_inventory_history_skus") or 0)
    return rows < 100 or skus < 50


def session_inventory_matrix_stats(sess) -> tuple[int, int, pd.Timestamp | None]:
    """Row count, SKU count, and max date from the session matrix (if any)."""
    df = getattr(sess, "daily_inventory_history_df", None)
    if df is None or getattr(df, "empty", True) or "OMS_SKU" not in df.columns:
        return 0, 0, None
    rows = int(len(df))
    skus = int(df["OMS_SKU"].astype(str).nunique())
    mx = inventory_history_max_date(df)
    df_ts = pd.Timestamp(mx).normalize() if mx is not None and pd.notna(mx) else None
    return rows, skus, df_ts


def inventory_history_matrix_cap_date(sess) -> pd.Timestamp | None:
    """Last date from the wide matrix upload — not extended by daily snapshot roll-forward."""
    meta = read_daily_inventory_history_disk_meta() or {}

    def _parse_cap(raw: str) -> pd.Timestamp | None:
        s = str(raw or "").strip()[:10]
        if len(s) != 10:
            return None
        try:
            return pd.Timestamp(s).normalize()
        except Exception:
            return None

    upload_caps: list[pd.Timestamp] = []
    disk_looks_real = not disk_inventory_meta_looks_placeholder(meta)
    cap_sources: list[str] = [str(getattr(sess, "daily_inventory_history_matrix_max_date", "") or "")]
    if disk_looks_real:
        cap_sources.extend(
            [
                str(meta.get("daily_inventory_history_matrix_max_date") or ""),
                str(meta.get("daily_inventory_history_max_date") or ""),
            ]
        )
    cap_sources.append(
        inventory_sheet_end_date_from_filename(
            str(
                getattr(sess, "daily_inventory_history_filename", "")
                or meta.get("daily_inventory_history_filename", "")
                or ""
            )
        )
    )
    for raw in cap_sources:
        ts = _parse_cap(raw)
        if ts is not None:
            upload_caps.append(ts)
    if upload_caps:
        return max(upload_caps)

    _, sess_skus, df_ts = session_inventory_matrix_stats(sess)
    if df_ts is not None and sess_skus >= 50:
        return df_ts
    disk_ts = _parse_cap(
        meta.get("daily_inventory_history_matrix_max_date")
        or meta.get("daily_inventory_history_max_date")
        or ""
    )
    if disk_ts is not None and not disk_inventory_meta_looks_placeholder(meta):
        return disk_ts
    return df_ts


def inventory_history_authoritative_cap_date(sess) -> pd.Timestamp:
    """Last day with authoritative inventory (wide matrix + daily snapshot columns)."""
    caps: list[pd.Timestamp] = []
    matrix_cap = inventory_history_matrix_cap_date(sess)
    if matrix_cap is not None:
        caps.append(matrix_cap)
    sess_mx = str(getattr(sess, "daily_inventory_history_matrix_max_date", "") or "").strip()[:10]
    if len(sess_mx) == 10:
        try:
            caps.append(pd.Timestamp(sess_mx).normalize())
        except Exception:
            pass
    snap = str(getattr(sess, "inventory_snapshot_date", "") or "").strip()[:10]
    if len(snap) == 10:
        try:
            caps.append(pd.Timestamp(snap).normalize())
        except Exception:
            pass
    df_max = inventory_history_max_date(getattr(sess, "daily_inventory_history_df", None))
    if df_max is not None:
        caps.append(df_max)
    if caps:
        return max(caps)
    return today_ist_timestamp()


def inventory_history_is_newer_than(
    incoming: pd.DataFrame | None,
    existing: pd.DataFrame | None,
    *,
    incoming_uploaded_at: str = "",
    existing_uploaded_at: str = "",
) -> bool:
    """True when ``incoming`` should replace ``existing`` (never downgrade on PO calc persist)."""
    if incoming is None or getattr(incoming, "empty", True):
        return False
    if existing is None or getattr(existing, "empty", True):
        return True
    in_at = upload_timestamp_epoch(incoming_uploaded_at)
    ex_at = upload_timestamp_epoch(existing_uploaded_at)
    if in_at > ex_at + 0.5:
        return True
    if ex_at > in_at + 0.5:
        return False
    in_max = inventory_history_max_date(incoming)
    ex_max = inventory_history_max_date(existing)
    if in_max is not None and ex_max is not None:
        if in_max > ex_max:
            return True
        if in_max < ex_max:
            return False
    if in_at > 0 or ex_at > 0:
        return in_at >= ex_at
    in_skus = int(incoming["OMS_SKU"].astype(str).nunique()) if "OMS_SKU" in incoming.columns else 0
    ex_skus = int(existing["OMS_SKU"].astype(str).nunique()) if "OMS_SKU" in existing.columns else 0
    return in_skus >= ex_skus


def filter_inventory_history_window(
    df: pd.DataFrame,
    *,
    days: int | None = None,
    end_date: str | None = None,
) -> pd.DataFrame:
    """Keep SKU-day rows in the last ``days`` calendar days ending on ``end_date`` (default today IST)."""
    if df is None or df.empty or "Date" not in df.columns:
        return df if df is not None else pd.DataFrame(columns=_TALL_COLS)
    span = int(days if days is not None else _DEFAULT_VIEW_DAYS)
    if span <= 0:
        return df.copy()
    if end_date:
        try:
            end = pd.Timestamp(str(end_date)[:10]).normalize()
        except Exception:
            end = today_ist_timestamp()
    else:
        end = pd.Timestamp(inventory_history_view_end_date(df))
    if pd.isna(end):
        end = today_ist_timestamp()
    start = end - pd.Timedelta(days=max(0, span - 1))
    work = df.copy()
    work["Date"] = pd.to_datetime(work["Date"], errors="coerce").dt.normalize()
    work = work.dropna(subset=["Date"])
    mask = (work["Date"] >= start) & (work["Date"] <= end)
    return work.loc[mask].reset_index(drop=True)


def filter_inventory_history_view(
    df: pd.DataFrame,
    *,
    days: int | None = None,
    end_date: str | None = None,
) -> pd.DataFrame:
    """Keep rows in the last ``days`` calendar days ending on the view anchor date."""
    return filter_inventory_history_window(df, days=days, end_date=end_date)


def should_skip_inventory_history_extend(
    sheet_max: pd.Timestamp | None,
    inv_window_end: pd.Timestamp,
    coverage_in_window: int,
    *,
    ads_window: int = 90,
) -> bool:
    """Whether PO calc can use the uploaded sheet without sales roll-forward.

    Partial-coverage sheets that still reach the window end may skip extend, but
    a sheet whose last column is weeks behind ``inv_window_end`` must roll forward
    even when it has 30+ snapshot days in the ADS window (stale May data vs June sales).
    """
    if pd.isna(sheet_max):
        return False
    sheet_max = pd.Timestamp(sheet_max).normalize()
    inv_window_end = pd.Timestamp(inv_window_end).normalize()
    sheet_covers_end = sheet_max >= inv_window_end - pd.Timedelta(days=1)
    if sheet_covers_end:
        return True
    sheet_stale_vs_window = sheet_max < inv_window_end - pd.Timedelta(days=1)
    if sheet_stale_vs_window:
        return False
    return coverage_in_window >= min(int(ads_window), 14)


def _sales_for_inventory_rollforward(sess) -> pd.DataFrame | None:
    sales = getattr(sess, "sales_df", None)
    if sales is not None and not getattr(sales, "empty", True):
        return sales
    try:
        from .po_calculate_run import _build_platform_sales_df

        built = _build_platform_sales_df(sess)
        if built is not None and not built.empty:
            return built
    except Exception:
        pass
    return None


def _variant_snapshot_qty_series(variant: pd.DataFrame) -> pd.Series | None:
    """Total on-hand per SKU row for daily snapshot → history append."""
    if variant is None or getattr(variant, "empty", True):
        return None
    if "Total_Inventory" in variant.columns:
        return pd.to_numeric(variant["Total_Inventory"], errors="coerce").fillna(0.0)
    try:
        from .inventory import recompute_inventory_totals

        work = recompute_inventory_totals(variant.copy())
        if "Total_Inventory" in work.columns:
            return pd.to_numeric(work["Total_Inventory"], errors="coerce").fillna(0.0)
    except Exception:
        pass
    if "OMS_Inventory" in variant.columns:
        return pd.to_numeric(variant["OMS_Inventory"], errors="coerce").fillna(0.0)
    return None


def refresh_inventory_history_rollforward(
    sess,
    *,
    cap_date: str | pd.Timestamp | None = None,
    sales_df: pd.DataFrame | None = None,
    include_snapshot: bool = True,
    max_history_days: int | None = None,
) -> dict:
    """Extend uploaded history with sales activity and optional daily snapshot."""
    from .daily_inventory_upload_run import _MAX_HISTORY_DAYS

    hist = getattr(sess, "daily_inventory_history_df", None)
    if hist is None or getattr(hist, "empty", True):
        return {"ok": False, "reason": "empty_history"}

    if include_snapshot:
        inferred = snapshot_dates_from_history(hist)
        if inferred and not getattr(sess, "daily_inventory_history_snapshot_dates", None):
            sess.daily_inventory_history_snapshot_dates = inferred
        hist = prune_non_snapshot_post_matrix_days(hist, sess)

    span = int(max_history_days if max_history_days is not None else _MAX_HISTORY_DAYS)
    sales = sales_df if sales_df is not None else _sales_for_inventory_rollforward(sess)

    snap = str(getattr(sess, "inventory_snapshot_date", "") or "").strip()[:10]
    sheet_max = inventory_history_max_date(hist)
    if cap_date is not None:
        cap_ts = pd.Timestamp(cap_date).normalize()
    elif include_snapshot and len(snap) == 10:
        cap_ts = pd.Timestamp(snap).normalize()
    elif sheet_max is not None:
        cap_ts = pd.Timestamp(sheet_max).normalize()
    else:
        cap_ts = inventory_history_authoritative_cap_date(sess)
    matrix_cap = inventory_history_matrix_cap_date(sess)
    # Clamp roll-forward only for PO-calc paths — daily snapshot uploads extend the matrix.
    if not include_snapshot and matrix_cap is not None:
        cap_ts = min(cap_ts, matrix_cap)
    extended = hist
    rolled = False
    # Daily snapshot upload: append one authoritative column — never sales-derived fill.
    if not include_snapshot and sheet_max is not None and sheet_max < cap_ts:
        extended = extend_history_with_sales(hist, sales_df=sales, cap_date=cap_ts)
        rolled = True

    merged = _ensure_source_column(extended)
    snapshot_appended = False
    if include_snapshot:
        wide_end = wide_matrix_upload_end_date(sess)
        snap_date = snap if len(snap) == 10 else str(cap_ts.date())
        snap_ts = pd.Timestamp(snap_date).normalize()
        if wide_end is not None and snap_ts > pd.Timestamp(wide_end).normalize():
            fill_cap = snap_ts - pd.Timedelta(days=1)
            wide_ts = pd.Timestamp(wide_end).normalize()
            if fill_cap > wide_ts:
                clip = merged[merged["Date"] <= wide_ts].copy()
                filled = extend_history_with_sales(clip, sales_df=sales, cap_date=fill_cap)
                snap_rows = merged[
                    (merged["Date"] > wide_ts)
                    & (merged["Source"].astype(str).str.lower() == "snapshot")
                ]
                merged = merge_inventory_history(filled, snap_rows)
                rolled = True
        variant = getattr(sess, "inventory_df_variant", None)
        if variant is not None and not getattr(variant, "empty", True) and "OMS_SKU" in variant.columns:
            snap_date = snap if len(snap) == 10 else str(cap_ts.date())
            snap_ts = pd.Timestamp(snap_date).normalize()
            work = variant.copy()
            work["OMS_SKU"] = work["OMS_SKU"].astype(str).str.strip().str.upper()
            qty = _variant_snapshot_qty_series(work)
            if qty is not None:
                work["Qty"] = qty.values
                work = work[work["OMS_SKU"].str.len() > 0]
                hist_skus = set(merged["OMS_SKU"].astype(str).str.strip().str.upper())
                work = work[work["OMS_SKU"].isin(hist_skus)]
                if not work.empty:
                    incoming = pd.DataFrame(
                        {
                            "OMS_SKU": work["OMS_SKU"],
                            "Date": snap_ts,
                            "Qty": work["Qty"],
                            "Source": "snapshot",
                        }
                    )
                    merged = merged[
                        ~((merged["Date"] == snap_ts) & (merged["Source"].astype(str) == "derived"))
                    ]
                    merged = merge_inventory_history(merged, incoming)
                    snapshot_appended = True
                    record_inventory_snapshot_date(sess, str(snap_ts.date()))
                    promote_daily_inventory_matrix_max_date(sess, str(snap_ts.date()))

    end_anchor = inventory_history_max_date(merged)
    end_s = str(end_anchor.date()) if end_anchor is not None else None
    merged = filter_inventory_history_window(merged, days=span, end_date=end_s)
    sess.daily_inventory_history_df = merged
    sess._quarterly_cache.clear()
    if end_anchor is not None:
        promote_daily_inventory_matrix_max_date(sess, str(pd.Timestamp(end_anchor).date()))

    dates_norm = pd.to_datetime(merged["Date"], errors="coerce").dt.normalize()
    min_d = dates_norm.min()
    max_d = dates_norm.max()
    return {
        "ok": True,
        "rolled_forward": rolled,
        "snapshot_appended": snapshot_appended,
        "rows": int(len(merged)),
        "skus": int(merged["OMS_SKU"].nunique()) if not merged.empty else 0,
        "days": int(dates_norm.nunique()) if not merged.empty else 0,
        "min_date": str(pd.Timestamp(min_d).date()) if pd.notna(min_d) else "",
        "max_date": str(pd.Timestamp(max_d).date()) if pd.notna(max_d) else "",
        "cap_date": str(cap_ts.date()),
    }


def append_snapshot_inventory_to_history(sess) -> dict:
    """
    After a daily snapshot inventory upload, roll history forward with sales and
    append the snapshot column (authoritative on-hand for that day).
    """
    variant = getattr(sess, "inventory_df_variant", None)
    if variant is None or getattr(variant, "empty", True):
        return {"appended": False, "reason": "empty_snapshot"}
    if "OMS_SKU" not in variant.columns:
        return {"appended": False, "reason": "no_sku_column"}
    if _variant_snapshot_qty_series(variant) is None:
        return {"appended": False, "reason": "no_total_inventory"}

    result = refresh_inventory_history_rollforward(sess, include_snapshot=True)
    if not result.get("ok"):
        return {"appended": False, "reason": result.get("reason", "refresh_failed")}
    snap = str(getattr(sess, "inventory_snapshot_date", "") or "").strip()[:10]
    return {
        "appended": bool(result.get("snapshot_appended")),
        "rolled_forward": bool(result.get("rolled_forward")),
        "snapshot_date": snap or result.get("max_date") or "",
        "rows": result.get("rows", 0),
        "skus": result.get("skus", 0),
        "days": result.get("days", 0),
        "max_date": result.get("max_date", ""),
    }


def _column_usecols_for_inventory_sheet(
    raw: bytes,
    sheet_name: str,
    max_days: int | None,
) -> list[int] | None:
    """Peek sheet headers; return column indices to read (SKU + recent date columns only)."""
    try:
        import openpyxl
    except ImportError:
        return None
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                return None
            ws = wb[sheet_name]
            rows: list[list] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 5:
                    break
                rows.append(list(row) if row else [])
        finally:
            wb.close()
    except Exception:
        return None
    if not rows:
        return None
    width = max(len(r) for r in rows)
    header_df = pd.DataFrame([r + [None] * (width - len(r)) for r in rows])
    sku_idx = _detect_sku_column(header_df)
    if sku_idx is None:
        return None
    parent_idx = _detect_parent_column(header_df, sku_idx)
    date_map, _ = _build_column_date_map(header_df, sku_idx, parent_idx)
    if max_days is not None and max_days > 0:
        date_map = _trim_date_map_to_window(date_map, max_days)
    if not date_map:
        return None
    return [sku_idx] + sorted(i for i in date_map if i != sku_idx)


def _read_inventory_history_sheet(
    raw: bytes,
    sheet_name: str,
    *,
    max_days: int | None = None,
) -> pd.DataFrame:
    """Read only SKU + recent date columns from a wide inventory sheet."""
    usecols = _column_usecols_for_inventory_sheet(raw, sheet_name, max_days)
    if usecols is None:
        return pd.read_excel(
            io.BytesIO(raw),
            sheet_name=sheet_name,
            header=None,
            dtype=str,
            engine="openpyxl",
        )
    return pd.read_excel(
        io.BytesIO(raw),
        sheet_name=sheet_name,
        header=None,
        usecols=usecols,
        dtype=str,
        engine="openpyxl",
    )


def _parse_one_sheet(
    df: pd.DataFrame,
    mapping: dict,
    sheet_name: str = "",
    *,
    max_days: int | None = None,
) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=_TALL_COLS)

    sku_idx = _detect_sku_column(df)
    if sku_idx is None:
        return pd.DataFrame(columns=_TALL_COLS)
    parent_idx = _detect_parent_column(df, sku_idx)

    date_map, header_row_count = _build_column_date_map(df, sku_idx, parent_idx)
    if max_days is not None and max_days > 0:
        date_map = _trim_date_map_to_window(date_map, max_days)
    if not date_map:
        # Sheet shape unrecognised — fail silently for this sheet, parser keeps trying others.
        return pd.DataFrame(columns=_TALL_COLS)

    # Slice to data rows only (skip label/date header rows).
    body = df.iloc[header_row_count:] if header_row_count > 0 else df
    if body.empty:
        return pd.DataFrame(columns=_TALL_COLS)

    date_positions = sorted(i for i in date_map if i < body.shape[1])
    if not date_positions:
        return pd.DataFrame(columns=_TALL_COLS)

    block = body.iloc[:, [sku_idx] + date_positions].copy()
    block.columns = ["_raw_sku"] + [
        pd.Timestamp(date_map[i]).strftime("%Y-%m-%d") for i in date_positions
    ]
    tall = block.melt(id_vars="_raw_sku", var_name="Date", value_name="Qty")
    tall["Date"] = pd.to_datetime(tall["Date"], errors="coerce").dt.normalize()
    tall = tall.dropna(subset=["Date"])
    tall = tall.dropna(subset=["_raw_sku"])
    tall["_raw_sku"] = tall["_raw_sku"].astype(str).str.strip()
    tall = tall[tall["_raw_sku"].str.len() > 0]
    tall = tall[~tall["_raw_sku"].str.lower().isin({"nan", "none", "<na>", "nat"})]
    if tall.empty:
        return pd.DataFrame(columns=_TALL_COLS)

    from .po_engine import canonical_oms_key

    unique_raw = tall["_raw_sku"].unique()
    canon_map = {r: canonical_oms_key(r, mapping) for r in unique_raw}
    tall["OMS_SKU"] = tall["_raw_sku"].map(canon_map)
    tall = tall[tall["OMS_SKU"].astype(str).str.len() > 0]
    tall["Qty"] = pd.to_numeric(tall["Qty"], errors="coerce")
    # Excel/pandas sometimes round-trips integer NA as ``iinfo(int64).min`` — those
    # show up as huge negatives. Coerce those back to NaN so they're treated as
    # "no snapshot for that day", not "zero stock".
    tall.loc[tall["Qty"] < -1.0, "Qty"] = pd.NA
    # Blank cells (= no snapshot was taken on that day, e.g. Sundays in many
    # warehouses) must NOT count as out-of-stock. Drop them so coverage_days
    # naturally reflects only the days actually sampled; the engine then
    # extrapolates Eff_Days from in-stock-rate × ADS_WINDOW.
    tall = tall.dropna(subset=["Qty"])
    if tall.empty:
        return pd.DataFrame(columns=_TALL_COLS)
    tall["Qty"] = tall["Qty"].astype(float).clip(lower=0.0)
    tall["Source"] = "uploaded"
    return tall[_STORE_COLS].reset_index(drop=True)


def parse_daily_inventory_history_dataframes(
    sheet_dfs: dict[str, pd.DataFrame],
    sku_mapping: Optional[dict] = None,
    *,
    max_days: int | None = None,
) -> pd.DataFrame:
    mapping = sku_mapping if sku_mapping is not None else {}
    parts: list[pd.DataFrame] = []
    for name, df in sheet_dfs.items():
        if df is None or df.empty:
            continue
        parts.append(_parse_one_sheet(df, mapping, sheet_name=name, max_days=max_days))
    if not parts:
        return pd.DataFrame(columns=_TALL_COLS)
    out = pd.concat(parts, ignore_index=True)
    if out.empty:
        return out
    # Multiple sheets can list the same SKU — keep the maximum stock seen across
    # sources for the same date (warehouse + FBA snapshots can both be > 0).
    out = (
        out.groupby(["OMS_SKU", "Date"], as_index=False)["Qty"].max()
        .sort_values(["OMS_SKU", "Date"])
        .reset_index(drop=True)
    )
    return out


def _sheet_looks_like_inventory(name: str) -> bool:
    n = (name or "").strip().lower()
    if not n:
        return False
    hints = ("oms", "amazon", "inventory", "fba", "warehouse", "stock", "on hand", "onhand")
    return any(h in n for h in hints)


def parse_daily_inventory_history_upload(
    file: BinaryIO,
    filename: str,
    sku_mapping: Optional[dict] = None,
    *,
    max_days: int | None = None,
    on_progress: Callable[[str], None] | None = None,
) -> pd.DataFrame:
    def _progress(msg: str) -> None:
        if on_progress is not None:
            on_progress(msg)

    name = (filename or "").lower()
    raw = file.read() if hasattr(file, "read") else file
    if not raw:
        return pd.DataFrame(columns=_TALL_COLS)
    if name.endswith(".csv"):
        _progress("Parsing CSV…")
        df = pd.read_csv(io.BytesIO(raw), dtype=str, low_memory=False)
        return parse_daily_inventory_history_dataframes(
            {"csv": df}, sku_mapping=sku_mapping, max_days=max_days,
        )
    _progress("Opening workbook…")
    name_l = (filename or "").lower()
    is_xlsx = name_l.endswith(".xlsx") or (not name_l.endswith(".xls") and not name_l.endswith(".csv"))
    if is_xlsx:
        xl = pd.ExcelFile(io.BytesIO(raw))
        sheet_names = list(xl.sheet_names)
        inv_names = [sn for sn in sheet_names if _sheet_looks_like_inventory(sn)]
        to_read = inv_names if inv_names else sheet_names
        sheets: dict[str, pd.DataFrame] = {}
        for i, sn in enumerate(to_read):
            _progress(f"Reading sheet {i + 1}/{len(to_read)}: {sn}…")
            try:
                sheets[sn] = _read_inventory_history_sheet(raw, sn, max_days=max_days)
            except Exception:
                try:
                    sheets[sn] = xl.parse(sn, dtype=str)
                except Exception:
                    continue
    else:
        xl = pd.ExcelFile(io.BytesIO(raw))
        sheet_names = list(xl.sheet_names)
        inv_names = [sn for sn in sheet_names if _sheet_looks_like_inventory(sn)]
        to_read = inv_names if inv_names else sheet_names
        sheets = {}
        for i, sn in enumerate(to_read):
            _progress(f"Reading sheet {i + 1}/{len(to_read)}: {sn}…")
            try:
                sheets[sn] = xl.parse(sn, dtype=str)
            except Exception:
                continue
    _progress("Melting date columns…")
    return parse_daily_inventory_history_dataframes(
        sheets, sku_mapping=sku_mapping, max_days=max_days,
    )


#: Operational threshold for "in stock" when counting Eff_Days. A day
#: counts toward Eff_Days when on-hand >= IN_STOCK_MIN_QTY (defaults to 1.0,
#: meaning any positive stock counts).
IN_STOCK_MIN_QTY: float = 1.0


def trim_inventory_history_for_po(
    inv_history: pd.DataFrame,
    window_start: pd.Timestamp,
    window_end: pd.Timestamp,
    po_skus: Optional[set] = None,
) -> pd.DataFrame:
    """Keep only SKU-day rows needed for one PO run (avoids processing multi-year baselines)."""
    if inv_history is None or inv_history.empty:
        return pd.DataFrame(columns=_TALL_COLS)
    ws = pd.Timestamp(window_start).normalize()
    we = pd.Timestamp(window_end).normalize()
    if ws > we:
        return pd.DataFrame(columns=_TALL_COLS)

    dates = pd.to_datetime(inv_history["Date"], errors="coerce").dt.normalize()
    mask = (dates >= ws) & (dates <= we)
    if po_skus:
        mask = mask & inv_history["OMS_SKU"].astype(str).isin(po_skus)
    if not bool(mask.any()):
        return pd.DataFrame(columns=_TALL_COLS)

    out = inv_history.loc[mask, ["OMS_SKU", "Date", "Qty"]].copy()
    out["Date"] = pd.to_datetime(out["Date"], errors="coerce").dt.normalize()
    out["Qty"] = pd.to_numeric(out["Qty"], errors="coerce")
    out = out.dropna(subset=["Date", "Qty", "OMS_SKU"])
    out = out[out["OMS_SKU"].astype(str).str.len() > 0]
    if out.empty:
        return pd.DataFrame(columns=_TALL_COLS)
    return out.reset_index(drop=True)


def effective_days_from_history(
    inv_history: pd.DataFrame,
    cutoff_start: pd.Timestamp,
    cutoff_end: pd.Timestamp,
    min_qty: float = IN_STOCK_MIN_QTY,
) -> pd.DataFrame:
    """
    Count days within ``[cutoff_start, cutoff_end]`` where each SKU had ``Qty >= min_qty``.

    By default ``min_qty == IN_STOCK_MIN_QTY`` (1.0) — any positive on-hand
    counts toward Eff_Days.

    Returns: ``OMS_SKU``, ``Eff_Days_Inventory`` (int).
    """
    if inv_history is None or inv_history.empty:
        return pd.DataFrame(columns=["OMS_SKU", "Eff_Days_Inventory"])
    df = inv_history.copy()
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df.dropna(subset=["Date"])
    if df.empty:
        return pd.DataFrame(columns=["OMS_SKU", "Eff_Days_Inventory"])
    cs = pd.Timestamp(cutoff_start).normalize()
    ce = pd.Timestamp(cutoff_end).normalize()
    if cs > ce:
        return pd.DataFrame(columns=["OMS_SKU", "Eff_Days_Inventory"])
    mask = (df["Date"] >= cs) & (df["Date"] <= ce)
    sub = df.loc[mask].copy()
    if sub.empty:
        return pd.DataFrame(columns=["OMS_SKU", "Eff_Days_Inventory"])
    sub["Qty"] = pd.to_numeric(sub["Qty"], errors="coerce")
    # Drop "no snapshot" rows so we never count blank cells as out-of-stock days.
    sub = sub.dropna(subset=["Qty"])
    if sub.empty:
        return pd.DataFrame(columns=["OMS_SKU", "Eff_Days_Inventory"])
    sub["_has_stock"] = (sub["Qty"] >= float(min_qty)).astype(int)
    out = (
        sub.groupby("OMS_SKU", as_index=False)["_has_stock"]
        .sum()
        .rename(columns={"_has_stock": "Eff_Days_Inventory"})
    )
    out["Eff_Days_Inventory"] = out["Eff_Days_Inventory"].astype(int)
    return out


def latest_inventory_qty_by_sku(
    history_df: pd.DataFrame,
    *,
    as_of: pd.Timestamp | None = None,
) -> pd.DataFrame:
    """Latest on-hand qty per SKU from the wide inventory history matrix."""
    if history_df is None or history_df.empty:
        return pd.DataFrame(columns=["OMS_SKU", "History_Qty"])
    df = history_df.copy()
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.normalize()
    df["Qty"] = pd.to_numeric(df["Qty"], errors="coerce")
    df = df.dropna(subset=["Date", "OMS_SKU", "Qty"])
    if df.empty:
        return pd.DataFrame(columns=["OMS_SKU", "History_Qty"])
    if as_of is not None:
        cap = pd.Timestamp(as_of).normalize()
        df = df[df["Date"] <= cap]
        if df.empty:
            return pd.DataFrame(columns=["OMS_SKU", "History_Qty"])
    latest_date = df["Date"].max()
    latest = df[df["Date"] == latest_date]
    out = (
        latest.groupby("OMS_SKU", as_index=False)["Qty"]
        .max()
        .rename(columns={"Qty": "History_Qty"})
    )
    out["OMS_SKU"] = out["OMS_SKU"].astype(str).str.strip().str.upper()
    return out


def overlay_inventory_variant_from_history(
    inv_df: pd.DataFrame,
    history_df: pd.DataFrame | None,
    *,
    snapshot_date: str | None = None,
    reference_date: str | None = None,
) -> tuple[pd.DataFrame, dict]:
    """
    When the wide history matrix is fresher than the daily snapshot, use its
    latest column for PO on-hand (Total_Inventory / OMS_Inventory).
    """
    from .inventory_staleness import data_is_stale, today_ist_iso

    meta: dict = {"applied": False, "skus_updated": 0, "reason": ""}
    if inv_df is None or getattr(inv_df, "empty", True):
        meta["reason"] = "empty_inventory"
        return inv_df, meta
    if history_df is None or getattr(history_df, "empty", True):
        meta["reason"] = "empty_history"
        return inv_df, meta

    ref = (reference_date or today_ist_iso()).strip()[:10]
    hist_max = inventory_history_max_date(history_df)
    if hist_max is None:
        meta["reason"] = "no_history_dates"
        return inv_df, meta
    hist_max_s = str(hist_max.date())
    if data_is_stale(ref, hist_max_s, max_expected_lag_days=1):
        meta["reason"] = "history_stale"
        return inv_df, meta

    snap_s = str(snapshot_date or "").strip()[:10]
    snap_stale = not snap_s or data_is_stale(ref, snap_s, max_expected_lag_days=1)
    if not snap_stale and snap_s >= hist_max_s:
        meta["reason"] = "snapshot_fresh"
        return inv_df, meta

    latest = latest_inventory_qty_by_sku(history_df, as_of=hist_max)
    if latest.empty:
        meta["reason"] = "no_latest_rows"
        return inv_df, meta

    out = inv_df.copy()
    if "OMS_SKU" not in out.columns:
        meta["reason"] = "no_sku_column"
        return inv_df, meta
    out["OMS_SKU"] = out["OMS_SKU"].astype(str).str.strip().str.upper()
    merged = out.merge(latest, on="OMS_SKU", how="left")
    has_hist = merged["History_Qty"].notna()
    if not bool(has_hist.any()):
        meta["reason"] = "no_overlap"
        return inv_df, meta

    hist_qty = pd.to_numeric(merged["History_Qty"], errors="coerce").fillna(0.0)
    # Matrix cells are total on-hand per day — update Total_Inventory only so we do not
    # mirror the same figure into OMS_Inventory (that made PO stock look doubled).
    if "Total_Inventory" in merged.columns:
        cur = pd.to_numeric(merged["Total_Inventory"], errors="coerce").fillna(0.0)
        merged["Total_Inventory"] = np.where(has_hist, hist_qty, cur)
    else:
        merged["Total_Inventory"] = np.where(has_hist, hist_qty, 0.0)
    merged.drop(columns=["History_Qty"], inplace=True)

    meta["applied"] = True
    meta["skus_updated"] = int(has_hist.sum())
    meta["history_as_of"] = hist_max_s
    meta["reason"] = "history_newer_than_snapshot"
    return merged, meta


def extend_history_with_sales(
    inv_history: pd.DataFrame,
    sales_df: Optional[pd.DataFrame],
    cap_date: Optional[pd.Timestamp] = None,
) -> pd.DataFrame:
    """Roll the uploaded baseline inventory forward using daily sales activity.

    User intent: upload the Daily Inventory History sheet *once*. From there,
    each day's sales upload should let the engine work out which SKUs were
    in stock without another inventory upload.

    For every SKU present in ``inv_history``, derive a synthetic snapshot
    for each day from ``(sheet_max_date + 1)`` up to ``cap_date`` (default:
    today, normalised). New on-hand is::

        new_qty = max(0, prev_qty - Σ Units_Effective for that day)

    Shipments are positive and pull stock down; refunds/cancellations are
    negative and push it back up. Stock is floored at 0 (we never report
    negative on-hand).

    SKUs without a baseline snapshot are skipped — we have no starting
    on-hand to roll forward.

    Returns the union of baseline + derived rows, with an extra ``Source``
    column (``"uploaded"`` vs ``"derived"``). Callers that only need the
    three core columns can drop ``Source`` after the merge.
    """
    if inv_history is None or inv_history.empty:
        return pd.DataFrame(columns=["OMS_SKU", "Date", "Qty", "Source"])

    base = inv_history.copy()
    base["Date"] = pd.to_datetime(base["Date"], errors="coerce").dt.normalize()
    base["Qty"] = pd.to_numeric(base["Qty"], errors="coerce")
    base = base.dropna(subset=["Date", "Qty", "OMS_SKU"])
    base = base[base["OMS_SKU"].astype(str).str.len() > 0]
    if base.empty:
        return pd.DataFrame(columns=["OMS_SKU", "Date", "Qty", "Source"])
    base["Qty"] = base["Qty"].astype(float).clip(lower=0.0)
    if "Source" not in base.columns:
        base["Source"] = "uploaded"

    sheet_max = base["Date"].max()
    if cap_date is None:
        cap_date = pd.Timestamp.now().normalize()
    cap_date = pd.Timestamp(cap_date).normalize()

    out_cols = ["OMS_SKU", "Date", "Qty", "Source"]
    if sheet_max >= cap_date:
        return base[out_cols].reset_index(drop=True)

    # Last seen snapshot per SKU = starting point for the roll-forward.
    last_snap = (
        base.sort_values(["OMS_SKU", "Date"])
        .groupby("OMS_SKU", as_index=False)
        .tail(1)
        .set_index("OMS_SKU")["Qty"]
    )
    sku_list = last_snap.index.to_numpy()

    # Aggregate net sales (signed Units_Effective) per (SKU, day) within window.
    sales_net = None
    if sales_df is not None and not sales_df.empty:
        s = sales_df.copy()
        sku_col = "Sku" if "Sku" in s.columns else "OMS_SKU"
        date_col = "TxnDate" if "TxnDate" in s.columns else "Date"
        eff_col = "Units_Effective" if "Units_Effective" in s.columns else "Quantity"
        if sku_col in s.columns and date_col in s.columns and eff_col in s.columns:
            s = s[[sku_col, date_col, eff_col]].copy()
            s.columns = ["OMS_SKU", "Date", "Net_Units"]
            s["Date"] = pd.to_datetime(s["Date"], errors="coerce").dt.normalize()
            s["Net_Units"] = pd.to_numeric(s["Net_Units"], errors="coerce").fillna(0.0)
            s = s.dropna(subset=["Date"])
            s = s[(s["Date"] > sheet_max) & (s["Date"] <= cap_date)]
            if not s.empty:
                sales_net = (
                    s.groupby(["OMS_SKU", "Date"], as_index=False)["Net_Units"].sum()
                )
                # Never roll inventory past the last day we actually have sales for.
                smax = pd.to_datetime(s["Date"], errors="coerce").max()
                if pd.notna(smax):
                    cap_date = min(cap_date, pd.Timestamp(smax).normalize())

    if sheet_max >= cap_date:
        return base[out_cols].reset_index(drop=True)

    # Only synthesize snapshots on days with sales activity — do not fill every
    # calendar day with flat on-hand (that inflated Eff_Days when today's sales
    # were not uploaded yet).
    if sales_net is not None and not sales_net.empty:
        active = sorted(
            pd.to_datetime(sales_net["Date"], errors="coerce")
            .dropna()
            .dt.normalize()
            .unique()
        )
        days = pd.DatetimeIndex(
            [d for d in active if sheet_max < pd.Timestamp(d) <= cap_date]
        )
    else:
        days = pd.DatetimeIndex([])

    if len(days) == 0:
        return base[out_cols].reset_index(drop=True)

    authoritative = base[base["Source"].astype(str).str.lower().isin(["snapshot", "uploaded"])]
    blocked = set(
        zip(
            authoritative["OMS_SKU"].astype(str).str.strip().str.upper(),
            authoritative["Date"].dt.normalize(),
        )
    )

    if sales_net is not None and not sales_net.empty:
        pivot = (
            sales_net.set_index(["OMS_SKU", "Date"])["Net_Units"]
            .unstack(fill_value=0.0)
            .reindex(index=sku_list, columns=days, fill_value=0.0)
        )
        net_matrix = pivot.to_numpy(dtype=float)  # shape: (n_sku, n_days)
    else:
        net_matrix = np.zeros((len(sku_list), len(days)), dtype=float)

    # Iterate days (small N — usually a handful) but vectorise across SKUs.
    prev_qty = last_snap.reindex(sku_list).fillna(0.0).to_numpy(dtype=float)
    derived_rows: list[pd.DataFrame] = []
    for di, d in enumerate(days):
        net_d = net_matrix[:, di]
        new_qty = np.maximum(0.0, prev_qty - net_d)
        # Only materialize rows for SKUs that had on-hand stock or net sales this day.
        # Writing explicit Qty=0 for every OOS SKU on every sales day poisoned the
        # wide matrix (all dashes) and forced Eff_Days_Inventory to 0 after merges.
        active = (prev_qty > 0) | (np.abs(net_d) > 1e-9)
        if np.any(active):
            day_ts = pd.Timestamp(d).normalize()
            sku_active = sku_list[active]
            keep = [
                i
                for i, sku in enumerate(sku_active)
                if (str(sku).strip().upper(), day_ts) not in blocked
            ]
            if keep:
                derived_rows.append(
                    pd.DataFrame(
                        {
                            "OMS_SKU": sku_active[keep],
                            "Date": day_ts,
                            "Qty": new_qty[active][keep],
                            "Source": "derived",
                        }
                    )
                )
        prev_qty = new_qty

    if not derived_rows:
        return base[out_cols].reset_index(drop=True)

    derived = pd.concat(derived_rows, ignore_index=True)
    full = pd.concat([base[out_cols], derived[out_cols]], ignore_index=True)
    return full.reset_index(drop=True)


def _sales_net_by_sku_day(
    sales_df: pd.DataFrame | None,
    *,
    start: pd.Timestamp | None = None,
    end: pd.Timestamp | None = None,
) -> pd.DataFrame:
    """Aggregate signed daily net units per SKU (shipments positive, returns negative)."""
    empty = pd.DataFrame(columns=["OMS_SKU", "Date", "Net_Units"])
    if sales_df is None or sales_df.empty:
        return empty
    s = sales_df.copy()
    sku_col = "Sku" if "Sku" in s.columns else "OMS_SKU"
    date_col = "TxnDate" if "TxnDate" in s.columns else "Date"
    eff_col = "Units_Effective" if "Units_Effective" in s.columns else "Quantity"
    if sku_col not in s.columns or date_col not in s.columns or eff_col not in s.columns:
        return empty
    s = s[[sku_col, date_col, eff_col]].copy()
    s.columns = ["OMS_SKU", "Date", "Net_Units"]
    s["Date"] = pd.to_datetime(s["Date"], errors="coerce").dt.normalize()
    s["Net_Units"] = pd.to_numeric(s["Net_Units"], errors="coerce").fillna(0.0)
    s["OMS_SKU"] = s["OMS_SKU"].astype(str).str.strip().str.upper()
    s = s.dropna(subset=["Date"])
    s = s[s["OMS_SKU"].str.len() > 0]
    if start is not None:
        s = s[s["Date"] >= pd.Timestamp(start).normalize()]
    if end is not None:
        s = s[s["Date"] <= pd.Timestamp(end).normalize()]
    if s.empty:
        return empty
    return s.groupby(["OMS_SKU", "Date"], as_index=False)["Net_Units"].sum()


def rollforward_inventory_day(
    prev_day: pd.DataFrame,
    target_date: pd.Timestamp,
    sales_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Vectorised one-day roll-forward from a prior snapshot column."""
    day = pd.Timestamp(target_date).normalize()
    base = prev_day.copy()
    base["OMS_SKU"] = base["OMS_SKU"].astype(str).str.strip().str.upper()
    base["Qty"] = pd.to_numeric(base["Qty"], errors="coerce").fillna(0.0)
    out = base.groupby("OMS_SKU", as_index=False)["Qty"].max()
    sales = _sales_net_by_sku_day(sales_df, start=day, end=day)
    if not sales.empty:
        out = out.merge(sales, on="OMS_SKU", how="left")
        out["Net_Units"] = out["Net_Units"].fillna(0.0)
        out["Qty"] = (out["Qty"] - out["Net_Units"]).clip(lower=0.0)
        out = out.drop(columns=["Net_Units"])
    out["Date"] = day
    out["Source"] = "derived"
    return out[["OMS_SKU", "Date", "Qty", "Source"]]


def project_inventory_calendar(
    inv_history: pd.DataFrame,
    start: pd.Timestamp,
    end: pd.Timestamp,
    sales_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Fill every calendar day in ``[start, end]`` with on-hand qty per SKU.

    Authoritative uploaded/snapshot cells win. Gap days roll forward from the
    previous day minus net sales (floored at zero).
    """
    if inv_history is None or inv_history.empty:
        return pd.DataFrame(columns=_STORE_COLS)
    cs = pd.Timestamp(start).normalize()
    ce = pd.Timestamp(end).normalize()
    if cs > ce:
        return pd.DataFrame(columns=_STORE_COLS)

    base = _ensure_source_column(inv_history)
    base["Date"] = pd.to_datetime(base["Date"], errors="coerce").dt.normalize()
    base["Qty"] = pd.to_numeric(base["Qty"], errors="coerce")
    base = base.dropna(subset=["Date", "OMS_SKU"])
    base["OMS_SKU"] = base["OMS_SKU"].astype(str).str.strip().str.upper()
    base = base[base["OMS_SKU"].str.len() > 0]
    if base.empty:
        return pd.DataFrame(columns=_STORE_COLS)

    calendar = pd.date_range(cs, ce, freq="D")
    skus = sorted(base.loc[base["Date"] <= ce, "OMS_SKU"].astype(str).unique())
    if not skus:
        return pd.DataFrame(columns=_STORE_COLS)

    auth = base.copy()
    auth["_src_rank"] = auth["Source"].astype(str).str.strip().str.lower().map(_SOURCE_RANK).fillna(1)
    auth = (
        auth.sort_values(["OMS_SKU", "Date", "_src_rank", "Qty"], ascending=[True, True, False, False])
        .drop_duplicates(subset=["OMS_SKU", "Date"], keep="first")
        .drop(columns=["_src_rank"])
    )

    pre = auth[auth["Date"] < cs].sort_values(["OMS_SKU", "Date"]).groupby("OMS_SKU", as_index=False).tail(1)
    prev = (
        pre.set_index("OMS_SKU")["Qty"].astype(float)
        if not pre.empty
        else pd.Series(dtype=float)
    ).reindex(skus).fillna(0.0)

    sales = _sales_net_by_sku_day(sales_df, start=cs - pd.Timedelta(days=365), end=ce)
    sales_by_day: dict[pd.Timestamp, pd.Series] = {}
    if not sales.empty:
        for day_ts, chunk in sales.groupby("Date"):
            sales_by_day[pd.Timestamp(day_ts).normalize()] = (
                chunk.set_index("OMS_SKU")["Net_Units"].astype(float)
            )

    out_frames: list[pd.DataFrame] = []
    for day in calendar:
        day_ts = pd.Timestamp(day).normalize()
        auth_day = auth[auth["Date"] == day_ts]
        if not auth_day.empty:
            out_frames.append(
                auth_day[["OMS_SKU", "Date", "Qty", "Source"]].copy()
            )
            prev = prev.copy()
            prev.loc[auth_day["OMS_SKU"].astype(str)] = (
                auth_day.set_index("OMS_SKU")["Qty"].astype(float)
            )
            continue
        net = sales_by_day.get(day_ts, pd.Series(0.0, index=skus)).reindex(skus).fillna(0.0)
        new_prev = (prev - net).clip(lower=0.0)
        active = (prev > 0.0) | (net.abs() > 1e-9) | (new_prev >= 0.0)
        out_frames.append(
            pd.DataFrame(
                {
                    "OMS_SKU": skus,
                    "Date": day_ts,
                    "Qty": new_prev.values,
                    "Source": "derived",
                }
            )
        )
        prev = new_prev

    if not out_frames:
        return pd.DataFrame(columns=_STORE_COLS)
    return _coalesce_history_rows(pd.concat(out_frames, ignore_index=True))


def repair_inventory_history_spikes(
    inv_history: pd.DataFrame,
    sales_df: pd.DataFrame | None = None,
    *,
    tolerance_units: float = 800.0,
    ratio_threshold: float = 1.04,
) -> tuple[pd.DataFrame, list[str]]:
    """Replace snapshot columns that jump up when sales imply a decrease."""
    if inv_history is None or inv_history.empty:
        return inv_history, []
    work = _ensure_source_column(inv_history)
    work["Date"] = pd.to_datetime(work["Date"], errors="coerce").dt.normalize()
    work["Qty"] = pd.to_numeric(work["Qty"], errors="coerce").fillna(0.0).clip(lower=0.0)
    dates = sorted(work["Date"].dropna().unique())
    if len(dates) < 2:
        return work.reset_index(drop=True), []

    sales = _sales_net_by_sku_day(sales_df)
    actions: list[str] = []
    out = work.copy()
    for i in range(1, len(dates)):
        d_prev = pd.Timestamp(dates[i - 1]).normalize()
        d_next = pd.Timestamp(dates[i]).normalize()
        prev_rows = out[out["Date"] == d_prev]
        next_rows = out[out["Date"] == d_next]
        if prev_rows.empty or next_rows.empty:
            continue
        t_prev = float(prev_rows["Qty"].sum())
        t_next = float(next_rows["Qty"].sum())
        if t_next <= t_prev:
            continue
        prev_skus = int(prev_rows["OMS_SKU"].astype(str).nunique())
        next_skus = int(next_rows["OMS_SKU"].astype(str).nunique())
        sku_ratio = next_skus / max(prev_skus, 1)
        gap_mask = (sales["Date"] > d_prev) & (sales["Date"] <= d_next) if not sales.empty else pd.Series(dtype=bool)
        gap_sales = float(sales.loc[gap_mask, "Net_Units"].sum()) if not sales.empty and bool(gap_mask.any()) else 0.0
        expected = max(0.0, t_prev - gap_sales)
        tol = max(50.0, float(gap_sales) * 0.25, t_prev * 0.005)
        similar_universe = 0.92 <= sku_ratio <= 1.12
        big_total_jump = t_next > t_prev * 1.12 and (t_next - t_prev) > max(5000.0, t_prev * 0.08)
        new_sku_surge = (next_skus - prev_skus) > 400 and t_next > t_prev * 1.08
        spike = (
            t_next > expected + tol
            and (big_total_jump and similar_universe or new_sku_surge)
        )
        if not spike:
            continue
        prev_rows = out[out["Date"] == d_prev]
        repaired = rollforward_inventory_day(prev_rows, d_next, sales_df=sales_df)
        out = out[out["Date"] != d_next]
        out = pd.concat([out, repaired], ignore_index=True)
        actions.append(f"repaired_spike:{d_next.date()}: {int(t_next)}→{int(repaired['Qty'].sum())}")
    if not actions:
        return work.reset_index(drop=True), []
    return _coalesce_history_rows(out), actions


def densify_inventory_history_for_view(
    inv_history: pd.DataFrame,
    *,
    days: int | None = None,
    end_date: str | None = None,
    sales_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Calendar window + per-SKU roll-forward for the inventory history UI."""
    if inv_history is None or inv_history.empty:
        return pd.DataFrame(columns=_STORE_COLS)
    span = int(days if days is not None else _DEFAULT_VIEW_DAYS)
    end = pd.Timestamp(inventory_history_view_end_date(inv_history, end_date)).normalize()
    start = end - pd.Timedelta(days=max(0, span - 1))
    trimmed = filter_inventory_history_window(inv_history, days=span, end_date=str(end.date()))
    return project_inventory_calendar(trimmed, start, end, sales_df=sales_df)


def coverage_days_within(
    inv_history: pd.DataFrame,
    cutoff_start: pd.Timestamp,
    cutoff_end: pd.Timestamp,
) -> int:
    """Number of distinct snapshot dates the history sheet provides within ``[cs, ce]``.

    Used by the PO engine to decide whether the sheet covers the full ADS window
    or only a sub-range — partial-coverage SKUs must be *extrapolated*, not
    capped at the sheet's day count (otherwise a sheet with 24 days collapses
    every SKU's Eff_Days to 24 even when the SKU was in-stock all 30 days).
    """
    if inv_history is None or inv_history.empty:
        return 0
    df = inv_history.copy()
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df.dropna(subset=["Date"])
    if df.empty:
        return 0
    cs = pd.Timestamp(cutoff_start).normalize()
    ce = pd.Timestamp(cutoff_end).normalize()
    if cs > ce:
        return 0
    mask = (df["Date"] >= cs) & (df["Date"] <= ce)
    sub = df.loc[mask]
    if sub.empty:
        return 0
    return int(sub["Date"].dt.normalize().nunique())


def inventory_history_summary(
    df: pd.DataFrame,
    *,
    days: int | None = None,
    end_date: str | None = None,
) -> dict:
    view = filter_inventory_history_view(df, days=days, end_date=end_date)
    if view is None or view.empty:
        return {
            "loaded": False,
            "rows": 0,
            "skus": 0,
            "days": 0,
            "min_date": "",
            "max_date": "",
            "window_days": int(days if days is not None else _DEFAULT_VIEW_DAYS),
            "window_end": str(end_date or today_ist_timestamp().date()),
        }
    dates = pd.to_datetime(view["Date"], errors="coerce").dt.normalize()
    min_d = dates.min()
    max_d = dates.max()
    return {
        "loaded": True,
        "rows": int(len(view)),
        "skus": int(view["OMS_SKU"].astype(str).nunique()),
        "days": int(dates.nunique()),
        "min_date": str(pd.Timestamp(min_d).date()) if pd.notna(min_d) else "",
        "max_date": str(pd.Timestamp(max_d).date()) if pd.notna(max_d) else "",
        "window_days": int(days if days is not None else _DEFAULT_VIEW_DAYS),
        "window_end": str(pd.Timestamp(max_d).date()) if pd.notna(max_d) else str(end_date or today_ist_timestamp().date()),
    }


def list_inventory_history_dates(df: pd.DataFrame, limit: int = 120) -> list[dict]:
    if df is None or df.empty:
        return []
    work = df.copy()
    work["Date"] = pd.to_datetime(work["Date"], errors="coerce").dt.normalize()
    work = work.dropna(subset=["Date"])
    if work.empty:
        return []
    grouped = (
        work.groupby("Date", as_index=False)
        .agg(rows=("OMS_SKU", "count"), skus=("OMS_SKU", "nunique"))
        .sort_values("Date", ascending=False)
    )
    if limit > 0:
        grouped = grouped.head(int(limit))
    return [
        {
            "date": str(pd.Timestamp(r["Date"]).date()),
            "rows": int(r["rows"]),
            "skus": int(r["skus"]),
        }
        for _, r in grouped.iterrows()
    ]


def inventory_rows_for_date(
    df: pd.DataFrame,
    date_iso: str,
    *,
    q: str = "",
    limit: int = 500,
    offset: int = 0,
) -> dict:
    if df is None or df.empty:
        return {
            "loaded": False,
            "date": date_iso,
            "rows": [],
            "total": 0,
            "limit": int(limit),
            "offset": int(offset),
        }
    try:
        target = pd.Timestamp(date_iso).normalize()
    except Exception:
        return {
            "loaded": True,
            "date": date_iso,
            "rows": [],
            "total": 0,
            "limit": int(limit),
            "offset": int(offset),
            "message": "Invalid date.",
        }

    work = df.copy()
    work["Date"] = pd.to_datetime(work["Date"], errors="coerce").dt.normalize()
    sub = work[work["Date"] == target].copy()
    if sub.empty:
        return {
            "loaded": True,
            "date": str(target.date()),
            "rows": [],
            "total": 0,
            "limit": int(limit),
            "offset": int(offset),
        }

    sub["OMS_SKU"] = sub["OMS_SKU"].astype(str).str.strip()
    sub["Qty"] = pd.to_numeric(sub["Qty"], errors="coerce").fillna(0.0)
    needle = (q or "").strip().upper()
    if needle:
        sub = sub[sub["OMS_SKU"].str.upper().str.contains(needle, na=False)]
    sub = sub.sort_values(["Qty", "OMS_SKU"], ascending=[False, True])
    total = int(len(sub))
    page = sub.iloc[max(0, int(offset)) : max(0, int(offset)) + max(1, int(limit))]
    rows = [
        {
            "sku": str(r["OMS_SKU"]),
            "qty": float(r["Qty"]),
            "in_stock": bool(float(r["Qty"]) >= IN_STOCK_MIN_QTY),
            "source": str(r.get("Source", "uploaded") or "uploaded"),
        }
        for _, r in page.iterrows()
    ]
    return {
        "loaded": True,
        "date": str(target.date()),
        "rows": rows,
        "total": total,
        "limit": int(limit),
        "offset": int(offset),
        "in_stock_min_qty": float(IN_STOCK_MIN_QTY),
    }


def inventory_history_wide_matrix(
    df: pd.DataFrame,
    *,
    q: str = "",
    limit: int = 150,
    offset: int = 0,
    days: int | None = None,
    end_date: str | None = None,
    sales_df: pd.DataFrame | None = None,
) -> dict:
    """Pivot tall history to Excel-style wide matrix: SKU rows × date columns."""
    empty = {
        "loaded": False,
        "dates": [],
        "date_totals": [],
        "rows": [],
        "total": 0,
        "limit": int(limit),
        "offset": int(offset),
        "in_stock_min_qty": float(IN_STOCK_MIN_QTY),
        "window_days": int(days if days is not None else _DEFAULT_VIEW_DAYS),
        "window_end": str(end_date or today_ist_timestamp().date()),
    }
    if df is None or df.empty:
        return empty

    work = densify_inventory_history_for_view(
        df, days=days, end_date=end_date, sales_df=sales_df
    )
    work["Date"] = pd.to_datetime(work["Date"], errors="coerce").dt.normalize()
    work = work.dropna(subset=["Date", "OMS_SKU"])
    if work.empty:
        return {**empty, "loaded": True}

    work["OMS_SKU"] = work["OMS_SKU"].astype(str).str.strip()
    work["Qty"] = pd.to_numeric(work["Qty"], errors="coerce").fillna(0.0)
    needle = (q or "").strip().upper()
    if needle:
        work = work[work["OMS_SKU"].str.upper().str.contains(needle, na=False)]
    if work.empty:
        return {**empty, "loaded": True}

    work = (
        work.sort_values("Qty", ascending=False)
        .drop_duplicates(subset=["OMS_SKU", "Date"], keep="first")
    )
    span = int(days if days is not None else _DEFAULT_VIEW_DAYS)
    view_end = pd.Timestamp(inventory_history_view_end_date(df, end_date)).normalize()
    view_start = view_end - pd.Timedelta(days=max(0, span - 1))
    dates_sorted = list(pd.date_range(view_start, view_end, freq="D"))
    date_strs = [str(pd.Timestamp(d).date()) for d in dates_sorted]
    date_totals = [
        float(work.loc[work["Date"] == d, "Qty"].sum()) for d in dates_sorted
    ]

    sku_list = sorted(work["OMS_SKU"].astype(str).unique())
    total = int(len(sku_list))
    start = max(0, int(offset))
    end = start + max(1, int(limit))
    page_skus = sku_list[start:end]
    if not len(page_skus):
        return {**empty, "loaded": True, "dates": date_strs, "date_totals": date_totals, "total": total}

    page_work = work[work["OMS_SKU"].isin(page_skus)]
    pivot = page_work.pivot(index="OMS_SKU", columns="Date", values="Qty")
    pivot = pivot.reindex(index=page_skus, columns=dates_sorted).fillna(0.0)

    rows = [
        {
            "sku": str(sku),
            "qtys": [float(row.get(d, 0.0) or 0.0) for d in dates_sorted],
        }
        for sku, row in pivot.iterrows()
    ]
    return {
        "loaded": True,
        "dates": date_strs,
        "date_totals": date_totals,
        "rows": rows,
        "total": total,
        "limit": int(limit),
        "offset": start,
        "in_stock_min_qty": float(IN_STOCK_MIN_QTY),
        "window_days": int(days if days is not None else _DEFAULT_VIEW_DAYS),
        "window_end": str(date_strs[-1]) if date_strs else str(end_date or today_ist_timestamp().date()),
    }


_DAILY_INV_META_FILENAME = "daily_inventory_history_meta.json"


def _warm_cache_dir() -> "Path":
    from pathlib import Path

    import os

    return Path(os.environ.get("WARM_CACHE_DIR", "/data/warm_cache"))


def daily_inventory_history_meta_bundle(sess) -> dict:
    """Serializable metadata for warm-cache / disk restore (shared across users)."""
    df = getattr(sess, "daily_inventory_history_df", None)
    rows, skus, df_ts = session_inventory_matrix_stats(sess)
    mx = inventory_history_max_date(df)
    min_d = None
    if df is not None and not getattr(df, "empty", True) and "Date" in df.columns:
        min_d = pd.to_datetime(df["Date"], errors="coerce").min()
    df_max_s = str(pd.Timestamp(mx).date()) if mx is not None and pd.notna(mx) else ""
    sess_matrix_max = str(getattr(sess, "daily_inventory_history_matrix_max_date", "") or "").strip()[:10]
    cap_ts = inventory_history_matrix_cap_date(sess)
    cap_s = str(cap_ts.date()) if cap_ts is not None else ""
    # Ceiling tracks upload/snapshot authority; max_date must reflect rows actually in the df.
    matrix_ceiling_s = cap_s or sess_matrix_max or df_max_s
    if sess_matrix_max and matrix_ceiling_s < sess_matrix_max:
        matrix_ceiling_s = sess_matrix_max
    if cap_s and matrix_ceiling_s < cap_s:
        matrix_ceiling_s = cap_s
    if df_max_s and matrix_ceiling_s < df_max_s:
        matrix_ceiling_s = df_max_s
    return {
        "daily_inventory_history_uploaded_at": str(
            getattr(sess, "daily_inventory_history_uploaded_at", "") or ""
        ),
        "daily_inventory_history_filename": str(
            getattr(sess, "daily_inventory_history_filename", "") or ""
        ),
        "daily_inventory_history_wide_end_date": str(
            getattr(sess, "daily_inventory_history_wide_end_date", "") or ""
        )[:10]
        or inventory_sheet_end_date_from_filename(
            str(getattr(sess, "daily_inventory_history_filename", "") or "")
        ),
        "daily_inventory_history_snapshot_dates": list(
            getattr(sess, "daily_inventory_history_snapshot_dates", None) or []
        ),
        "daily_inventory_history_rows": rows,
        "daily_inventory_history_skus": skus,
        "daily_inventory_history_max_date": df_max_s,
        "daily_inventory_history_matrix_max_date": matrix_ceiling_s or df_max_s,
        "daily_inventory_history_min_date": str(pd.Timestamp(min_d).date()) if pd.notna(min_d) else "",
    }


def apply_daily_inventory_history_meta(sess, meta: dict) -> None:
    """Copy upload metadata onto a session (not the dataframe)."""
    if not isinstance(meta, dict):
        return
    for key in (
        "daily_inventory_history_uploaded_at",
        "daily_inventory_history_filename",
        "daily_inventory_history_matrix_max_date",
        "daily_inventory_history_wide_end_date",
    ):
        if meta.get(key):
            setattr(sess, key, str(meta[key]))
    snap_dates = meta.get("daily_inventory_history_snapshot_dates")
    if isinstance(snap_dates, list) and snap_dates:
        sess.daily_inventory_history_snapshot_dates = [str(d)[:10] for d in snap_dates if d]
    cap = str(meta.get("daily_inventory_history_matrix_max_date") or meta.get("daily_inventory_history_max_date") or "")
    if len(cap) >= 10:
        promote_daily_inventory_matrix_max_date(sess, cap[:10])


def read_daily_inventory_history_disk_meta() -> dict | None:
    try:
        path = _warm_cache_dir() / _DAILY_INV_META_FILENAME
        if not path.is_file():
            return None
        import json

        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def _load_daily_inventory_df_from_disk() -> pd.DataFrame | None:
    try:
        path = _warm_cache_dir() / "daily_inventory_history_df.parquet"
        if not path.is_file():
            return None
        df = pd.read_parquet(path)
        return df if df is not None and not getattr(df, "empty", True) else None
    except Exception:
        return None


def ensure_latest_daily_inventory_authoritative(sess) -> bool:
    """Prefer the newest uploaded inventory matrix on disk/warm — never merge stale copies."""
    try:
        reconcile_inventory_history_disk_integrity(repair=True)
    except Exception:
        pass
    changed = False
    sess_at = upload_timestamp_epoch(str(getattr(sess, "daily_inventory_history_uploaded_at", "") or ""))
    disk_meta = read_daily_inventory_history_disk_meta() or {}
    disk_at = upload_timestamp_epoch(str(disk_meta.get("daily_inventory_history_uploaded_at") or ""))
    disk_df = _load_daily_inventory_df_from_disk()

    try:
        import backend.main as _main

        wc = (_main._warm_cache or {}).get("daily_inventory_history_df")
        wc_meta = (_main._warm_cache or {}).get(_main._DAILY_INV_META_WARM_KEY)
        wc_at = upload_timestamp_epoch(
            str((wc_meta or {}).get("daily_inventory_history_uploaded_at") or "")
        )
    except Exception:
        wc = None
        wc_meta = None
        wc_at = 0.0

    cur = getattr(sess, "daily_inventory_history_df", None)
    best_df = cur
    best_at = sess_at
    best_meta: dict = {}

    for candidate_df, candidate_at, candidate_meta in (
        (disk_df, disk_at, disk_meta),
        (wc, wc_at, wc_meta if isinstance(wc_meta, dict) else {}),
    ):
        if candidate_df is None or getattr(candidate_df, "empty", True):
            continue
        if candidate_at > best_at + 0.5 or inventory_history_is_newer_than(
            candidate_df,
            best_df,
            incoming_uploaded_at=str((candidate_meta or {}).get("daily_inventory_history_uploaded_at") or ""),
            existing_uploaded_at=str(getattr(sess, "daily_inventory_history_uploaded_at", "") or ""),
        ):
            best_df = candidate_df
            best_at = candidate_at
            best_meta = candidate_meta or {}

    if best_df is not None and not getattr(best_df, "empty", True) and best_df is not cur:
        sess.daily_inventory_history_df = best_df.copy()
        changed = True
    if best_meta:
        apply_daily_inventory_history_meta(sess, best_meta)
        cap = inventory_history_matrix_cap_date(sess)
        if cap is not None:
            promote_daily_inventory_matrix_max_date(sess, str(cap.date()))
        changed = True
    elif getattr(sess, "daily_inventory_history_df", None) is not None:
        cap = inventory_history_matrix_cap_date(sess)
        if cap is not None:
            prev = str(getattr(sess, "daily_inventory_history_matrix_max_date", "") or "")[:10]
            promote_daily_inventory_matrix_max_date(sess, str(cap.date()))
            if str(cap.date()) != prev:
                changed = True
    return changed


def ensure_daily_inventory_coverage_light(sess) -> bool:
    """Attach daily inventory matrix from warm cache for coverage / staleness checks."""
    ensure_latest_daily_inventory_authoritative(sess)
    df = getattr(sess, "daily_inventory_history_df", None)
    if df is not None and not getattr(df, "empty", True):
        try:
            reconcile_daily_inventory_meta_if_session_newer(sess)
        except Exception:
            pass
        cap = inventory_history_matrix_cap_date(sess)
        if cap is not None:
            promote_daily_inventory_matrix_max_date(sess, str(cap.date()))
        return True
    try:
        import backend.main as _main

        if not _main._warm_cache:
            _main.bootstrap_warm_cache_if_empty()
        wc = (_main._warm_cache or {}).get("daily_inventory_history_df")
        if wc is not None and not getattr(wc, "empty", True):
            sess.daily_inventory_history_df = wc
            meta = (_main._warm_cache or {}).get(_main._DAILY_INV_META_WARM_KEY)
            if isinstance(meta, dict) and meta:
                apply_daily_inventory_history_meta(sess, meta)
            cap = inventory_history_matrix_cap_date(sess)
            if cap is not None:
                promote_daily_inventory_matrix_max_date(sess, str(cap.date()))
            return True
    except Exception:
        pass
    disk = read_daily_inventory_history_disk_meta() or {}
    if disk_inventory_meta_looks_placeholder(disk):
        return False
    apply_daily_inventory_history_meta(sess, disk)
    cap = inventory_history_matrix_cap_date(sess)
    if cap is not None:
        promote_daily_inventory_matrix_max_date(sess, str(cap.date()))
    return int(disk.get("daily_inventory_history_rows") or 0) > 0


def reconcile_daily_inventory_meta_if_session_newer(sess) -> bool:
    """Overwrite stale placeholder disk meta when the session holds the real matrix."""
    sess_rows, sess_skus, _ = session_inventory_matrix_stats(sess)
    if sess_skus < 50 or sess_rows < 500:
        return False
    disk = read_daily_inventory_history_disk_meta() or {}
    if not disk_inventory_meta_looks_placeholder(disk):
        disk_rows = int(disk.get("daily_inventory_history_rows") or 0)
        if disk_rows >= int(sess_rows * 0.9):
            return False
    return persist_daily_inventory_history_meta(sess)


def persist_daily_inventory_history_meta(sess) -> bool:
    meta = daily_inventory_history_meta_bundle(sess)
    if not meta.get("daily_inventory_history_uploaded_at") and not meta.get("daily_inventory_history_rows"):
        return False
    df = getattr(sess, "daily_inventory_history_df", None)
    if df is not None and not getattr(df, "empty", True):
        mx = inventory_history_max_date(df)
        meta_max = str(meta.get("daily_inventory_history_max_date") or "")[:10]
        if mx is not None and meta_max and meta_max != str(mx.date()):
            meta["daily_inventory_history_max_date"] = str(mx.date())
    try:
        import json

        path = _warm_cache_dir()
        path.mkdir(parents=True, exist_ok=True)
        (path / _DAILY_INV_META_FILENAME).write_text(
            json.dumps(meta, default=str),
            encoding="utf-8",
        )
        return True
    except Exception:
        return False


# Disk reconcile / production scripts must never trim the on-disk matrix to 30d.
INVENTORY_HISTORY_DISK_RECONCILE_DAYS = 0


def _pipeline_inventory_history_path():
    from pathlib import Path

    return _warm_cache_dir() / "pipeline" / "inventory_history_snapshot.parquet"


def persist_upload_pipeline_snapshot(df: pd.DataFrame | None) -> bool:
    """Keep a durable copy of the last wide-matrix upload (survives stale-cache overwrites)."""
    if df is None or getattr(df, "empty", True):
        return False
    try:
        from ..services.helpers import _coerce_df_for_parquet

        path = _pipeline_inventory_history_path()
        path.parent.mkdir(parents=True, exist_ok=True)
        _coerce_df_for_parquet(df).to_parquet(path, index=False)
        return True
    except Exception:
        return False


def iter_inventory_history_parquet_candidates() -> list:
    """All on-disk daily inventory history parquets (pipeline + warm cache + github bundles)."""
    from pathlib import Path

    cache = _warm_cache_dir()
    roots: list[Path] = []
    if cache.is_dir():
        roots.append(cache)
    gh_root = cache.parent / "github_cache"
    if gh_root.is_dir():
        for sub in sorted(gh_root.iterdir(), reverse=True):
            if sub.is_dir():
                roots.append(sub)
    out: list[Path] = []
    seen: set[str] = set()
    pipeline_p = _pipeline_inventory_history_path()
    if pipeline_p.is_file():
        out.append(pipeline_p)
        seen.add(str(pipeline_p.resolve()))
    for root in roots:
        p = root / "daily_inventory_history_df.parquet"
        key = str(p.resolve())
        if p.is_file() and key not in seen:
            seen.add(key)
            out.append(p)
    return out


def merge_inventory_history_candidates(paths: list) -> pd.DataFrame | None:
    """Union every readable on-disk candidate into one history frame."""
    merged: pd.DataFrame | None = None
    for p in paths:
        try:
            df = pd.read_parquet(p)
        except Exception:
            continue
        if df is None or getattr(df, "empty", True):
            continue
        merged = merge_inventory_history(merged, df)
    return merged


def _inventory_history_unique_days(df: pd.DataFrame | None) -> int:
    if df is None or getattr(df, "empty", True) or "Date" not in df.columns:
        return 0
    return int(pd.to_datetime(df["Date"], errors="coerce").dt.normalize().nunique())


def restore_inventory_history_from_best_disk_backups(
    current: pd.DataFrame | None = None,
) -> pd.DataFrame | None:
    """Merge all on-disk candidates when the union is broader than ``current`` alone."""
    paths = iter_inventory_history_parquet_candidates()
    if not paths:
        return None
    merged = merge_inventory_history_candidates(paths)
    if merged is None or merged.empty:
        return None
    if current is None or getattr(current, "empty", True):
        return merged
    cur_days = _inventory_history_unique_days(current)
    merged_days = _inventory_history_unique_days(merged)
    if merged_days > cur_days:
        return merged
    if inventory_history_is_newer_than(merged, current):
        return merged
    cur_max = inventory_history_max_date(current)
    merged_max = inventory_history_max_date(merged)
    if (
        merged_max is not None
        and cur_max is not None
        and merged_max >= cur_max
        and merged_days >= cur_days
        and len(merged) > len(current)
    ):
        return merged
    return None


def reconcile_inventory_history_disk_integrity(*, repair: bool = True) -> dict:
    """Detect and repair meta/parquet drift; never let metadata claim dates absent from rows."""
    import json

    path = _warm_cache_dir()
    hist_path = path / "daily_inventory_history_df.parquet"
    meta_path = path / _DAILY_INV_META_FILENAME
    report: dict = {"ok": True, "repaired": False, "actions": []}

    if not hist_path.is_file():
        report["ok"] = False
        report["reason"] = "missing_parquet"
        return report

    try:
        df = pd.read_parquet(hist_path)
    except Exception as exc:
        report["ok"] = False
        report["reason"] = f"parquet_read_failed:{exc}"
        return report

    meta = {}
    if meta_path.is_file():
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
        except Exception:
            meta = {}

    df_max = inventory_history_max_date(df)
    df_max_s = str(pd.Timestamp(df_max).date()) if df_max is not None and pd.notna(df_max) else ""
    meta_max_s = str(meta.get("daily_inventory_history_max_date") or "")[:10]
    rows = int(len(df))
    skus = int(df["OMS_SKU"].astype(str).nunique()) if "OMS_SKU" in df.columns else 0
    meta_rows = int(meta.get("daily_inventory_history_rows") or 0)
    wide_end_s = str(meta.get("daily_inventory_history_wide_end_date") or "")[:10]
    matrix_max_s = str(meta.get("daily_inventory_history_matrix_max_date") or "")[:10]

    if meta_max_s and df_max_s and meta_max_s != df_max_s:
        report["actions"].append(f"meta_max {meta_max_s} != parquet_max {df_max_s}")
    if wide_end_s and df_max_s and wide_end_s > df_max_s:
        report["actions"].append(f"parquet_max {df_max_s} behind wide_end {wide_end_s}")
    if matrix_max_s and df_max_s and matrix_max_s > df_max_s:
        report["actions"].append(f"parquet_max {df_max_s} behind matrix_max {matrix_max_s}")
    if meta_rows and rows and abs(meta_rows - rows) > max(500, int(rows * 0.05)):
        report["actions"].append(f"meta_rows {meta_rows} != parquet_rows {rows}")

    if repair and report["actions"]:
        restored = restore_inventory_history_from_best_disk_backups(df)
        if restored is not None and (
            inventory_history_is_newer_than(restored, df)
            or _inventory_history_unique_days(restored) > _inventory_history_unique_days(df)
        ):
            from ..services.helpers import _coerce_df_for_parquet

            df = restored
            _coerce_df_for_parquet(df).to_parquet(hist_path, index=False)
            report["actions"].append("parquet_restored_from_backups")
            df_max = inventory_history_max_date(df)
            df_max_s = str(pd.Timestamp(df_max).date()) if df_max is not None and pd.notna(df_max) else ""
            rows = int(len(df))
            skus = int(df["OMS_SKU"].astype(str).nunique()) if "OMS_SKU" in df.columns else 0

    if not repair or not report["actions"]:
        report["parquet_max_date"] = df_max_s
        report["meta_max_date"] = meta_max_s
        report["parquet_rows"] = rows
        return report

    repaired = dict(meta) if isinstance(meta, dict) else {}
    repaired["daily_inventory_history_max_date"] = df_max_s
    repaired["daily_inventory_history_rows"] = rows
    repaired["daily_inventory_history_skus"] = skus
    if df_max_s and (
        not str(repaired.get("daily_inventory_history_matrix_max_date") or "")[:10]
        or str(repaired.get("daily_inventory_history_matrix_max_date") or "")[:10] < df_max_s
    ):
        # Do not inflate matrix_max above parquet unless snapshot_dates explicitly newer in meta.
        snap_dates = list(repaired.get("daily_inventory_history_snapshot_dates") or [])
        snap_max = max(snap_dates) if snap_dates else df_max_s
        if snap_max > df_max_s:
            repaired["daily_inventory_history_matrix_max_date"] = snap_max
        else:
            repaired["daily_inventory_history_matrix_max_date"] = df_max_s
    min_d = pd.to_datetime(df["Date"], errors="coerce").min() if "Date" in df.columns else None
    if pd.notna(min_d):
        repaired["daily_inventory_history_min_date"] = str(pd.Timestamp(min_d).date())

    meta_path.write_text(json.dumps(repaired, default=str, indent=2), encoding="utf-8")
    report["repaired"] = True
    report["parquet_max_date"] = df_max_s
    report["meta_max_date"] = df_max_s
    report["parquet_rows"] = rows
    return report


def persist_inventory_history_authoritative(sess, df: pd.DataFrame | None = None) -> bool:
    """Atomically persist inventory history parquet + meta from one dataframe."""
    import json

    from ..services.helpers import _coerce_df_for_parquet

    work = df if df is not None else getattr(sess, "daily_inventory_history_df", None)
    if work is None or getattr(work, "empty", True):
        return False
    sess.daily_inventory_history_df = work
    cache = _warm_cache_dir()
    cache.mkdir(parents=True, exist_ok=True)
    hist_path = cache / "daily_inventory_history_df.parquet"
    if hist_path.is_file():
        try:
            old = pd.read_parquet(hist_path)
            disk_meta = read_daily_inventory_history_disk_meta() or {}
            if not inventory_history_is_newer_than(
                work,
                old,
                incoming_uploaded_at=str(
                    daily_inventory_history_meta_bundle(sess).get("daily_inventory_history_uploaded_at") or ""
                ),
                existing_uploaded_at=str(disk_meta.get("daily_inventory_history_uploaded_at") or ""),
            ):
                return False
        except Exception:
            pass
    _coerce_df_for_parquet(work).to_parquet(hist_path, index=False)
    meta = daily_inventory_history_meta_bundle(sess)
    (cache / _DAILY_INV_META_FILENAME).write_text(json.dumps(meta, default=str, indent=2), encoding="utf-8")
    try:
        import backend.main as _main

        if not _main._warm_cache:
            _main._warm_cache = {}
        _main._warm_cache["daily_inventory_history_df"] = work.copy()
        _main._warm_cache[_main._DAILY_INV_META_WARM_KEY] = meta
    except Exception:
        pass
    return True


def upload_timestamp_epoch(value: str) -> float:
    """Parse uploaded_at strings (IST local or UTC ISO) for comparison."""
    s = str(value or "").strip()
    if not s:
        return 0.0
    try:
        from datetime import datetime

        if s.endswith("Z"):
            return datetime.fromisoformat(s.replace("Z", "+00:00")).timestamp()
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            return dt.replace(tzinfo=_IST).timestamp()
        return dt.timestamp()
    except Exception:
        return 0.0


def daily_inventory_meta_is_newer(meta: dict | None, sess) -> bool:
    """True when shared meta is strictly newer than what the session holds."""
    if not isinstance(meta, dict) or not meta:
        return False
    disk_at = upload_timestamp_epoch(str(meta.get("daily_inventory_history_uploaded_at") or ""))
    sess_at = upload_timestamp_epoch(str(getattr(sess, "daily_inventory_history_uploaded_at", "") or ""))
    if disk_at > sess_at + 0.5:
        return True
    disk_max = str(meta.get("daily_inventory_history_max_date") or "").strip()
    if not disk_max:
        return False
    df = getattr(sess, "daily_inventory_history_df", None)
    sess_max = inventory_history_max_date(df)
    if sess_max is None:
        return True
    try:
        return pd.Timestamp(disk_max).normalize() > pd.Timestamp(sess_max).normalize()
    except Exception:
        return False


__all__ = [
    "parse_daily_inventory_history_dataframes",
    "parse_daily_inventory_history_upload",
    "effective_days_from_history",
    "latest_inventory_qty_by_sku",
    "overlay_inventory_variant_from_history",
    "coverage_days_within",
    "extend_history_with_sales",
    "should_skip_inventory_history_extend",
    "refresh_inventory_history_rollforward",
    "append_snapshot_inventory_to_history",
    "recanonicalize_inventory_history_skus",
    "drop_zero_derived_rows",
    "inventory_history_view_end_date",
    "inventory_history_matrix_cap_date",
    "inventory_sheet_end_date_from_filename",
    "record_inventory_snapshot_date",
    "prune_non_snapshot_post_matrix_days",
    "snapshot_dates_from_history",
    "wide_matrix_upload_end_date",
    "_ensure_source_column",
    "_coalesce_history_rows",
    "ensure_latest_daily_inventory_authoritative",
    "disk_inventory_meta_looks_placeholder",
    "session_inventory_matrix_stats",
    "reconcile_daily_inventory_meta_if_session_newer",
    "ensure_daily_inventory_coverage_light",
    "inventory_history_is_newer_than",
    "filter_inventory_history_window",
    "filter_inventory_history_view",
    "inventory_history_summary",
    "list_inventory_history_dates",
    "inventory_rows_for_date",
    "inventory_history_wide_matrix",
    "daily_inventory_history_meta_bundle",
    "apply_daily_inventory_history_meta",
    "read_daily_inventory_history_disk_meta",
    "persist_daily_inventory_history_meta",
    "persist_inventory_history_authoritative",
    "reconcile_inventory_history_disk_integrity",
    "restore_inventory_history_from_best_disk_backups",
    "persist_upload_pipeline_snapshot",
    "iter_inventory_history_parquet_candidates",
    "INVENTORY_HISTORY_DISK_RECONCILE_DAYS",
    "upload_timestamp_epoch",
    "daily_inventory_meta_is_newer",
]


# Suppress unused np import lint when the module shrinks during refactors.
_unused_np = np  # type: ignore[unused-ignore]
